from __future__ import annotations

import csv
import json
import re
import shutil
import time
import urllib.parse
import urllib.request
import urllib.error
import xml.etree.ElementTree as ET
from collections import OrderedDict
from dataclasses import dataclass, field
from datetime import datetime
from email.utils import parsedate_to_datetime
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


ROOT = Path(__file__).resolve().parents[1]
CATALOGUE = ROOT / "catalogue.xlsx"
PREVIEW_CSV = ROOT / "dry_run_catalogue_preview.csv"
CHANGE_LOG = ROOT / "library_changes.md"
RAW_ARXIV_CACHE = ROOT / "raw_arxiv.json"

QUERY = '("de novo protein design" OR "computational protein design" OR "generative protein design")'
TODAY = datetime.now().date().isoformat()
NOW_STAMP = datetime.now().strftime("%Y%m%d-%H%M%S")
NOW_LOG = datetime.now().strftime("%Y-%m-%d %H:%M")

MAX_RECORDS = 100
PUBMED_RETMAX = 70
ARXIV_RETMAX = 50
ARXIV_DELAY_SECONDS = 3.2
ARXIV_MAX_RETRIES = 3
ARXIV_BACKOFF_SECONDS = [30, 60, 120]
ARXIV_CACHE_MAX_AGE_SECONDS = 24 * 60 * 60

USER_AGENT = "ResearchLibraryMetadataMaintenance/1.0 (local catalogue script)"

COLUMNS = [
    "id",
    "title",
    "authors",
    "year",
    "journal",
    "journal_abbrev",
    "doi",
    "pmid",
    "arxiv_id",
    "publication_type",
    "abstract",
    "keywords",
    "auto_tags",
    "manual_tags",
    "suggested_topic",
    "topic_folder",
    "pdf_status",
    "pdf_filename",
    "source",
    "url",
    "date_added",
    "date_updated",
    "notes",
    "uncertainty",
]

PROTECTED_EXISTING_FIELDS = {"manual_tags", "notes", "topic_folder"}


@dataclass
class Record:
    values: OrderedDict[str, str] = field(default_factory=OrderedDict)

    def __getitem__(self, key: str) -> str:
        return self.values.get(key, "")

    def __setitem__(self, key: str, value: str) -> None:
        self.values[key] = clean_text(value)


def clean_text(value: object) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def normalize_title(title: str) -> str:
    return re.sub(r"\s+", " ", re.sub(r"[^a-z0-9]+", " ", title.lower())).strip()


def normalize_doi(doi: str) -> str:
    doi = clean_text(doi).lower()
    doi = re.sub(r"^https?://(dx\.)?doi\.org/", "", doi)
    return doi


def request_text(url: str, attempts: int = 3, delay: int = 8) -> str:
    last_error: Exception | None = None
    for attempt in range(1, attempts + 1):
        req = urllib.request.Request(url, headers={"User-Agent": USER_AGENT})
        try:
            with urllib.request.urlopen(req, timeout=45) as response:
                return response.read().decode("utf-8", errors="replace")
        except Exception as exc:  # network and rate-limit recovery
            last_error = exc
            if attempt < attempts:
                time.sleep(delay)
    raise RuntimeError(f"Request failed after {attempts} attempts: {url}") from last_error


def parse_retry_after(value: str | None) -> int | None:
    if not value:
        return None
    value = value.strip()
    if value.isdigit():
        return max(0, int(value))
    try:
        retry_at = parsedate_to_datetime(value)
    except (TypeError, ValueError):
        return None
    if retry_at.tzinfo is None:
        return None
    delay = int((retry_at - datetime.now(retry_at.tzinfo)).total_seconds())
    return max(0, delay)


def read_cached_arxiv_response(url: str) -> str | None:
    if not RAW_ARXIV_CACHE.exists():
        return None
    try:
        cached = json.loads(RAW_ARXIV_CACHE.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return None
    if cached.get("url") != url or not cached.get("raw_response"):
        return None
    fetched_at = cached.get("fetched_at")
    try:
        fetched_dt = datetime.fromisoformat(fetched_at)
    except (TypeError, ValueError):
        return None
    age_seconds = (datetime.now() - fetched_dt).total_seconds()
    if age_seconds > ARXIV_CACHE_MAX_AGE_SECONDS:
        return None
    return cached["raw_response"]


def write_cached_arxiv_response(url: str, raw_response: str) -> None:
    payload = {
        "query": QUERY,
        "url": url,
        "fetched_at": datetime.now().isoformat(timespec="seconds"),
        "format": "atom_xml",
        "raw_response": raw_response,
    }
    RAW_ARXIV_CACHE.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def request_arxiv_text(url: str) -> str:
    cached = read_cached_arxiv_response(url)
    if cached is not None:
        return cached

    last_error: Exception | None = None
    time.sleep(ARXIV_DELAY_SECONDS)
    for attempt in range(ARXIV_MAX_RETRIES + 1):
        req = urllib.request.Request(url, headers={"User-Agent": USER_AGENT})
        try:
            with urllib.request.urlopen(req, timeout=45) as response:
                raw = response.read().decode("utf-8", errors="replace")
                write_cached_arxiv_response(url, raw)
                return raw
        except urllib.error.HTTPError as exc:
            last_error = exc
            if exc.code != 429 or attempt >= ARXIV_MAX_RETRIES:
                break
            retry_after = parse_retry_after(exc.headers.get("Retry-After"))
            delay = retry_after if retry_after is not None else ARXIV_BACKOFF_SECONDS[attempt]
            time.sleep(max(delay, ARXIV_DELAY_SECONDS))
        except Exception as exc:
            last_error = exc
            break
    raise RuntimeError(f"arXiv request failed after rate-limit compliant retries: {url}") from last_error


def node_text(node: ET.Element | None) -> str:
    if node is None:
        return ""
    return clean_text("".join(node.itertext()))


def child_text(node: ET.Element | None, path: str) -> str:
    if node is None:
        return ""
    return node_text(node.find(path))


def infer_publication_type(pub_types: Iterable[str]) -> str:
    joined = "; ".join([p for p in pub_types if p])
    low = joined.lower()
    if "review" in low:
        return "review"
    if "editorial" in low:
        return "editorial"
    if "letter" in low:
        return "letter"
    if "comment" in low:
        return "commentary"
    if "meta-analysis" in low:
        return "meta-analysis"
    return joined or "research article"


def auto_tags(title: str, abstract: str, keywords: str, publication_type: str) -> str:
    hay = f"{title} {abstract} {keywords} {publication_type}".lower()
    rules = [
        ("de novo protein design", ["de novo protein design", "de-novo protein design"]),
        ("computational protein design", ["computational protein design", "protein design"]),
        ("generative protein design", ["generative protein design", "generative model", "diffusion", "language model"]),
        ("AI/ML", ["machine learning", "deep learning", "neural network", "artificial intelligence"]),
        ("protein structure prediction", ["structure prediction", "alphafold", "rosettafold"]),
        ("binder design", ["binder", "binding protein", "protein binder"]),
        ("enzyme design", ["enzyme design", "catalyst", "catalytic"]),
        ("protein engineering", ["protein engineering", "engineered protein"]),
        ("review", ["review"]),
    ]
    tags: list[str] = []
    for tag, patterns in rules:
        if any(pattern in hay for pattern in patterns) and tag not in tags:
            tags.append(tag)
    return "; ".join(tags)


def make_record(**kwargs: str) -> Record:
    row = OrderedDict((column, "") for column in COLUMNS)
    for key, value in kwargs.items():
        if key in row:
            row[key] = clean_text(value)
    row["auto_tags"] = auto_tags(row["title"], row["abstract"], row["keywords"], row["publication_type"])
    row["suggested_topic"] = row["suggested_topic"] or "de novo protein design"
    row["topic_folder"] = row["topic_folder"] or "Unclassified"
    row["pdf_status"] = row["pdf_status"] or "not_downloaded"
    row["date_added"] = row["date_added"] or TODAY
    row["date_updated"] = row["date_updated"] or TODAY
    return Record(row)


def fetch_pubmed_records() -> list[Record]:
    term = urllib.parse.quote(QUERY)
    esearch_url = (
        "https://eutils.ncbi.nlm.nih.gov/entrez/eutils/esearch.fcgi"
        f"?db=pubmed&retmode=json&retmax={PUBMED_RETMAX}&sort=relevance&term={term}"
    )
    esearch = json.loads(request_text(esearch_url))
    pmids = esearch.get("esearchresult", {}).get("idlist", [])
    if not pmids:
        return []

    efetch_url = (
        "https://eutils.ncbi.nlm.nih.gov/entrez/eutils/efetch.fcgi"
        f"?db=pubmed&retmode=xml&id={','.join(pmids)}"
    )
    root = ET.fromstring(request_text(efetch_url))
    records: list[Record] = []
    for pubmed_article in root.findall("PubmedArticle"):
        medline = pubmed_article.find("MedlineCitation")
        article = medline.find("Article") if medline is not None else None
        pubmed_data = pubmed_article.find("PubmedData")
        pmid = child_text(medline, "PMID")
        title = child_text(article, "ArticleTitle")
        journal = child_text(article, "Journal/Title")
        journal_abbrev = child_text(article, "Journal/ISOAbbreviation")
        year = child_text(article, "Journal/JournalIssue/PubDate/Year")
        if not year:
            medline_date = child_text(article, "Journal/JournalIssue/PubDate/MedlineDate")
            match = re.search(r"\d{4}", medline_date)
            year = match.group(0) if match else ""

        abstract = " ".join(node_text(n) for n in article.findall("Abstract/AbstractText")) if article is not None else ""
        authors = []
        if article is not None:
            for author in article.findall("AuthorList/Author"):
                name = " ".join(part for part in [child_text(author, "LastName"), child_text(author, "ForeName")] if part)
                if name:
                    authors.append(name)
        author_text = "; ".join(authors[:6]) + ("; et al." if len(authors) > 6 else "")

        doi = ""
        if pubmed_data is not None:
            for article_id in pubmed_data.findall("ArticleIdList/ArticleId"):
                if article_id.attrib.get("IdType") == "doi":
                    doi = node_text(article_id)
                    break

        keywords = []
        if medline is not None:
            keywords.extend(node_text(k) for k in medline.findall("KeywordList/Keyword"))
            keywords.extend(node_text(m.find("DescriptorName")) for m in medline.findall("MeshHeadingList/MeshHeading"))
        keywords_text = "; ".join(dict.fromkeys(k for k in keywords if k))

        pub_types = [node_text(p) for p in article.findall("PublicationTypeList/PublicationType")] if article is not None else []
        publication_type = infer_publication_type(pub_types)

        records.append(
            make_record(
                id=pmid or doi,
                title=title,
                authors=author_text,
                year=year,
                journal=journal,
                journal_abbrev=journal_abbrev,
                doi=doi,
                pmid=pmid,
                arxiv_id="",
                publication_type=publication_type,
                abstract=abstract,
                keywords=keywords_text,
                source="PubMed",
                url=f"https://pubmed.ncbi.nlm.nih.gov/{pmid}/" if pmid else "",
            )
        )
    return records


def fetch_arxiv_records() -> list[Record]:
    search_query = urllib.parse.quote('all:"de novo protein design" OR all:"computational protein design" OR all:"generative protein design"')
    url = (
        "https://export.arxiv.org/api/query"
        f"?search_query={search_query}&start=0&max_results={ARXIV_RETMAX}&sortBy=relevance&sortOrder=descending"
    )
    try:
        root = ET.fromstring(request_arxiv_text(url))
    except RuntimeError as exc:
        return load_arxiv_records_from_existing_preview(str(exc))
    ns = {"atom": "http://www.w3.org/2005/Atom", "arxiv": "http://arxiv.org/schemas/atom"}
    records: list[Record] = []
    for entry in root.findall("atom:entry", ns):
        id_url = child_text(entry, "atom:id")
        arxiv_id = re.sub(r"v\d+$", "", id_url.rsplit("/abs/", 1)[-1])
        title = child_text(entry, "atom:title")
        abstract = child_text(entry, "atom:summary")
        published = child_text(entry, "atom:published")
        year_match = re.search(r"\d{4}", published)
        year = year_match.group(0) if year_match else ""
        authors = [child_text(author, "atom:name") for author in entry.findall("atom:author", ns)]
        author_text = "; ".join(authors[:6]) + ("; et al." if len(authors) > 6 else "")
        doi = child_text(entry, "arxiv:doi")
        categories = [cat.attrib.get("term", "") for cat in entry.findall("atom:category", ns)]
        records.append(
            make_record(
                id=arxiv_id or doi or normalize_title(title),
                title=title,
                authors=author_text,
                year=year,
                journal="arXiv",
                journal_abbrev="arXiv",
                doi=doi,
                pmid="",
                arxiv_id=arxiv_id,
                publication_type="preprint",
                abstract=abstract,
                keywords="; ".join(c for c in categories if c),
                source="arXiv",
                url=id_url,
            )
        )
    return records


def load_arxiv_records_from_existing_preview(reason: str) -> list[Record]:
    if not PREVIEW_CSV.exists():
        raise RuntimeError("arXiv API failed and no existing dry_run_catalogue_preview.csv is available for fallback.")
    records: list[Record] = []
    with PREVIEW_CSV.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        for row in reader:
            source = clean_text(row.get("source", ""))
            if "arxiv" not in source.lower():
                continue
            record = make_record(
                id=row.get("arxiv_id") or row.get("id") or normalize_title(row.get("title", "")),
                title=row.get("title", ""),
                authors=row.get("authors", ""),
                year=row.get("year", ""),
                journal=row.get("journal", "arXiv") or "arXiv",
                journal_abbrev=row.get("journal_abbrev", "arXiv") or "arXiv",
                doi=row.get("doi", ""),
                pmid="",
                arxiv_id=row.get("arxiv_id", ""),
                publication_type=row.get("publication_type", "preprint") or "preprint",
                abstract=row.get("abstract", ""),
                keywords=row.get("keywords", ""),
                source="arXiv",
                url=row.get("url", ""),
            )
            uncertainty = row.get("uncertainty", "")
            fallback_note = f"arXiv API rate-limited during confirmed run; reused prior preview metadata. API error: {reason}"
            record["uncertainty"] = " | ".join(x for x in [uncertainty, fallback_note] if x)
            records.append(record)
    if not records:
        raise RuntimeError("arXiv API failed and existing preview contains no arXiv records.")
    return records


def merge_sources(existing: str, incoming: str) -> str:
    values: list[str] = []
    for source in (existing, incoming):
        for part in source.split(";"):
            part = part.strip()
            if part and part not in values:
                values.append(part)
    return "; ".join(values)


def merge_record(existing: Record, incoming: Record, reason: str) -> None:
    for column in COLUMNS:
        if column in PROTECTED_EXISTING_FIELDS:
            continue
        if column == "source":
            existing[column] = merge_sources(existing[column], incoming[column])
        elif column == "date_added":
            existing[column] = existing[column] or incoming[column] or TODAY
        elif column == "date_updated":
            existing[column] = TODAY
        elif not existing[column] and incoming[column]:
            existing[column] = incoming[column]
    note = f"Possible duplicate merged by {reason}: {incoming['source']} record {incoming['id']}"
    existing["uncertainty"] = " | ".join(x for x in [existing["uncertainty"], note] if x)


def deduplicate(records: list[Record]) -> tuple[list[Record], list[tuple[str, str, str]]]:
    kept: list[Record] = []
    duplicate_log: list[tuple[str, str, str]] = []
    by_doi: dict[str, Record] = {}
    by_pmid: dict[str, Record] = {}
    by_arxiv: dict[str, Record] = {}
    by_title: dict[str, Record] = {}

    for record in records:
        keys = [
            ("DOI", normalize_doi(record["doi"]), by_doi),
            ("PMID", record["pmid"], by_pmid),
            ("arXiv ID", record["arxiv_id"], by_arxiv),
            ("normalized title", normalize_title(record["title"]), by_title),
        ]
        duplicate = None
        reason = ""
        for label, key, index in keys:
            if key and key in index:
                duplicate = index[key]
                reason = label
                break
        if duplicate is not None:
            duplicate_log.append((record["title"], duplicate["title"], reason))
            merge_record(duplicate, record, reason)
            continue

        kept.append(record)
        if record["doi"]:
            by_doi[normalize_doi(record["doi"])] = record
        if record["pmid"]:
            by_pmid[record["pmid"]] = record
        if record["arxiv_id"]:
            by_arxiv[record["arxiv_id"]] = record
        if record["title"]:
            by_title[normalize_title(record["title"])] = record
        if len(kept) >= MAX_RECORDS:
            break
    return kept, duplicate_log


def read_existing_catalogue() -> tuple[Workbook | None, object | None, list[str], list[dict[str, str]]]:
    if not CATALOGUE.exists():
        return None, None, COLUMNS[:], []
    wb = load_workbook(CATALOGUE)
    ws = wb.active
    headers = [clean_text(cell.value) for cell in ws[1]]
    rows: list[dict[str, str]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        rows.append({header: clean_text(value) for header, value in zip(headers, row)})
    return wb, ws, headers, rows


def existing_indexes(rows: list[dict[str, str]]) -> dict[str, dict[str, dict[str, str]]]:
    indexes = {"doi": {}, "pmid": {}, "arxiv_id": {}, "title": {}}
    for row in rows:
        if row.get("doi"):
            indexes["doi"][normalize_doi(row["doi"])] = row
        if row.get("pmid"):
            indexes["pmid"][row["pmid"]] = row
        if row.get("arxiv_id"):
            indexes["arxiv_id"][row["arxiv_id"]] = row
        if row.get("title"):
            indexes["title"][normalize_title(row["title"])] = row
    return indexes


def write_preview_csv(records: list[Record]) -> None:
    with PREVIEW_CSV.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=COLUMNS)
        writer.writeheader()
        for record in records:
            writer.writerow({column: record[column] for column in COLUMNS})


def style_sheet(ws) -> None:
    header_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    widths = {
        "A": 18, "B": 60, "C": 42, "D": 10, "E": 32, "F": 18, "G": 28, "H": 14,
        "I": 16, "J": 18, "K": 80, "L": 40, "M": 36, "N": 18, "O": 24, "P": 18,
        "Q": 18, "R": 26, "S": 18, "T": 42, "U": 14, "V": 14, "W": 42, "X": 50,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def append_or_update_catalogue(records: list[Record]) -> tuple[int, int, str]:
    backup_path = ""
    wb, ws, headers, existing_rows = read_existing_catalogue()
    if CATALOGUE.exists():
        backup = ROOT / f"catalogue.backup.{NOW_STAMP}.xlsx"
        shutil.copy2(CATALOGUE, backup)
        backup_path = backup.name

    if wb is None or ws is None:
        wb = Workbook()
        ws = wb.active
        ws.title = "catalogue"
        headers = COLUMNS[:]
        ws.append(headers)
        existing_rows = []
    else:
        for column in COLUMNS:
            if column not in headers:
                headers.append(column)
                ws.cell(row=1, column=len(headers), value=column)

    header_to_col = {header: index + 1 for index, header in enumerate(headers)}
    indexes = existing_indexes(existing_rows)
    added = 0
    updated = 0

    for record in records:
        matched = None
        for field, key in [
            ("doi", normalize_doi(record["doi"])),
            ("pmid", record["pmid"]),
            ("arxiv_id", record["arxiv_id"]),
            ("title", normalize_title(record["title"])),
        ]:
            if key and key in indexes[field]:
                matched = indexes[field][key]
                break

        if matched is None:
            ws.append([record[column] if column in COLUMNS else "" for column in headers])
            added += 1
            row_dict = {column: record[column] for column in headers}
            existing_rows.append(row_dict)
            indexes = existing_indexes(existing_rows)
            continue

        row_number = existing_rows.index(matched) + 2
        changed = False
        for column in headers:
            if column in PROTECTED_EXISTING_FIELDS:
                continue
            if column == "source":
                new_value = merge_sources(matched.get(column, ""), record[column])
            elif column == "date_updated":
                new_value = TODAY
            elif column == "date_added":
                new_value = matched.get(column, "") or record[column] or TODAY
            else:
                new_value = matched.get(column, "") or record[column]
            if clean_text(matched.get(column, "")) != clean_text(new_value):
                ws.cell(row=row_number, column=header_to_col[column], value=new_value)
                matched[column] = new_value
                changed = True
        if changed:
            updated += 1

    style_sheet(ws)
    wb.save(CATALOGUE)
    return added, updated, backup_path


def append_change_log(added: int, updated: int, backup_path: str, duplicates: list[tuple[str, str, str]]) -> None:
    backup_text = backup_path or "None; catalogue.xlsx did not previously exist."
    text = f"""
## {NOW_LOG}

Action:
Created/updated preliminary de novo protein design catalogue from PubMed and arXiv metadata.

Files changed:
- scripts/search_literature.py
- dry_run_catalogue_preview.csv
- catalogue.xlsx
- library_changes.md

Catalogue rows changed:
- Added rows: {added}
- Updated existing rows: {updated}
- Dry-run duplicate merges before write: {len(duplicates)}

Reason:
User confirmed proposed_plan.md and requested catalogue creation/update using PubMed Entrez API, arXiv API, and openpyxl.

Uncertainty:
arXiv API availability can vary due to rate limiting. Records are metadata-derived and topic_folder/manual_tags/notes were not overwritten. Backup: {backup_text}
"""
    with CHANGE_LOG.open("a", encoding="utf-8") as handle:
        handle.write(text.lstrip())


def main() -> None:
    pubmed = fetch_pubmed_records()
    time.sleep(3)
    arxiv = fetch_arxiv_records()
    records, duplicates = deduplicate(pubmed + arxiv)
    write_preview_csv(records)
    added, updated, backup_path = append_or_update_catalogue(records)
    append_change_log(added, updated, backup_path, duplicates)
    print(
        json.dumps(
            {
                "pubmed_records": len(pubmed),
                "arxiv_records": len(arxiv),
                "deduplicated_records": len(records),
                "duplicates_merged": len(duplicates),
                "catalogue_rows_added": added,
                "catalogue_rows_updated": updated,
                "backup": backup_path,
            },
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
