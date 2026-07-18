from __future__ import annotations

from copy import deepcopy
import json
from pathlib import Path

from openpyxl import load_workbook

from conftest import write_text_pdf
from lam.config import Settings
from lam.models import (
    CatalogueRecord,
    IdentifierCandidate,
    MetadataLookupResult,
    MetadataLookupStatus,
    MetadataRecord,
    PdfInspection,
    WorkflowStatus,
)
from lam.providers.unavailable import UnavailableMetadataService
from lam.services.catalogue_service import CatalogueService
from lam.services.pdf_service import PdfService
from lam.utils.journal import journal_is_variant, journals_equivalent, normalize_journal_name
from lam.utils.uncertainty import confirmation_for, confirmed_value, parse_user_confirmations
from lam.workflows.inbox_register import InboxRegisterWorkflow


class RecordingMetadataService:
    def __init__(self, resolver):
        self.resolver = resolver
        self.requests = []

    def lookup(self, request):
        self.requests.append(deepcopy(request))
        return self.resolver(request)


def found(metadata: MetadataRecord, confidence: str = "exact_identifier"):
    return MetadataLookupResult(
        MetadataLookupStatus.FOUND,
        records=[metadata.to_dict()],
        best_record=metadata.to_dict(),
        confidence=confidence,
        providers_used=list(metadata.source),
        selection_reason="Exact test identity.",
    )


def unavailable(_request):
    return MetadataLookupResult(
        MetadataLookupStatus.UNAVAILABLE,
        selection_reason="Provider unavailable.",
    )


def inspection(**overrides):
    values = {
        "relative_path": "Inbox/test.pdf",
        "filename": "Test Journal, 2025 - Test Identity Paper.pdf",
        "size": 100,
        "mtime_ns": 200,
        "is_readable": True,
        "local_metadata": {},
    }
    values.update(overrides)
    return PdfInspection(**values)


def provisional_record(**overrides):
    values = {
        "paper_uuid": "00000000-0000-4000-8000-000000000123",
        "title": "Test Identity Paper",
        "authors": "Alice Smith",
        "year": "2025",
        "journal": "Test Journal",
        "doi": "",
        "pmid": "",
        "uncertainty": "",
    }
    values.update(overrides)
    return CatalogueRecord(2, values)


def seed_provisional_blocker(root: Path, relative_path: str, row_number: int = 2) -> None:
    workbook = load_workbook(root / "catalogue.xlsx")
    sheet = workbook["Catalogue"]
    headers = {cell.value: cell.column for cell in sheet[1]}
    paper_uuid = sheet.cell(row_number, headers["paper_uuid"]).value
    source = root / relative_path
    stat = source.stat()
    state_dir = root / ".library_state"
    state_dir.mkdir(exist_ok=True)
    (state_dir / "inbox_blockers.json").write_text(
        json.dumps(
            {
                "version": 1,
                "files": [
                    {
                        "stable_file_id": (
                            f"{relative_path}|{stat.st_size}|{stat.st_mtime_ns}"
                        ),
                        "source_path": relative_path,
                        "paper_uuid": paper_uuid,
                        "size": stat.st_size,
                        "mtime_ns": stat.st_mtime_ns,
                        "issue_keys": ["metadata_identity_unconfirmed"],
                    }
                ],
            },
            indent=2,
        )
        + "\n",
        encoding="utf-8",
    )


def test_01_naked_user_confirmed_defaults_to_paper_identity():
    item = parse_user_confirmations("USER_CONFIRMED")[0]
    assert (item.field, item.value) == ("paper_identity", "")


def test_02_colon_only_user_confirmed_defaults_to_paper_identity():
    item = parse_user_confirmations("USER_CONFIRMED:")[0]
    assert (item.field, item.value) == ("paper_identity", "")


def test_03_field_only_confirmation_is_valid():
    item = confirmation_for("USER_CONFIRMED: field=paper_identity", "paper_identity")
    assert item is not None and item.value == ""


def test_04_empty_value_confirmation_is_valid():
    assert confirmation_for(
        "USER_CONFIRMED: field=paper_identity; value=", "paper_identity"
    ) is not None


def test_05_nonempty_confirmation_value_is_returned():
    assert confirmed_value(
        "USER_CONFIRMED: field=paper_identity; value=PMID:12345678",
        "paper_identity",
    ) == "PMID:12345678"


def test_06_confirmation_parser_preserves_original_line():
    raw = "USER_CONFIRMED: field=paper_identity; value=; note=keep"
    assert parse_user_confirmations(raw)[0].raw == raw


def test_07_bare_confirmation_resolves_paper_identity_blocker(library_factory):
    root = library_factory(
        [{"id": "LOCAL:x", "title": "Example", "uncertainty": "USER_CONFIRMED:"}]
    )
    service = CatalogueService(root / "catalogue.xlsx")
    record = service.load()[0]
    outcome = service.ensure_review_blocker(
        record, "paper_identity", "Pending", issue_key="metadata_identity_unconfirmed"
    )
    assert outcome == "confirmed"
    assert record.get("uncertainty") == "USER_CONFIRMED:"


def test_08_confirmation_cannot_suppress_hard_conflict(library_factory):
    root = library_factory(
        [{"id": "LOCAL:x", "title": "Example", "uncertainty": "USER_CONFIRMED"}]
    )
    service = CatalogueService(root / "catalogue.xlsx")
    record = service.load()[0]
    outcome = service.ensure_review_blocker(
        record,
        "paper_identity",
        "DOI conflict",
        issue_key="metadata_identifier_conflict",
        conflict_with_confirmation=True,
    )
    assert outcome == "added"
    assert "metadata_identifier_conflict" in record.get("uncertainty")


def test_09_deleted_blocker_is_a_one_time_clearance(library_factory):
    root = library_factory([{"id": "LOCAL:x", "title": "Example"}])
    first = CatalogueService(root / "catalogue.xlsx")
    record = first.load()[0]
    first.ensure_review_blocker(
        record, "paper_identity", "Pending", issue_key="metadata_identity_unconfirmed"
    )
    previous = first.snapshot_payload()
    first.update_fields(record, {"uncertainty": ""})
    second = CatalogueService(root / "catalogue.xlsx")
    second.load()
    second.configure_review_state(previous)
    assert second.ensure_review_blocker(
        second.records[0], "paper_identity", "Pending", issue_key="metadata_identity_unconfirmed"
    ) == "cleared"


def test_10_changed_identifier_allows_new_blocker_after_clearance(library_factory):
    root = library_factory([{"id": "LOCAL:x", "title": "Example"}])
    first = CatalogueService(root / "catalogue.xlsx")
    record = first.load()[0]
    first.ensure_review_blocker(record, "paper_identity", "Pending", issue_key="same")
    previous = first.snapshot_payload()
    first.update_fields(record, {"uncertainty": ""})
    first.save_atomic()
    second = CatalogueService(root / "catalogue.xlsx")
    record = second.load()[0]
    second.configure_review_state(previous)
    assert second.ensure_review_blocker(record, "paper_identity", "Pending", issue_key="same") == "cleared"
    cleared_snapshot = second.snapshot_payload()
    second.update_fields(record, {"doi": "10.1000/new"})
    second.save_atomic()
    third = CatalogueService(root / "catalogue.xlsx")
    record = third.load()[0]
    third.configure_review_state(cleared_snapshot)
    assert third.ensure_review_blocker(record, "paper_identity", "Pending", issue_key="same") == "added"


def test_11_resolving_confirmation_preserves_free_text(library_factory):
    root = library_factory(
        [{"id": "LOCAL:x", "title": "Example", "uncertainty": "keep me\nUSER_CONFIRMED"}]
    )
    service = CatalogueService(root / "catalogue.xlsx")
    record = service.load()[0]
    service.ensure_review_blocker(record, "paper_identity", "Pending", issue_key="pending")
    assert "keep me" in record.get("uncertainty")
    assert "USER_CONFIRMED" in record.get("uncertainty")


def test_12_catalogue_pmid_is_first_lookup_request():
    record = provisional_record(pmid="12345678", doi="10.1000/test")
    requests = InboxRegisterWorkflow._lookup_requests(record, inspection(), "Inbox/test.pdf")
    assert requests[0].pmid == "12345678"


def test_13_catalogue_doi_precedes_title_when_pmid_absent():
    record = provisional_record(doi="10.1000/test")
    requests = InboxRegisterWorkflow._lookup_requests(record, inspection(), "Inbox/test.pdf")
    assert requests[0].doi == "10.1000/test"


def test_14_pmid_request_carries_catalogue_context():
    record = provisional_record(pmid="12345678")
    request = InboxRegisterWorkflow._lookup_requests(record, inspection(), "Inbox/test.pdf")[0]
    assert request.paper_uuid == "00000000-0000-4000-8000-000000000123"
    assert request.authors == "Alice Smith"
    assert request.year == "2025"


def test_15_request_marks_shorthand_user_confirmation():
    record = provisional_record(uncertainty="USER_CONFIRMED:")
    request = InboxRegisterWorkflow._lookup_requests(record, inspection(), "Inbox/test.pdf")[0]
    assert request.user_confirmed_identity is True


def test_16_catalogue_pdf_identifier_conflict_is_detected():
    record = provisional_record(doi="10.1000/catalogue")
    pdf = inspection(
        doi_candidates=[
            IdentifierCandidate("10.1000/pdf", 1, "doi", "high", "first_page")
        ]
    )
    _, _, conflicts = InboxRegisterWorkflow.build_merged_identity_evidence(record, None, pdf)
    assert "doi_catalogue_pdf_conflict" in conflicts


def test_17_provider_incomplete_record_uses_catalogue_authors_and_year():
    record = provisional_record(doi="10.1000/test")
    provider = MetadataRecord(
        title="Test Identity Paper", doi="10.1000/test", source=["unpaywall"]
    )
    merged, sources, conflicts = InboxRegisterWorkflow.build_merged_identity_evidence(
        record, provider, inspection(), lookup_confidence="exact_identifier"
    )
    assert merged.authors == ["Alice Smith"]
    assert merged.year == "2025"
    assert sources["authors"] == "catalogue"
    assert conflicts == []


def test_18_exact_provider_plus_merged_fields_is_durable():
    record = provisional_record(doi="10.1000/test")
    provider = MetadataRecord(title="Test Identity Paper", doi="10.1000/test")
    pdf = inspection()
    merged, sources, conflicts = InboxRegisterWorkflow.build_merged_identity_evidence(
        record, provider, pdf, lookup_confidence="exact_identifier"
    )
    durable, reason = InboxRegisterWorkflow.validate_durable_identity(
        merged,
        sources=sources,
        conflicts=conflicts,
        user_confirmed_identity=False,
        lookup_confidence="exact_identifier",
        inspection=pdf,
    )
    assert durable is True and reason == "provider_identifier_merged"


def test_19_merged_evidence_without_core_fields_is_incomplete():
    pdf = inspection(filename="unknown.pdf")
    merged, sources, conflicts = InboxRegisterWorkflow.build_merged_identity_evidence(
        provisional_record(title="", authors="", year="", journal=""), None, pdf
    )
    durable, reason = InboxRegisterWorkflow.validate_durable_identity(
        merged,
        sources=sources,
        conflicts=conflicts,
        user_confirmed_identity=False,
        lookup_confidence="",
        inspection=pdf,
    )
    assert durable is False and reason == "durable_identity_incomplete"


def test_20_user_confirmation_releases_completeness_block():
    record = provisional_record(uncertainty="USER_CONFIRMED")
    pdf = inspection()
    merged, sources, conflicts = InboxRegisterWorkflow.build_merged_identity_evidence(
        record, None, pdf
    )
    durable, reason = InboxRegisterWorkflow.validate_durable_identity(
        merged,
        sources=sources,
        conflicts=conflicts,
        user_confirmed_identity=True,
        lookup_confidence="",
        inspection=pdf,
    )
    assert durable is True and reason == "user_confirmed_merged"


def test_21_user_confirmation_does_not_release_identifier_conflict():
    record = provisional_record(uncertainty="USER_CONFIRMED")
    pdf = inspection()
    merged, sources, _ = InboxRegisterWorkflow.build_merged_identity_evidence(record, None, pdf)
    durable, reason = InboxRegisterWorkflow.validate_durable_identity(
        merged,
        sources=sources,
        conflicts=["doi_catalogue_pdf_conflict"],
        user_confirmed_identity=True,
        lookup_confidence="",
        inspection=pdf,
    )
    assert durable is False and reason == "metadata_identifier_conflict"


def test_22_first_page_extractor_structures_doi_authors_and_abstract(library_factory):
    root = library_factory([])
    path = root / "Inbox" / "Test Journal, 2025, Review - Local Evidence Paper.pdf"
    write_text_pdf(
        path,
        [
            "Local Evidence Paper\nAlice Smith; Bob Jones\nTest Journal (2025) 12:1\n"
            "https://doi.org/10.1000/local\nAbstract\n"
            + "This is a sufficiently detailed abstract sentence describing the methods and results. " * 2
            + "\nKeywords testing"
        ],
    )
    result = PdfService(Settings.from_root(root)).inspect(path, ocr_mode="never")
    assert result.local_metadata["doi"] == "10.1000/local"
    assert result.local_metadata["authors"] == ("Alice Smith", "Bob Jones")
    assert result.local_metadata["abstract"].startswith("This is a sufficiently")


def test_23_high_quality_local_metadata_does_not_replace_unconfirmed_identity(library_factory):
    name = "Test Journal, 2025 - Local Evidence Paper.pdf"
    root = library_factory(
        [{"id": "LOCAL:x", "pdf_status": "inbox", "pdf_filename": name, "pdf_relative_path": f"Inbox/{name}"}]
    )
    write_text_pdf(
        root / "Inbox" / name,
        ["Local Evidence Paper\nAlice Smith; Bob Jones\nTest Journal (2025) 12:1\nAbstract\n" + "Long abstract text for reliable local evidence. " * 4 + "\nKeywords test"],
    )
    settings = Settings.from_root(root)
    settings.ensure_runtime_directories()
    result = InboxRegisterWorkflow(settings, UnavailableMetadataService()).run()
    assert result.status == WorkflowStatus.NEEDS_REVIEW
    assert (root / "Inbox" / name).exists()
    assert not (root / "Registered" / name).exists()
    assert result.details["files"][0]["result_status"] == "provisional"


def test_24_low_quality_local_metadata_stays_provisional(library_factory):
    root = library_factory(
        [{"id": "LOCAL:x", "pdf_status": "inbox", "pdf_filename": "scan.pdf", "pdf_relative_path": "Inbox/scan.pdf"}]
    )
    write_text_pdf(root / "Inbox" / "scan.pdf", ["Unclear page"])
    settings = Settings.from_root(root)
    settings.ensure_runtime_directories()
    result = InboxRegisterWorkflow(settings, UnavailableMetadataService()).run()
    assert result.status == WorkflowStatus.NEEDS_REVIEW
    assert (root / "Inbox" / "scan.pdf").exists()


def test_25_local_fallback_does_not_overwrite_existing_author(library_factory):
    name = "Test Journal, 2025 - Local Evidence Paper.pdf"
    root = library_factory(
        [{
            "id": "LOCAL:x",
            "title": "Local Evidence Paper",
            "authors": "User Author",
            "year": "2025",
            "journal": "Test Journal",
            "pdf_status": "inbox",
            "pdf_filename": name,
            "pdf_relative_path": f"Inbox/{name}",
            "uncertainty": "USER_CONFIRMED",
        }]
    )
    write_text_pdf(
        root / "Inbox" / name,
        ["Local Evidence Paper\nDifferent Author; Second Author\nTest Journal (2025) 12:1\nAbstract\n" + "Reliable abstract text. " * 8 + "\nKeywords test"],
    )
    settings = Settings.from_root(root)
    settings.ensure_runtime_directories()
    InboxRegisterWorkflow(settings, UnavailableMetadataService()).run()
    sheet = load_workbook(root / "catalogue.xlsx")["Catalogue"]
    headers = {cell.value: cell.column for cell in sheet[1]}
    assert sheet.cell(2, headers["authors"]).value == "User Author"


def test_26_structure_long_name_has_same_base_journal():
    value = normalize_journal_name("Structure (London, England : 1993)")
    assert value.base_name == "structure"
    assert journals_equivalent("Structure", value.original)


def test_27_structure_long_name_is_a_nonblocking_variant():
    assert journal_is_variant("Structure", "Structure (London, England : 1993)")


def test_28_different_journal_names_remain_different():
    assert not journals_equivalent("Structure", "Cell")


def test_29_pmid_success_does_not_fall_back_to_title(library_factory):
    name = "Test Journal, 2025 - Test Identity Paper.pdf"
    root = library_factory(
        [{
            "id": "LOCAL:x",
            "title": "Test Identity Paper",
            "authors": "Alice Smith",
            "year": "2025",
            "journal": "Test Journal",
            "pmid": "12345678",
            "pdf_status": "inbox",
            "pdf_filename": name,
            "pdf_relative_path": f"Inbox/{name}",
            "uncertainty": "NEEDS_REVIEW: field=paper_identity; issue_key=metadata_identity_unconfirmed; issue=Pending",
        }]
    )
    write_text_pdf(root / "Inbox" / name, ["Test Identity Paper\nAlice Smith\n2025"])
    seed_provisional_blocker(root, f"Inbox/{name}")
    metadata = MetadataRecord(
        canonical_id="PMID:12345678",
        title="Test Identity Paper",
        authors=["Alice Smith"],
        year="2025",
        journal="Test Journal",
        pmid="12345678",
        source=["pubmed"],
    )
    service = RecordingMetadataService(lambda request: found(metadata))
    settings = Settings.from_root(root)
    settings.ensure_runtime_directories()
    InboxRegisterWorkflow(settings, service).run()
    assert len(service.requests) == 1
    assert service.requests[0].pmid == "12345678"


def test_30_incomplete_unpaywall_record_merges_and_registers(library_factory):
    name = "Test Journal, 2025 - Test Identity Paper.pdf"
    root = library_factory(
        [{
            "id": "LOCAL:preserved",
            "title": "Test Identity Paper",
            "authors": "Alice Smith",
            "year": "2025",
            "journal": "Test Journal",
            "doi": "10.1000/merged",
            "pdf_status": "inbox",
            "pdf_filename": name,
            "pdf_relative_path": f"Inbox/{name}",
            "uncertainty": "USER_CONFIRMED:",
        }]
    )
    write_text_pdf(root / "Inbox" / name, ["Test Identity Paper\nAlice Smith\n2025"])
    metadata = MetadataRecord(
        canonical_id="DOI:10.1000/merged",
        title="Test Identity Paper",
        year="2025",
        journal="Test Journal",
        doi="10.1000/merged",
        source=["unpaywall"],
    )
    settings = Settings.from_root(root)
    settings.ensure_runtime_directories()
    result = InboxRegisterWorkflow(
        settings, RecordingMetadataService(lambda request: found(metadata))
    ).run()
    assert result.status == WorkflowStatus.SUCCESS
    sheet = load_workbook(root / "catalogue.xlsx")["Catalogue"]
    headers = {cell.value: cell.column for cell in sheet[1]}
    assert sheet.cell(2, headers["paper_uuid"]).value
    assert sheet.cell(2, headers["doi"]).value == "10.1000/merged"
    assert "id" not in headers
    assert "USER_CONFIRMED:" in sheet.cell(2, headers["uncertainty"]).value


def test_31_journal_variant_generates_note_without_blocking(library_factory):
    name = "Structure, 2002 - Test Identity Paper.pdf"
    root = library_factory(
        [{
            "id": "LOCAL:x",
            "title": "Test Identity Paper",
            "authors": "Alice Smith",
            "year": "2002",
            "journal": "Structure",
            "pmid": "12345678",
            "source": "local_pdf",
            "pdf_status": "inbox",
            "pdf_filename": name,
            "pdf_relative_path": f"Inbox/{name}",
        }]
    )
    write_text_pdf(root / "Inbox" / name, ["Test Identity Paper\nAlice Smith\n2002"])
    seed_provisional_blocker(root, f"Inbox/{name}")
    metadata = MetadataRecord(
        canonical_id="PMID:12345678",
        title="Test Identity Paper",
        authors=["Alice Smith"],
        year="2002",
        journal="Structure (London, England : 1993)",
        journal_abbrev="Structure",
        pmid="12345678",
        source=["pubmed"],
    )
    settings = Settings.from_root(root)
    settings.ensure_runtime_directories()
    result = InboxRegisterWorkflow(
        settings, RecordingMetadataService(lambda request: found(metadata))
    ).run()
    assert result.status == WorkflowStatus.SUCCESS
    sheet = load_workbook(root / "catalogue.xlsx")["Catalogue"]
    headers = {cell.value: cell.column for cell in sheet[1]}
    assert sheet.cell(2, headers["journal"]).value == "Structure (London, England : 1993)"
    assert sheet.cell(2, headers["journal_abbrev"]).value == "Structure"
    assert sheet.cell(2, headers["uncertainty"]).value in (None, "")


def test_32_genuinely_different_provider_journal_blocks(library_factory):
    name = "Structure, 2002 - Test Identity Paper.pdf"
    root = library_factory(
        [{
            "id": "LOCAL:x",
            "title": "Test Identity Paper",
            "authors": "Alice Smith",
            "year": "2002",
            "journal": "Structure",
            "pmid": "12345678",
            "source": "local_pdf",
            "pdf_status": "inbox",
            "pdf_filename": name,
            "pdf_relative_path": f"Inbox/{name}",
            "uncertainty": "USER_CONFIRMED",
        }]
    )
    write_text_pdf(root / "Inbox" / name, ["Test Identity Paper\nAlice Smith\n2002"])
    seed_provisional_blocker(root, f"Inbox/{name}")
    metadata = MetadataRecord(
        canonical_id="PMID:12345678",
        title="Test Identity Paper",
        authors=["Alice Smith"],
        year="2002",
        journal="Cell",
        pmid="12345678",
        source=["pubmed"],
    )
    settings = Settings.from_root(root)
    settings.ensure_runtime_directories()
    result = InboxRegisterWorkflow(
        settings, RecordingMetadataService(lambda request: found(metadata))
    ).run()
    assert result.status == WorkflowStatus.NEEDS_REVIEW
    assert (root / "Inbox" / name).exists()
    assert "metadata_journal_conflict" in result.details["files"][0]["issue_keys"]
