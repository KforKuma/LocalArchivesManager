from __future__ import annotations

import json
from pathlib import Path

from openpyxl import load_workbook

from lam.cli import main
from lam.config import Settings
from lam.exceptions import CatalogueError
from lam.services.catalogue_service import CatalogueService
from lam.workflows.paper_delete import PaperDeleteWorkflow
from lam.workflows.recovery import RecoveryWorkflow
from lam.workflows.cleanup import CleanupWorkflow


PAPER_UUID = "12345678-1234-4234-9234-123456789abc"


def _library(current_library_factory, *, with_file: bool = True) -> Path:
    relative = "Topics/Topic_A/paper.pdf"
    root = current_library_factory(
        rows=[
            {
                "paper_uuid": PAPER_UUID,
                "record_origin": "pdf",
                "document_expectation": "required",
                "title": "Delete safely",
                "topic_folder": "Topic_A",
                "uncertainty": "USER_CONFIRMED: field=topic_folder; value=Topic_A",
            }
        ],
        documents=[
            {
                "document_id": f"{PAPER_UUID}:main",
                "paper_uuid": PAPER_UUID,
                "document_type": "main",
                "filename": "paper.pdf",
                "relative_path": relative,
                "extension": ".pdf",
                "file_status": "filed",
            }
        ],
        files={relative: b"paper bytes"} if with_file else None,
    )
    topic = root / "Topics" / "Topic_A"
    topic.mkdir(parents=True, exist_ok=True)
    (topic / "summary.md").write_text("do not inspect or remove", encoding="utf-8")
    return root


def test_delete_dry_run_keeps_catalogue_and_file(current_library_factory):
    root = _library(current_library_factory)
    before = (root / "catalogue.xlsx").read_bytes()

    result = PaperDeleteWorkflow(Settings.from_root(root)).run(
        paper_uuid=PAPER_UUID,
        dry_run=True,
    )

    assert result.status.value == "success"
    assert result.counts["managed_files"] == 1
    assert (root / "catalogue.xlsx").read_bytes() == before
    assert (root / "Topics" / "Topic_A" / "paper.pdf").is_file()
    assert not (root / ".library_state" / "trash" / "index.jsonl").exists()


def test_delete_apply_moves_complete_entity_to_trash(current_library_factory):
    root = _library(current_library_factory)

    result = PaperDeleteWorkflow(Settings.from_root(root)).run(
        paper_uuid=PAPER_UUID,
        dry_run=False,
    )

    assert result.status.value in {"success", "needs_review"}
    deletion_id = result.details["deletion_id"]
    trash = root / ".library_state" / "trash" / deletion_id
    manifest = json.loads((trash / "manifest.json").read_text(encoding="utf-8"))
    assert manifest["status"] == "committed"
    assert manifest["paper_uuid"] == PAPER_UUID
    assert (trash / "catalogue_record.json").is_file()
    assert (trash / "document_records.json").is_file()
    assert b"paper bytes" in (trash / "files" / "0001" / "paper.pdf").read_bytes()
    assert not (root / "Topics" / "Topic_A" / "paper.pdf").exists()
    assert (root / "Topics" / "Topic_A" / "summary.md").is_file()
    workbook = load_workbook(root / "catalogue.xlsx", read_only=True)
    assert workbook["Catalogue"].max_row == 1
    assert workbook["Documents"].max_row == 1
    tombstone = root / ".library_state" / "trash" / "tombstones" / f"{deletion_id}.json"
    assert tombstone.is_file()


def test_delete_allows_document_already_missing(current_library_factory):
    root = _library(current_library_factory, with_file=False)

    result = PaperDeleteWorkflow(Settings.from_root(root)).run(
        paper_uuid=PAPER_UUID,
        dry_run=False,
    )

    assert result.counts["missing_files"] == 1
    manifest = json.loads(
        (
            root
            / ".library_state"
            / "trash"
            / result.details["deletion_id"]
            / "manifest.json"
        ).read_text(encoding="utf-8")
    )
    assert manifest["files"][0]["missing_at_deletion"] is True


def test_delete_rolls_files_back_when_catalogue_commit_fails(
    current_library_factory,
    monkeypatch,
):
    root = _library(current_library_factory)

    def fail_save(_self):
        raise CatalogueError("simulated commit failure")

    monkeypatch.setattr(CatalogueService, "save_atomic", fail_save)
    try:
        PaperDeleteWorkflow(Settings.from_root(root)).run(
            paper_uuid=PAPER_UUID,
            dry_run=False,
        )
    except CatalogueError:
        pass
    else:  # pragma: no cover
        raise AssertionError("expected simulated Catalogue failure")

    assert b"paper bytes" in (root / "Topics" / "Topic_A" / "paper.pdf").read_bytes()
    workbook = load_workbook(root / "catalogue.xlsx", read_only=True)
    assert workbook["Catalogue"].max_row == 2
    trash = root / ".library_state" / "trash"
    assert not [path for path in trash.iterdir() if path.is_dir()] if trash.exists() else True


def test_agent_caller_cannot_apply_delete(current_library_factory, capsys):
    root = _library(current_library_factory)

    code = main(
        [
            "--root",
            str(root),
            "--json",
            "--caller",
            "agent",
            "delete",
            "--paper-uuid",
            PAPER_UUID,
            "--apply",
        ]
    )
    payload = json.loads(capsys.readouterr().out)

    assert code == 10
    assert payload["errors"][0]["type"] == "ConfigurationError"
    assert (root / "Topics" / "Topic_A" / "paper.pdf").is_file()


def test_trash_list_and_recovery_preserve_entity_ids(current_library_factory):
    root = _library(current_library_factory)
    settings = Settings.from_root(root)
    deleted = PaperDeleteWorkflow(settings).run(
        paper_uuid=PAPER_UUID,
        dry_run=False,
    )
    deletion_id = deleted.details["deletion_id"]

    listed = RecoveryWorkflow(settings).run(dry_run=True, list_trash=True)
    preview = RecoveryWorkflow(settings).run(
        dry_run=True,
        trash_id=deletion_id,
    )
    restored = RecoveryWorkflow(settings).run(
        dry_run=False,
        trash_id=deletion_id,
    )

    assert listed.counts["trash_entries"] == 1
    assert preview.completed[0]["action"] == "would_restore_trash_entity"
    assert restored.status.value in {"success", "needs_review"}
    service = CatalogueService(root / "catalogue.xlsx")
    rows = service.load()
    assert [row.get("paper_uuid") for row in rows] == [PAPER_UUID]
    assert rows[0].get("topic_folder") == "Topic_A"
    assert service.documents[0].get("document_id") == f"{PAPER_UUID}:main"
    assert (root / "Topics" / "Topic_A" / "paper.pdf").is_file()
    manifest = json.loads(
        (
            root / ".library_state" / "trash" / deletion_id / "manifest.json"
        ).read_text(encoding="utf-8")
    )
    assert manifest["status"] == "recovered"


def test_cleanup_purges_only_aged_trash_payload_and_keeps_tombstone(
    current_library_factory,
):
    root = _library(current_library_factory)
    settings = Settings.from_root(root)
    deleted = PaperDeleteWorkflow(settings).run(paper_uuid=PAPER_UUID, dry_run=False)
    deletion_id = deleted.details["deletion_id"]
    trash = root / ".library_state" / "trash" / deletion_id
    manifest_path = trash / "manifest.json"
    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    manifest["deleted_at"] = "2000-01-01T00:00:00+00:00"
    manifest_path.write_text(json.dumps(manifest), encoding="utf-8")

    preview = CleanupWorkflow(settings).run(
        dry_run=True,
        purge_trash=True,
        older_than="30d",
    )
    applied = CleanupWorkflow(settings).run(
        dry_run=False,
        purge_trash=True,
        older_than="30d",
    )

    assert any(item["kind"] == "trash_entry" for item in preview.completed)
    assert applied.status.value == "success"
    assert not trash.exists()
    assert (
        root / ".library_state" / "trash" / "tombstones" / f"{deletion_id}.json"
    ).is_file()
