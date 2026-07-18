from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


LEGACY_CATALOGUE_FIELDS = (
    "id",
    "title",
    "authors",
    "doi",
    "pmid",
    "manual_tags",
    "topic_folder",
    "pdf_status",
    "pdf_filename",
    "pdf_relative_path",
    "date_updated",
    "notes",
    "uncertainty",
    "custom_column",
    "year",
    "journal",
    "journal_abbrev",
    "publication_type",
)


def create_legacy_library(
    root: Path,
    rows: list[dict[str, object]],
    files: dict[str, bytes] | None = None,
) -> Path:
    """Create an isolated legacy workbook for migration/recovery tests only."""
    root.mkdir()
    for name in ("Inbox", "Registered", "Topics"):
        (root / name).mkdir()
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Catalogue"
    sheet.append(LEGACY_CATALOGUE_FIELDS)
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
    for row in rows:
        sheet.append([row.get(header) for header in LEGACY_CATALOGUE_FIELDS])
    notes = workbook.create_sheet("Other sheet")
    notes["A1"] = "preserve me"
    workbook.save(root / "catalogue.xlsx")
    workbook.close()
    for relative, marker in (files or {}).items():
        path = root / relative
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_bytes(b"%PDF-1.4\n" + marker + b"\n%%EOF\n")
    return root

