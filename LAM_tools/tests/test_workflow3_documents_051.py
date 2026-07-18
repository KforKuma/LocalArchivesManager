from __future__ import annotations

import hashlib
from pathlib import Path

from openpyxl import load_workbook

from lam.config import Settings
from lam.models import WorkflowStatus
from lam.workflows.daily_check import DailyCheckWorkflow
from lam.workflows.inbox_register import InboxRegisterWorkflow


UNKNOWN_UUID = "00000000-0000-4000-8000-000000000001"


class ForbiddenMetadataService:
    def __init__(self) -> None:
        self.calls = 0

    def lookup(self, *args, **kwargs):
        self.calls += 1
        raise AssertionError("Supplementary and exact local registration must not query metadata")


def _sheet_rows(path: Path, sheet_name: str) -> list[dict[str, object]]:
    workbook = load_workbook(path, data_only=False)
    sheet = workbook[sheet_name]
    headers = [
        str(cell.value).strip() if cell.value is not None else "" for cell in sheet[1]
    ]
    rows = [
        {
            header: sheet.cell(row=row_number, column=column).value
            for column, header in enumerate(headers, start=1)
            if header
        }
        for row_number in range(2, sheet.max_row + 1)
        if any(
            sheet.cell(row=row_number, column=column).value not in (None, "")
            for column in range(1, sheet.max_column + 1)
        )
    ]
    workbook.close()
    return rows


def _current_library(library_factory, *, with_main: bool) -> tuple[Path, str]:
    paper_uuid = "00000000-0000-4000-8000-000000000101"
    row = {
        "paper_uuid": paper_uuid,
        "record_origin": "pdf",
        "document_expectation": "required",
        "pmid": "1",
        "title": "Coordinated Paper",
        "authors": "Example Author",
        "year": "2025",
        "journal": "Test Journal",
        "journal_abbrev": "Test J",
        "topic_folder": "Topic_A",
        "publication_type": "",
    }
    marker = b"existing-main"
    files = {"Registered/main.pdf": marker} if with_main else None
    documents = (
        [
            {
                "document_id": f"{paper_uuid}:main",
                "paper_uuid": paper_uuid,
                "document_type": "main",
                "filename": "main.pdf",
                "relative_path": "Registered/main.pdf",
                "extension": ".pdf",
                "sha256": hashlib.sha256(
                    b"%PDF-1.4\n" + marker + b"\n%%EOF\n"
                ).hexdigest(),
                "file_status": "registered",
                "source": "fixture",
            }
        ]
        if with_main
        else []
    )
    root = library_factory([row], documents, files)
    DailyCheckWorkflow(Settings.from_root(root)).run(dry_run=False)
    return root, paper_uuid


def _run(root: Path, metadata_service: ForbiddenMetadataService, *, dry_run=False):
    settings = Settings.from_root(root)
    settings.ensure_runtime_directories()
    return InboxRegisterWorkflow(
        settings,
        metadata_service=metadata_service,
    ).run(
        dry_run=dry_run,
        filename_only=True,
        ocr_mode="never",
    )


def test_uuid_supplementary_pdf_and_csv_register_without_metadata_lookup(
    current_library_factory,
):
    root, paper_uuid = _current_library(current_library_factory, with_main=True)
    pdf_source = root / "Inbox" / f"{paper_uuid}__supp.pdf"
    csv_source = root / "Inbox" / f"{paper_uuid}__table01.csv"
    pdf_source.write_bytes(b"supplementary-pdf")
    csv_source.write_bytes(b"supplementary-table")
    metadata = ForbiddenMetadataService()

    result = _run(root, metadata)

    assert metadata.calls == 0
    assert result.status == WorkflowStatus.SUCCESS
    assert result.changed_files == 2
    assert result.counts["supplementary_registered"] == 2
    assert not pdf_source.exists()
    assert not csv_source.exists()
    documents = _sheet_rows(root / "catalogue.xlsx", "Documents")
    by_id = {row["document_id"]: row for row in documents}
    generic_id = f"{paper_uuid}:supp:generic:01"
    table_id = f"{paper_uuid}:supp:table:01"
    assert set(by_id) == {f"{paper_uuid}:main", generic_id, table_id}
    assert by_id[generic_id]["document_type"] == "supplementary"
    assert by_id[generic_id]["supplementary_type"] == "Supplementary"
    assert by_id[generic_id]["filename"] == (
        "Test J, 2025 - Coordinated Paper - Supplementary.pdf"
    )
    assert by_id[table_id]["supplementary_type"] == "Table"
    assert by_id[table_id]["sequence"] == 1
    assert by_id[table_id]["filename"] == (
        "Test J, 2025 - Coordinated Paper - Table01.csv"
    )
    for document_id in (generic_id, table_id):
        relative = str(by_id[document_id]["relative_path"])
        assert relative.startswith("Registered/")
        assert (root / relative).is_file()


def test_unknown_uuid_supplementary_is_blocked_and_left_in_inbox(current_library_factory):
    root, paper_uuid = _current_library(current_library_factory, with_main=True)
    source = root / "Inbox" / f"{UNKNOWN_UUID}__supp.pdf"
    source.write_bytes(b"unknown-parent")
    metadata = ForbiddenMetadataService()

    result = _run(root, metadata)

    assert metadata.calls == 0
    assert result.status == WorkflowStatus.NEEDS_REVIEW
    assert source.is_file()
    assert any(
        item.get("file") == f"Inbox/{source.name}"
        and item.get("issue") == "supplementary_uuid_not_found"
        for item in result.needs_review
    )
    documents = _sheet_rows(root / "catalogue.xlsx", "Documents")
    assert {row["document_id"] for row in documents} == {f"{paper_uuid}:main"}


def test_main_and_same_stem_supplementaries_register_together_with_one_final_check(
    current_library_factory,
    monkeypatch,
):
    root, paper_uuid = _current_library(current_library_factory, with_main=False)
    stem = "Test J, 2025 - Coordinated Paper"
    main_source = root / "Inbox" / f"{stem}.pdf"
    table_source = root / "Inbox" / f"{stem}_table1.csv"
    figure_source = root / "Inbox" / f"{stem}_figure02.pdf"
    main_source.write_bytes(b"new-main")
    table_source.write_bytes(b"table-one")
    figure_source.write_bytes(b"figure-two")
    metadata = ForbiddenMetadataService()

    from lam.workflows import progressive_register

    calls: list[bool] = []
    original = progressive_register.DailyCheckWorkflow.run

    def counted(self, *, dry_run=False, final_check=False):
        calls.append(final_check)
        return original(self, dry_run=dry_run, final_check=final_check)

    monkeypatch.setattr(progressive_register.DailyCheckWorkflow, "run", counted)

    result = _run(root, metadata)

    assert metadata.calls == 0
    assert calls == [True]
    assert result.status == WorkflowStatus.SUCCESS
    assert result.changed_files == 3
    assert not main_source.exists()
    assert not table_source.exists()
    assert not figure_source.exists()
    documents = _sheet_rows(root / "catalogue.xlsx", "Documents")
    by_type = {(row["document_type"], row["supplementary_type"]): row for row in documents}
    assert set(by_type) == {
        ("main", None),
        ("supplementary", "Table"),
        ("supplementary", "Figure"),
    }
    assert by_type[("main", None)]["document_id"] == f"{paper_uuid}:main"
    assert by_type[("supplementary", "Table")]["sequence"] == 1
    assert by_type[("supplementary", "Figure")]["sequence"] == 2
    assert all((root / str(row["relative_path"])).is_file() for row in documents)


def test_exact_sha_main_duplicate_is_blocked_before_metadata_lookup(current_library_factory):
    root, paper_uuid = _current_library(current_library_factory, with_main=True)
    existing = root / "Registered" / "main.pdf"
    source = root / "Inbox" / "duplicate.pdf"
    source.write_bytes(existing.read_bytes())
    metadata = ForbiddenMetadataService()

    result = _run(root, metadata)

    assert metadata.calls == 0
    assert result.status == WorkflowStatus.NEEDS_REVIEW
    assert source.is_file()
    assert any(
        item.get("file") == "Inbox/duplicate.pdf"
        and item.get("issue") == "duplicate_file_exact"
        and item.get("existing_document_id") == f"{paper_uuid}:main"
        for item in result.needs_review
    )
    assert len(_sheet_rows(root / "catalogue.xlsx", "Documents")) == 1


def test_workflow3_documents_dry_run_changes_no_catalogue_files_or_state(
    current_library_factory,
):
    root, paper_uuid = _current_library(current_library_factory, with_main=True)
    source = root / "Inbox" / f"{paper_uuid}__table01.csv"
    source.write_bytes(b"dry-run-table")
    metadata = ForbiddenMetadataService()
    catalogue_path = root / "catalogue.xlsx"
    catalogue_before = catalogue_path.read_bytes()
    documents_before = _sheet_rows(catalogue_path, "Documents")
    backups_before = {path.name for path in root.glob("catalogue.backup.*.xlsx")}
    journals_before = {
        path.relative_to(root).as_posix()
        for path in (root / ".library_state" / "runs").glob(
            "*/operation_journal.json"
        )
    }
    state_before = {
        name: (root / ".library_state" / name).read_bytes()
        for name in (
            "catalogue_snapshot.json",
            "file_manifest.json",
            "last_diff.json",
            "snapshot_commit.json",
        )
    }

    result = _run(root, metadata, dry_run=True)

    assert metadata.calls == 0
    assert result.status == WorkflowStatus.SUCCESS
    assert result.changed_files == 0
    assert source.is_file()
    assert catalogue_path.read_bytes() == catalogue_before
    assert _sheet_rows(catalogue_path, "Documents") == documents_before
    assert {path.name for path in root.glob("catalogue.backup.*.xlsx")} == backups_before
    assert {
        path.relative_to(root).as_posix()
        for path in (root / ".library_state" / "runs").glob(
            "*/operation_journal.json"
        )
    } == journals_before
    for name, content in state_before.items():
        assert (root / ".library_state" / name).read_bytes() == content
    target = root / "Registered" / "Test J, 2025 - Coordinated Paper - Table01.csv"
    assert not target.exists()
    assert any(
        item.get("action") == "would_register_supplementary"
        for item in result.completed
    )
