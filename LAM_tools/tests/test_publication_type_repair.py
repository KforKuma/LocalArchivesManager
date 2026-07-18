from __future__ import annotations

import json

from openpyxl import load_workbook

from lam.config import Settings
from lam.providers.unavailable import UnavailableMetadataService
from lam.workflows.inbox_register import InboxRegisterWorkflow
from lam.workflows.publication_type_repair import PublicationTypeRepairWorkflow

from conftest import write_text_pdf


def _row(old_name: str, *, publication_type: str) -> dict[str, object]:
    return {
        "id": "P1",
        "title": "Mucosal Profiling of Pediatric-Onset Colitis and IBD Reveals Common Pathogenics and Therapeutic Implications",
        "year": "2019",
        "journal": "Cell",
        "journal_abbrev": "Cell",
        "publication_type": publication_type,
        "pdf_status": "registered",
        "pdf_filename": old_name,
        "pdf_relative_path": f"Registered/{old_name}",
    }


def test_repair_command_renames_registered_pdf_and_normalizes_catalogue(library_factory):
    old_name = (
        "Cell, 2019, Journal Article; Research Support, Non-U.S. Gov't; journal-article - "
        "Mucosal Profiling of Pediatric-Onset Colitis and IBD Reveals Common Pathogenics and Therapeutic.pdf"
    )
    root = library_factory([_row(old_name, publication_type="Journal Article; Research Support, Non-U.S. Gov't; journal-article")])
    write_text_pdf(root / "Registered" / old_name, ["paper"])
    result = PublicationTypeRepairWorkflow(Settings.from_root(root)).run()
    expected = (
        "Cell, 2019 - Mucosal Profiling of Pediatric-Onset Colitis and IBD Reveals Common "
        "Pathogenics and Therapeutic Implications.pdf"
    )
    assert result.changed_files == 1
    assert (root / "Registered" / expected).is_file()
    assert not (root / "Registered" / old_name).exists()
    workbook = load_workbook(root / "catalogue.xlsx")
    sheet = workbook["Catalogue"]
    headers = {cell.value: cell.column for cell in sheet[1]}
    assert sheet.cell(2, headers["publication_type"]).value is None
    documents = workbook["Documents"]
    document_headers = {cell.value: cell.column for cell in documents[1]}
    assert documents.cell(2, document_headers["filename"]).value == expected
    assert documents.cell(2, document_headers["relative_path"]).value == (
        f"Registered/{expected}"
    )
    assert workbook["Other sheet"]["A1"].value == "preserve me"
    assert list(root.glob("catalogue.backup.*.xlsx"))
    journal = next((root / ".library_state" / "runs").glob("*-publication-types/operation_journal.json"))
    assert json.loads(journal.read_text(encoding="utf-8"))["status"] == "final_check_committed"


def test_repair_collision_never_overwrites_different_target(library_factory):
    old_name = "Test J, 2025, Journal Article; journal-article - Collision Paper.pdf"
    target_name = "Test J, 2025 - Collision Paper.pdf"
    row = {
        "id": "P1",
        "title": "Collision Paper",
        "year": "2025",
        "journal_abbrev": "Test J",
        "publication_type": "Journal Article; journal-article",
        "pdf_status": "registered",
        "pdf_filename": old_name,
        "pdf_relative_path": f"Registered/{old_name}",
    }
    root = library_factory([row], {f"Registered/{old_name}": b"old", f"Registered/{target_name}": b"different"})
    result = PublicationTypeRepairWorkflow(Settings.from_root(root)).run()
    assert (root / "Registered" / old_name).is_file()
    assert (root / "Registered" / target_name).read_bytes().endswith(b"different\n%%EOF\n")
    assert any(item["issue"] == "publication_type_repair_collision" for item in result.needs_review)


def test_repair_dry_run_changes_no_catalogue_files_or_state(library_factory):
    old_name = "Test J, 2025, Journal Article; journal-article - Dry Run Paper.pdf"
    row = {
        "id": "P1",
        "title": "Dry Run Paper",
        "year": "2025",
        "journal_abbrev": "Test J",
        "publication_type": "Journal Article; journal-article",
        "pdf_status": "registered",
        "pdf_filename": old_name,
        "pdf_relative_path": f"Registered/{old_name}",
    }
    root = library_factory([row], {f"Registered/{old_name}": b"paper"})
    before = (root / "catalogue.xlsx").read_bytes()
    result = PublicationTypeRepairWorkflow(Settings.from_root(root)).run(dry_run=True)
    assert result.changed_files == 1
    assert (root / "catalogue.xlsx").read_bytes() == before
    assert (root / "Registered" / old_name).is_file()
    assert not list(root.glob("catalogue.backup.*.xlsx"))
    assert not (root / ".library_state" / "catalogue_snapshot.json").exists()
    assert not (root / ".library_state" / "runs").exists()
    detail = result.details["repairs"][0]
    assert detail["old_publication_type"] == "Journal Article; journal-article"
    assert detail["new_publication_type"] is None
    assert detail["new_filename"] == "Test J, 2025 - Dry Run Paper.pdf"


def test_workflow3_new_registration_uses_canonical_type(library_factory):
    root = library_factory(
        [
            {
                "id": "P1",
                "title": "A Canonical Review",
                "year": "2025",
                "journal_abbrev": "Test J",
                "publication_type": "Journal Article; Review; Research Support; journal-article",
            }
        ]
    )
    source_name = "Test J, 2025, Review - A Canonical Review.pdf"
    write_text_pdf(root / "Inbox" / source_name, ["A Canonical Review\n2025"])
    result = InboxRegisterWorkflow(
        Settings.from_root(root), UnavailableMetadataService()
    ).run()
    assert result.changed_files == 1
    assert (root / "Registered" / "Test J, 2025, Review - A Canonical Review.pdf").is_file()
    workbook = load_workbook(root / "catalogue.xlsx")
    sheet = workbook["Catalogue"]
    headers = {cell.value: cell.column for cell in sheet[1]}
    assert sheet.cell(2, headers["publication_type"]).value == "Review"
