from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook, load_workbook

from lam.config import Settings
from lam.models import WorkflowStatus
from lam.schema import CATALOGUE_051_FIELDS, DOCUMENT_FIELDS
from lam.workflows.catalogue_filing import CatalogueFilingWorkflow


PAPER_UUID = "12345678-1234-4234-9234-123456789abc"


def _create_dual_library(
    root: Path,
    *,
    topic: str = "Topic_A",
    document_paths: tuple[tuple[str, str, str, str], ...],
) -> Path:
    root.mkdir()
    for name in ("Inbox", "Registered", "Topics"):
        (root / name).mkdir()
    catalogue_headers = list(CATALOGUE_051_FIELDS)
    for legacy in (
        "id",
        "record_uid",
        "pdf_status",
        "pdf_filename",
        "pdf_relative_path",
    ):
        if legacy not in catalogue_headers:
            catalogue_headers.append(legacy)

    workbook = Workbook()
    catalogue = workbook.active
    catalogue.title = "Catalogue"
    catalogue.append(catalogue_headers)
    main_path = document_paths[0][1]
    catalogue_values = {
        "paper_uuid": PAPER_UUID,
        "title": "Grouped paper",
        "topic_folder": topic,
        "source": "pubmed",
        "uncertainty": "USER_CONFIRMED: field=topic_folder; value=" + topic,
        "date_added": "2026-01-01",
        "date_updated": "2026-01-01",
        "id": "PMID:1",
        "record_uid": PAPER_UUID,
        "pdf_status": "registered",
        "pdf_filename": Path(main_path).name,
        "pdf_relative_path": main_path,
    }
    catalogue.append([catalogue_values.get(header) for header in catalogue_headers])

    documents = workbook.create_sheet("Documents")
    documents.append(list(DOCUMENT_FIELDS))
    for document_id, relative, document_type, supplementary_type in document_paths:
        values = {
            "document_id": document_id,
            "paper_uuid": PAPER_UUID,
            "document_type": document_type,
            "supplementary_type": supplementary_type,
            "sequence": (
                document_id.rsplit(":", 1)[-1] if document_type == "supplementary" else None
            ),
            "filename": Path(relative).name,
            "relative_path": relative,
            "extension": Path(relative).suffix,
            "sha256": "",
            "file_status": "registered",
            "source": "local_pdf",
            "uncertainty": "",
            "date_added": "2026-01-01",
            "date_updated": "2026-01-01",
        }
        documents.append([values.get(header) for header in DOCUMENT_FIELDS])
    workbook.save(root / "catalogue.xlsx")
    workbook.close()

    for _document_id, relative, _document_type, _supplementary_type in document_paths:
        path = root / relative
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_bytes(f"content:{path.name}".encode("utf-8"))
    return root


def _rows(path: Path, sheet_name: str) -> list[dict[str, object]]:
    workbook = load_workbook(path)
    sheet = workbook[sheet_name]
    headers = [cell.value for cell in sheet[1]]
    rows = [
        {
            str(header): sheet.cell(row=row_number, column=column).value
            for column, header in enumerate(headers, start=1)
            if header
        }
        for row_number in range(2, sheet.max_row + 1)
    ]
    workbook.close()
    return rows


def test_dual_workflow4_moves_every_document_and_journals_document_ids(
    tmp_path: Path, monkeypatch
):
    documents = (
        (f"{PAPER_UUID}:main", "Registered/main.pdf", "main", ""),
        (f"{PAPER_UUID}:supp:table:01", "Registered/table01.csv", "supplementary", "Table"),
        (f"{PAPER_UUID}:supp:data:02", "Registered/data02.xls", "supplementary", "Data"),
        (f"{PAPER_UUID}:supp:other:03", "Registered/other03.xlsx", "supplementary", "Other"),
    )
    root = _create_dual_library(tmp_path / "library", document_paths=documents)
    from lam.workflows import catalogue_filing

    final_check_calls = 0
    original = catalogue_filing.DailyCheckWorkflow.run

    def counted(self, *args, **kwargs):
        nonlocal final_check_calls
        final_check_calls += 1
        return original(self, *args, **kwargs)

    monkeypatch.setattr(catalogue_filing.DailyCheckWorkflow, "run", counted)

    result = CatalogueFilingWorkflow(Settings.from_root(root)).run()

    assert result.status == WorkflowStatus.SUCCESS
    assert result.changed_files == 4
    assert final_check_calls == 1
    assert len(list(root.glob("catalogue.backup.*.xlsx"))) == 1
    for _document_id, relative, _document_type, _supplementary_type in documents:
        filename = Path(relative).name
        assert not (root / relative).exists()
        assert (root / "Topics" / "Topic_A" / filename).is_file()

    document_rows = _rows(root / "catalogue.xlsx", "Documents")
    assert {row["document_id"] for row in document_rows} == {
        item[0] for item in documents
    }
    assert all(row["file_status"] == "filed" for row in document_rows)
    assert {
        row["relative_path"] for row in document_rows
    } == {
        f"Topics/Topic_A/{Path(item[1]).name}" for item in documents
    }

    # The compatibility fields remain read-only; Documents is authoritative.
    catalogue_row = _rows(root / "catalogue.xlsx", "Catalogue")[0]
    assert catalogue_row["pdf_status"] == "registered"
    assert catalogue_row["pdf_relative_path"] == "Registered/main.pdf"

    journal_path = next(
        (root / ".library_state" / "runs").glob("*-filing/operation_journal.json")
    )
    journal = json.loads(journal_path.read_text(encoding="utf-8"))
    assert journal["status"] == "final_check_committed"
    assert {item["document_id"] for item in journal["operations"]} == {
        item[0] for item in documents
    }
    assert all(
        item["execution_state"] == "final_check_committed"
        and "catalogue_committed" in item["stages"]
        for item in journal["operations"]
    )


def test_dual_workflow4_collision_blocks_the_entire_paper_group(tmp_path: Path):
    documents = (
        (f"{PAPER_UUID}:main", "Registered/main.pdf", "main", ""),
        (f"{PAPER_UUID}:supp:table:01", "Registered/table01.csv", "supplementary", "Table"),
    )
    root = _create_dual_library(tmp_path / "library", document_paths=documents)
    collision = root / "Topics" / "Topic_A" / "table01.csv"
    collision.parent.mkdir(parents=True)
    collision.write_bytes(b"different target")

    result = CatalogueFilingWorkflow(Settings.from_root(root)).run()

    assert result.status == WorkflowStatus.NEEDS_REVIEW
    assert result.changed_files == 0
    assert (root / "Registered" / "main.pdf").is_file()
    assert (root / "Registered" / "table01.csv").is_file()
    assert collision.read_bytes() == b"different target"
    rows = _rows(root / "catalogue.xlsx", "Documents")
    assert {row["relative_path"] for row in rows} == {
        "Registered/main.pdf",
        "Registered/table01.csv",
    }
    table = next(row for row in rows if row["filename"] == "table01.csv")
    assert "document_target_collision" in str(table["uncertainty"])
    assert any(
        item.get("document_id") == f"{PAPER_UUID}:supp:table:01"
        and item.get("issue") == "target_collision"
        for item in result.needs_review
    )


def test_dual_workflow4_dry_run_plans_whole_group_without_changes(tmp_path: Path):
    documents = (
        (f"{PAPER_UUID}:main", "Registered/main.pdf", "main", ""),
        (f"{PAPER_UUID}:supp:table:01", "Registered/table01.csv", "supplementary", "Table"),
    )
    root = _create_dual_library(tmp_path / "library", document_paths=documents)
    workbook_before = (root / "catalogue.xlsx").read_bytes()

    result = CatalogueFilingWorkflow(Settings.from_root(root)).run(dry_run=True)

    assert result.status == WorkflowStatus.SUCCESS
    assert result.details["planned_operations"] == 2
    assert {item["document_id"] for item in result.completed} == {
        item[0] for item in documents
    }
    assert result.changed_files == 0
    assert (root / "catalogue.xlsx").read_bytes() == workbook_before
    assert (root / "Registered" / "main.pdf").is_file()
    assert (root / "Registered" / "table01.csv").is_file()
    assert not (root / "Topics" / "Topic_A").exists()
    assert not list(root.glob("catalogue.backup.*.xlsx"))
    assert not (root / ".library_state" / "runs").exists()
