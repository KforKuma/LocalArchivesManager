from __future__ import annotations

import hashlib
import uuid
from pathlib import Path

from openpyxl import load_workbook

from lam.config import Settings
from lam.models import WorkflowStatus
from lam.services.catalogue_preflight_service import CataloguePreflightService
from lam.workflows.document_migration import DocumentMigrationWorkflow


def _set_catalogue_value(path: Path, field_name: str, row: int, value: object) -> None:
    workbook = load_workbook(path)
    sheet = workbook["Catalogue"]
    headers = {
        str(cell.value).strip(): cell.column
        for cell in sheet[1]
        if cell.value is not None and str(cell.value).strip()
    }
    column = headers.get(field_name)
    if column is None:
        column = sheet.max_column + 1
        sheet.cell(row=1, column=column).value = field_name
    sheet.cell(row=row, column=column).value = value
    workbook.save(path)
    workbook.close()


def _sheet_rows(path: Path, sheet_name: str) -> list[dict[str, object]]:
    workbook = load_workbook(path, data_only=False)
    sheet = workbook[sheet_name]
    headers = [
        str(cell.value).strip() if cell.value is not None else ""
        for cell in sheet[1]
    ]
    rows = [
        {
            header: sheet.cell(row=row_number, column=column).value
            for column, header in enumerate(headers, start=1)
            if header
        }
        for row_number in range(2, sheet.max_row + 1)
        if any(
            sheet.cell(row=row_number, column=column).value not in (None, "")
            for column in range(1, sheet.max_column + 1)
        )
    ]
    workbook.close()
    return rows


def test_legacy_preflight_for_migration_leaves_no_probe(legacy_library_factory):
    root = legacy_library_factory([])
    path = root / "catalogue.xlsx"

    result = CataloguePreflightService(path).before_modification()

    assert result.schema_mode == "legacy"
    assert result.path == path
    assert result.token.size == path.stat().st_size
    assert result.token.mtime_ns == path.stat().st_mtime_ns
    assert not list(root.glob(".*.preflight-*.tmp.xlsx"))


def test_migrate_documents_dry_run_does_not_change_workbook(legacy_library_factory):
    root = legacy_library_factory(
        [
            {
                "id": "PMID:1",
                "title": "Example",
                "topic_folder": "Topic_A",
                "pdf_status": "registered",
                "pdf_filename": "paper.pdf",
                "pdf_relative_path": "Registered/paper.pdf",
            }
        ],
        {"Registered/paper.pdf": b"content"},
    )
    path = root / "catalogue.xlsx"
    before = path.read_bytes()

    result = DocumentMigrationWorkflow(Settings.from_root(root)).run(dry_run=True)

    assert result.status == WorkflowStatus.SUCCESS
    assert result.counts["main_documents"] == 1
    assert result.counts["paper_uuid_generated"] == 1
    generated_uuid = result.details["plan"][0]["document_id"].removesuffix(":main")
    assert uuid.UUID(generated_uuid).version == 4
    assert path.read_bytes() == before
    workbook = load_workbook(path, read_only=True)
    assert workbook.sheetnames == ["Catalogue", "Other sheet"]
    workbook.close()
    assert not list(root.glob("catalogue.backup.*.xlsx"))


def test_migrate_documents_apply_creates_dual_model_reuses_record_uid_and_hash(
    legacy_library_factory,
):
    record_uid = "12345678-1234-4234-9234-123456789abc"
    root = legacy_library_factory(
        [
            {
                "id": "PMID:2",
                "title": "Migrated paper",
                "topic_folder": "Topic_A",
                "pdf_status": "registered",
                "pdf_filename": "paper.pdf",
                "pdf_relative_path": "Registered/paper.pdf",
                "source": "pubmed",
            }
        ],
        {"Registered/paper.pdf": b"main-document"},
    )
    path = root / "catalogue.xlsx"
    _set_catalogue_value(path, "record_uid", 2, record_uid)
    _set_catalogue_value(path, "source", 2, "pubmed")
    expected_sha256 = hashlib.sha256(
        (root / "Registered" / "paper.pdf").read_bytes()
    ).hexdigest()

    result = DocumentMigrationWorkflow(Settings.from_root(root)).run(dry_run=False)

    assert result.counts["record_uid_reused"] == 1
    assert result.counts["paper_uuid_generated"] == 0
    catalogue_row = _sheet_rows(path, "Catalogue")[0]
    document = _sheet_rows(path, "Documents")[0]
    assert catalogue_row["paper_uuid"] == record_uid
    assert catalogue_row["record_uid"] == record_uid
    assert document == {
        "document_id": f"{record_uid}:main",
        "paper_uuid": record_uid,
        "document_type": "main",
        "supplementary_type": None,
        "sequence": None,
        "filename": "paper.pdf",
        "relative_path": "Registered/paper.pdf",
        "extension": ".pdf",
        "sha256": expected_sha256,
        "file_status": "registered",
        "source": "pubmed",
        "uncertainty": None,
        "date_added": document["date_added"],
        "date_updated": document["date_updated"],
    }
    assert document["date_added"]
    assert document["date_updated"]
    assert list(root.glob("catalogue.backup.*.xlsx"))


def test_migrate_documents_apply_is_idempotent(legacy_library_factory):
    root = legacy_library_factory(
        [
            {
                "id": "PMID:3",
                "title": "Idempotent paper",
                "topic_folder": "Topic_A",
                "pdf_status": "registered",
                "pdf_filename": "paper.pdf",
                "pdf_relative_path": "Registered/paper.pdf",
            }
        ],
        {"Registered/paper.pdf": b"same"},
    )
    workflow = DocumentMigrationWorkflow(Settings.from_root(root))
    first = workflow.run(dry_run=False)
    after_first = (root / "catalogue.xlsx").read_bytes()
    backups_after_first = list(root.glob("catalogue.backup.*.xlsx"))

    second = DocumentMigrationWorkflow(Settings.from_root(root)).run(dry_run=False)

    assert first.changed_rows == 1
    assert second.status == WorkflowStatus.NO_CHANGES
    assert second.skipped == [{"reason": "already_migrated"}]
    assert second.counts == {"catalogue_rows": 1, "documents": 1}
    assert (root / "catalogue.xlsx").read_bytes() == after_first
    assert list(root.glob("catalogue.backup.*.xlsx")) == backups_after_first
    assert len(_sheet_rows(root / "catalogue.xlsx", "Documents")) == 1


def test_migrate_documents_records_missing_file_issue_only_in_documents(
    legacy_library_factory,
):
    root = legacy_library_factory(
        [
            {
                "id": "PMID:4",
                "title": "Missing paper",
                "topic_folder": "Topic_A",
                "pdf_status": "registered",
                "pdf_filename": "missing.pdf",
                "pdf_relative_path": "Registered/missing.pdf",
                "uncertainty": "USER_CONFIRMED: field=title; value=Missing paper",
            }
        ]
    )

    result = DocumentMigrationWorkflow(Settings.from_root(root)).run(dry_run=False)

    document = _sheet_rows(root / "catalogue.xlsx", "Documents")[0]
    catalogue_row = _sheet_rows(root / "catalogue.xlsx", "Catalogue")[0]
    assert result.counts["missing_files"] == 1
    assert document["file_status"] == "missing"
    assert document["sha256"] in (None, "")
    assert "document_file_missing" in str(document["uncertainty"])
    assert catalogue_row["uncertainty"] == (
        "USER_CONFIRMED: field=title; value=Missing paper"
    )
