from __future__ import annotations

import inspect

from openpyxl import load_workbook
from pypdf import PdfWriter

from lam.config import Settings
from lam.models import (
    CandidateConfidence,
    DocumentAnalysisRequest,
    IdentifierCandidate,
    OcrInspection,
    OcrTextBlock,
    MetadataRecord,
    PdfInspection,
    PdfVisualType,
    TitleCandidate,
    WorkflowStatus,
)
from lam.providers.unavailable import UnavailableMetadataService
from lam.services.document_analysis_service import (
    DocumentAnalysisService,
    EasyOcrRegionBackend,
    NativePdfBackend,
)
from lam.services.ocr_service import OcrService
from lam.services.catalogue_service import CatalogueService
from lam.services.record_canonicalization_service import RegisteredRecordCanonicalizer
from lam.utils.candidate_cleaning import (
    classify_title_candidate,
    merge_title_blocks,
)
from lam.utils.identifiers import merge_doi_fragments, parse_doi_candidate
from lam.workflows.inbox_register import InboxRegisterWorkflow
import lam.workflows.progressive_register as progressive_register


def _block(text: str, y: float, *, left: float = 20, width: float = 500):
    return OcrTextBlock(
        [[left, y], [left + width, y], [left + width, y + 24], [left, y + 24]],
        text,
        0.95,
    )


def test_document_analysis_backends_share_one_contract(library_factory):
    root = library_factory([])
    source = root / "Inbox" / "analysis.pdf"
    source.write_bytes(b"placeholder")
    settings = Settings.from_root(root)
    request = DocumentAnalysisRequest(
        file_path=source,
        pdf_visual_type=PdfVisualType.NATIVE_TEXT,
        requested_fields={"title", "doi", "year"},
        native_text="A Reliable Native PDF Research Title\nPublished 2024",
        metadata_title="A Reliable Title - Anna's Archive",
    )
    native = NativePdfBackend(settings).analyze(request)
    assert native.backend == "native"
    assert any(item.query_eligible for item in native.title_candidates)
    contaminated = next(
        item
        for item in native.title_candidates
        if "metadata_title_contaminated" in item.rejection_reasons
    )
    assert contaminated.confidence == CandidateConfidence.REJECTED

    class FixedOcr:
        def inspect_first_page(self, *args, **kwargs):
            return OcrInspection(
                status="success",
                title_candidates=[
                    TitleCandidate(
                        "A Reliable OCR Research Title",
                        "high",
                        "ocr_title_merged",
                        1,
                    )
                ],
                year_candidates=["2024"],
            )

    easy = EasyOcrRegionBackend(settings, FixedOcr()).analyze(request)
    assert easy.backend == "easyocr"
    assert easy.status == "success"
    assert easy.title_candidates[0].field == "title"
    assert easy.raw_result.status == "success"


def test_uninstalled_advanced_backend_is_explainable(library_factory):
    root = library_factory([])
    request = DocumentAnalysisRequest(file_path=root / "Inbox" / "missing.pdf")
    result = DocumentAnalysisService(Settings.from_root(root)).analyze(
        request, backend_name="docling"
    )
    assert result.status == "backend_unavailable"
    assert result.backend == "docling"
    assert "document_analysis_backend_unavailable" in result.warnings


def test_workflow3_has_no_direct_easyocr_dependency():
    source = inspect.getsource(progressive_register)
    assert "import easyocr" not in source
    assert "OcrService(" not in source


def test_title_pollution_and_publisher_navigation_are_rejected():
    for value, source in (
        ("A room for living - Anna's Archive", "metadata"),
        ("ELSEVIER journal homepage: www.elsevier.com/locate/jep", "ocr"),
        ("Contents lists available at ScienceDirect", "ocr"),
    ):
        candidate = classify_title_candidate(value, source=source)
        assert candidate.confidence == CandidateConfidence.REJECTED
        assert candidate.rejection_reasons


def test_two_and_three_line_title_merge_and_author_stop():
    candidates, metrics = merge_title_blocks(
        [
            _block("ELSEVIER journal homepage", 10),
            _block("A room for living: Private and public aspects", 50),
            _block("in the experience", 80),
            _block("of the living room", 110),
            _block("Anna Smith", 150),
        ]
    )
    values = [item.value for item in candidates if item.query_eligible]
    assert (
        "A room for living: Private and public aspects in the experience "
        "of the living room"
    ) in values
    assert all("ELSEVIER" not in value for value in values)
    assert all("Anna Smith" not in value for value in values)
    assert metrics["title_lines_merged"] >= 2


def test_hyphenated_line_repairs_but_normal_line_keeps_space():
    repaired, metrics = merge_title_blocks(
        [_block("A study of experi-", 10), _block("ence in living rooms", 40)]
    )
    assert any("experience in living rooms" in item.value for item in repaired)
    assert metrics["hyphenation_repaired"] >= 1
    normal, _ = merge_title_blocks(
        [_block("A study of private", 10), _block("and public rooms", 40)]
    )
    assert any("private and public" in item.value for item in normal)


def test_doi_parser_rejects_prefix_short_and_long_candidates():
    assert parse_doi_candidate("https://doi.org/10.1016/j.jenvp.2008.05.001").complete
    prefix = parse_doi_candidate("10.1016/j")
    assert prefix.status == "prefix_only"
    assert prefix.rejection_reasons == ("doi_prefix_only",)
    short = parse_doi_candidate("10.1000/a")
    assert short.status == "prefix_only"
    long = parse_doi_candidate("10.1000/" + "a" * 220)
    assert long.status == "rejected"
    assert "doi_length_rejected" in long.rejection_reasons


def test_adjacent_doi_and_footer_url_fragments_merge():
    doi = merge_doi_fragments(["10.1016/j.jenvp.", "2008.05.001"])
    assert any(item.value == "10.1016/j.jenvp.2008.05.001" for item in doi)
    footer = merge_doi_fragments(
        ["https://doi.org/10.1016/", "j.jenvp.2008.05.001"]
    )
    assert any(item.value == "10.1016/j.jenvp.2008.05.001" for item in footer)


def test_limited_ocr_doi_separator_correction():
    corrected = OcrService._correct_ocr_doi_text(
        "https://doi.org/10.1016/j-jenvp.2008.05.001"
    )
    assert "10.1016/j.jenvp.2008.05.001" in corrected
    assert parse_doi_candidate(corrected).complete


def test_doi_prefix_is_only_auxiliary_support_for_provider_record():
    inspection = PdfInspection(
        relative_path="Inbox/a.pdf",
        filename="a.pdf",
        size=1,
        mtime_ns=1,
        title_candidates=[
            TitleCandidate(
                "A room for living Private and public aspects",
                "high",
                "ocr_title_merged",
            )
        ],
        year_candidates=["2008"],
        ocr_result=OcrInspection(doi_prefix_only=["10.1016/j"]),
    )
    metadata = MetadataRecord(
        canonical_id="DOI:10.1016/j.jenvp.2008.05.001",
        title="A room for living Private and public aspects",
        year="2008",
        doi="10.1016/j.jenvp.2008.05.001",
    )
    assert InboxRegisterWorkflow._metadata_supports_inspection(metadata, inspection)
    assert not inspection.doi_candidates


def test_corrected_doi_accepts_adjacent_publication_year():
    inspection = PdfInspection(
        relative_path="Inbox/a.pdf",
        filename="a.pdf",
        size=1,
        mtime_ns=1,
        title_candidates=[
            TitleCandidate(
                "A room for living Private and public aspects",
                "high",
                "filename",
            )
        ],
        year_candidates=["2008"],
        doi_candidates=[
            IdentifierCandidate(
                "10.1016/j.jenvp.2008.05.001",
                source_type="ocr_corrected",
            )
        ],
    )
    metadata = MetadataRecord(
        title="A room for living Private and public aspects",
        year="2009",
        doi="10.1016/j.jenvp.2008.05.001",
    )

    assert InboxRegisterWorkflow._metadata_supports_inspection(metadata, inspection)


def test_confirmed_provider_replaces_historical_prefix_only_doi(library_factory):
    root = library_factory(
        [
            {
                "title": "A room for living Private and public aspects",
                "year": "2008",
                "doi": "10.1016/j",
                "source": "local_pdf",
                "uncertainty": (
                    "NEEDS_REVIEW: field=paper_identity; "
                    "issue_key=metadata_identity_unconfirmed"
                ),
            }
        ]
    )
    catalogue = CatalogueService(root / "catalogue.xlsx")
    record = catalogue.load()[0]
    provider = MetadataRecord(
        canonical_id="DOI:10.1016/j.jenvp.2008.05.001",
        title="A room for living Private and public aspects",
        year="2008",
        doi="10.1016/j.jenvp.2008.05.001",
        source=["crossref"],
    )
    result = RegisteredRecordCanonicalizer().canonicalize(
        catalogue, record, provider, merged=provider
    )
    assert not result.conflicts
    assert record.get("doi") == "10.1016/j.jenvp.2008.05.001"


def test_polluted_metadata_is_not_queried_or_conflicted(library_factory):
    root = library_factory([])
    source = root / "Inbox" / "scan.pdf"
    writer = PdfWriter()
    writer.add_blank_page(width=612, height=792)
    writer.add_metadata({"/Title": "A paper - Anna's Archive"})
    with source.open("wb") as handle:
        writer.write(handle)
    result = InboxRegisterWorkflow(
        Settings.from_root(root), UnavailableMetadataService()
    ).run(ocr_mode="never")
    item = result.details["files"][0]
    assert result.details["metadata_lookup_requests"] == 0
    assert "pdf_text_ocr_conflict" not in item["issue_keys"]
    assert any(
        "metadata_title_contaminated" in candidate["rejection_reasons"]
        for candidate in item["rejected_title_candidates"]
    )


def test_provisional_inbox_document_prevents_unmatched_final_check(library_factory):
    root = library_factory([])
    source = root / "Inbox" / "unknown.pdf"
    writer = PdfWriter()
    writer.add_blank_page(width=612, height=792)
    with source.open("wb") as handle:
        writer.write(handle)
    result = InboxRegisterWorkflow(
        Settings.from_root(root), UnavailableMetadataService()
    ).run(ocr_mode="never")
    assert result.status == WorkflowStatus.NEEDS_REVIEW
    assert not any(
        item.get("issue") == "unmatched_local_document"
        for item in result.needs_review
    )
    workbook = load_workbook(root / "catalogue.xlsx", read_only=True)
    documents = workbook["Documents"]
    headers = {cell.value: cell.column for cell in documents[1]}
    assert documents.max_row == 2
    assert documents.cell(2, headers["relative_path"]).value == "Inbox/unknown.pdf"
    assert documents.cell(2, headers["file_status"]).value == "inbox"
