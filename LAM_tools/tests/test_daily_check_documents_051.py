from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook

from lam.config import Settings
from lam.models import WorkflowStatus
from lam.schema import CATALOGUE_051_FIELDS, DOCUMENT_FIELDS
from lam.workflows.daily_check import DailyCheckWorkflow


PAPER_UUID = "11111111-1111-4111-8111-111111111111"
CATALOGUE_HEADERS = (
    *CATALOGUE_051_FIELDS,
    "id",
    "record_uid",
    "pdf_status",
    "pdf_filename",
    "pdf_relative_path",
)


def _library(
    tmp_path: Path,
    *,
    documents: list[dict[str, object]],
    files: dict[str, bytes] | None = None,
    catalogue_uncertainty: str = "USER_CONFIRMED: field=title; value=Example",
) -> Path:
    root = tmp_path / "library"
    root.mkdir()
    for directory in ("Inbox", "Registered", "Topics"):
        (root / directory).mkdir()

    workbook = Workbook()
    catalogue = workbook.active
    catalogue.title = "Catalogue"
    catalogue.append(CATALOGUE_HEADERS)
    catalogue_values = {
        "paper_uuid": PAPER_UUID,
        "id": "PMID:1",
        "title": "Example",
        "topic_folder": "Topic_A",
        "pdf_status": "not_downloaded",
        "pdf_filename": "legacy.pdf",
        "pdf_relative_path": "Registered/legacy.pdf",
        "date_updated": "2000-01-01",
        "uncertainty": catalogue_uncertainty,
    }
    catalogue.append([catalogue_values.get(field) for field in CATALOGUE_HEADERS])
    sheet = workbook.create_sheet("Documents")
    sheet.append(DOCUMENT_FIELDS)
    for document in documents:
        sheet.append([document.get(field) for field in DOCUMENT_FIELDS])
    workbook.save(root / "catalogue.xlsx")

    for relative, content in (files or {}).items():
        path = root / relative
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_bytes(content)
    return root


def _document_rows(root: Path) -> dict[str, dict[str, object]]:
    sheet = load_workbook(root / "catalogue.xlsx")["Documents"]
    headers = {
        str(sheet.cell(1, column).value): column
        for column in range(1, sheet.max_column + 1)
    }
    return {
        str(sheet.cell(row, headers["document_id"]).value): {
            field: sheet.cell(row, column).value for field, column in headers.items()
        }
        for row in range(2, sheet.max_row + 1)
    }


def _catalogue_row(root: Path) -> dict[str, object]:
    sheet = load_workbook(root / "catalogue.xlsx")["Catalogue"]
    headers = {
        str(sheet.cell(1, column).value): column
        for column in range(1, sheet.max_column + 1)
    }
    return {
        field: sheet.cell(2, column).value for field, column in headers.items()
    }


def _run(root: Path):
    settings = Settings.from_root(root)
    settings.ensure_runtime_directories()
    return DailyCheckWorkflow(settings).run()


def test_daily_check_matches_each_supported_document_and_only_updates_documents(
    tmp_path,
):
    documents = [
        {
            "document_id": f"{PAPER_UUID}:main",
            "paper_uuid": PAPER_UUID,
            "document_type": "main",
            "filename": "paper.pdf",
            "relative_path": "Registered/paper.pdf",
            "extension": ".wrong",
            "file_status": "missing",
        },
        {
            "document_id": f"{PAPER_UUID}:supp:table:01",
            "paper_uuid": PAPER_UUID,
            "document_type": "supplementary",
            "supplementary_type": "Table",
            "sequence": "01",
            "filename": "table.csv",
            "relative_path": "Registered/table.csv",
            "extension": ".wrong",
            "file_status": "missing",
        },
        {
            "document_id": f"{PAPER_UUID}:supp:data:01",
            "paper_uuid": PAPER_UUID,
            "document_type": "supplementary",
            "supplementary_type": "Data",
            "sequence": "01",
            "filename": "data.xlsx",
            "relative_path": "Topics/Topic_A/data.xlsx",
            "extension": ".wrong",
            "file_status": "registered",
        },
        {
            "document_id": f"{PAPER_UUID}:supp:methods:01",
            "paper_uuid": PAPER_UUID,
            "document_type": "supplementary",
            "supplementary_type": "Methods",
            "sequence": "01",
            "filename": "methods.xls",
            "relative_path": "Inbox/methods.xls",
            "extension": ".wrong",
            "file_status": "missing",
        },
    ]
    root = _library(
        tmp_path,
        documents=documents,
        files={
            "Registered/paper.pdf": b"pdf",
            "Registered/table.csv": b"csv",
            "Topics/Topic_A/data.xlsx": b"xlsx",
            "Inbox/methods.xls": b"xls",
        },
    )

    result = _run(root)

    assert result.status == WorkflowStatus.SUCCESS
    assert result.counts["managed_documents"] == 4
    assert result.changed_rows == 4
    assert result.needs_review == []
    assert result.completed
    assert {item.get("sheet") for item in result.completed} == {"Documents"}
    rows = _document_rows(root)
    assert rows[f"{PAPER_UUID}:main"]["file_status"] == "registered"
    assert rows[f"{PAPER_UUID}:main"]["extension"] == ".pdf"
    assert rows[f"{PAPER_UUID}:supp:table:01"]["file_status"] == "registered"
    assert rows[f"{PAPER_UUID}:supp:table:01"]["extension"] == ".csv"
    assert rows[f"{PAPER_UUID}:supp:data:01"]["file_status"] == "filed"
    assert rows[f"{PAPER_UUID}:supp:data:01"]["extension"] == ".xlsx"
    assert rows[f"{PAPER_UUID}:supp:methods:01"]["file_status"] == "inbox"
    assert rows[f"{PAPER_UUID}:supp:methods:01"]["extension"] == ".xls"

    catalogue = _catalogue_row(root)
    assert catalogue["pdf_status"] == "not_downloaded"
    assert catalogue["pdf_filename"] == "legacy.pdf"
    assert catalogue["pdf_relative_path"] == "Registered/legacy.pdf"
    assert catalogue["date_updated"] == "2000-01-01"
    assert catalogue["uncertainty"] == "USER_CONFIRMED: field=title; value=Example"


def test_missing_document_updates_document_uncertainty_not_catalogue(tmp_path):
    root = _library(
        tmp_path,
        documents=[
            {
                "document_id": f"{PAPER_UUID}:main",
                "paper_uuid": PAPER_UUID,
                "document_type": "main",
                "filename": "missing.pdf",
                "relative_path": "Registered/missing.pdf",
                "extension": ".pdf",
                "file_status": "registered",
                "uncertainty": "user file note",
            }
        ],
        catalogue_uncertainty="USER_CONFIRMED: field=title; value=Keep",
    )

    result = _run(root)

    assert result.status == WorkflowStatus.NEEDS_REVIEW
    assert result.needs_review == [
        {
            "sheet": "Documents",
            "row": 2,
            "document_id": f"{PAPER_UUID}:main",
            "issue": "document_file_missing",
        }
    ]
    row = _document_rows(root)[f"{PAPER_UUID}:main"]
    assert row["file_status"] == "missing"
    assert row["relative_path"] == "Registered/missing.pdf"
    assert row["uncertainty"].splitlines() == [
        "user file note",
        "document_file_missing",
    ]
    assert _catalogue_row(root)["uncertainty"] == (
        "USER_CONFIRMED: field=title; value=Keep"
    )


def test_unmatched_supported_file_is_reported_without_polluting_catalogue(tmp_path):
    root = _library(
        tmp_path,
        documents=[],
        files={"Registered/orphan.csv": b"orphan"},
        catalogue_uncertainty="manual catalogue note",
    )

    result = _run(root)

    assert result.status == WorkflowStatus.NEEDS_REVIEW
    assert result.needs_review == [
        {
            "file": "Registered/orphan.csv",
            "issue": "unmatched_local_document",
        }
    ]
    assert result.changed_rows == 0
    assert _catalogue_row(root)["uncertainty"] == "manual catalogue note"
