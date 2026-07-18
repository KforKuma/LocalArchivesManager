from __future__ import annotations

import json
import uuid
from pathlib import Path

from openpyxl import load_workbook

from lam.cli import build_parser, main
from lam.command_registry import COMMANDS
from lam.models import WorkflowResult
from lam.schema import CATALOGUE_FIELDS, DOCUMENT_FIELDS


PUBLIC_COMMANDS = {
    "init",
    "check",
    "register",
    "search",
    "file",
    "delete",
    "export",
    "review",
    "status",
    "recover",
    "migrate",
    "cleanup",
    "doctor",
    "commands",
}


def _payload(capsys):
    return json.loads(capsys.readouterr().out.strip())


def test_public_help_and_registry_expose_only_055_commands(capsys):
    parser = build_parser()
    action = next(item for item in parser._actions if item.dest == "command")
    assert set(action.choices) == PUBLIC_COMMANDS
    assert {item.name for item in COMMANDS} == PUBLIC_COMMANDS
    help_text = parser.format_help()
    for old in (
        "normalize-records",
        "migrate-documents",
        "migrate-identifiers",
        "migrate-topics",
        "repair-publication-types",
    ):
        assert old not in help_text


def test_init_apply_creates_exact_empty_library(tmp_path, capsys):
    root = tmp_path / "new-library"
    code = main(["--root", str(root), "--json", "init", "--apply"])
    payload = _payload(capsys)
    assert code == 0
    assert payload["command"] == "init"
    assert payload["details"]["details"]["uses_network"] is False
    assert payload["details"]["details"]["uses_ocr"] is False
    for name in ("Inbox", "Registered", "Topics", ".library_state"):
        assert (root / name).is_dir()
    assert (root / "library_changes.md").is_file()
    env_example = (root / ".env.example").read_text(encoding="utf-8")
    assert "NCBI_EMAIL=" in env_example
    assert "NCBI_API_KEY=" in env_example
    workbook = load_workbook(root / "catalogue.xlsx", read_only=True)
    assert tuple(cell.value for cell in workbook["Catalogue"][1]) == CATALOGUE_FIELDS
    assert tuple(cell.value for cell in workbook["Documents"][1]) == DOCUMENT_FIELDS


def test_init_refuses_existing_library(current_library_factory, capsys):
    root = current_library_factory()
    code = main(["--root", str(root), "--json", "init", "--apply"])
    payload = _payload(capsys)
    assert code == 10
    assert payload["errors"]


def test_status_commands_and_alias_have_canonical_identity(
    current_library_factory, capsys
):
    root = current_library_factory()
    assert main(["--root", str(root), "--json", "commands"]) == 0
    alias = _payload(capsys)
    assert alias["command"] == "commands"
    assert alias["canonical_command"] == "status commands"
    assert alias["report_path"] is None

    assert main(["--root", str(root), "--json", "status", "commands"]) == 0
    canonical = _payload(capsys)
    assert canonical["command"] == "status commands"
    assert canonical["canonical_command"] == "status commands"
    assert canonical["report_path"] is None
    lines = []
    for path in (root / ".library_state" / "invocations").glob("*.jsonl"):
        lines.extend(path.read_text(encoding="utf-8").splitlines())
    invocations = [json.loads(line) for line in lines]
    assert sum(item["command"] == "commands" for item in invocations) == 1
    assert sum(item["command"] == "status commands" for item in invocations) == 1


def test_doctor_alias_has_one_invocation_and_environment_canonical_identity(
    current_library_factory, capsys, monkeypatch
):
    root = current_library_factory()

    def safe_doctor(self, *, initialize_ocr_models=False):
        result = WorkflowResult("doctor")
        result.details = {
            "uses_network": initialize_ocr_models,
            "may_download_models": initialize_ocr_models,
        }
        return result

    monkeypatch.setattr("lam.cli.DoctorWorkflow.run", safe_doctor)
    assert main(["--root", str(root), "--json", "doctor"]) == 0
    payload = _payload(capsys)
    assert payload["command"] == "doctor"
    assert payload["canonical_command"] == "status environment"
    lines = []
    for path in (root / ".library_state" / "invocations").glob("*.jsonl"):
        lines.extend(path.read_text(encoding="utf-8").splitlines())
    invocations = [json.loads(line) for line in lines]
    assert sum(item["command"] == "doctor" for item in invocations) == 1


def test_status_library_and_config_work_before_initialization(tmp_path, capsys):
    root = tmp_path / "uninitialized"
    code = main(["--root", str(root), "--json", "status", "library"])
    payload = _payload(capsys)
    assert code == 2
    assert payload["details"]["details"]["initialized"] is False
    code = main(["--root", str(root), "--json", "status", "config"])
    payload = _payload(capsys)
    assert code in {0, 3}
    assert payload["details"]["details"]["secrets_exposed"] is False


def test_migrate_identifiers_current_schema_is_no_changes(
    current_library_factory, capsys
):
    root = current_library_factory()
    code = main(
        ["--root", str(root), "--json", "migrate", "identifiers", "--dry-run"]
    )
    payload = _payload(capsys)
    assert code == 3
    assert payload["command"] == "migrate identifiers"
    assert payload["details"]["details"]["schema_detection"]["classification"] == "current"


def test_review_preserves_user_fields_and_clears_resolved_document_blocker(
    current_library_factory, capsys
):
    paper_uuid = str(uuid.uuid4())
    document_id = str(uuid.uuid4())
    root = current_library_factory(
        rows=[
            {
                "paper_uuid": paper_uuid,
                "title": "Keep title",
                "manual_tags": "keep-tag",
                "topic_folder": "Keep topic",
                "notes": "keep note",
                "uncertainty": "USER_CONFIRMED: field=title; value=Keep title",
            }
        ],
        documents=[
            {
                "document_id": document_id,
                "paper_uuid": paper_uuid,
                "document_type": "main",
                "filename": "paper.pdf",
                "relative_path": "Registered/paper.pdf",
                "extension": ".pdf",
                "file_status": "missing",
                "uncertainty": "document_file_missing\nuser free text",
            }
        ],
        files={"Registered/paper.pdf": b"same"},
    )
    code = main(
        [
            "--root",
            str(root),
            "--json",
            "review",
            "--document-id",
            document_id,
            "--apply",
        ]
    )
    _payload(capsys)
    assert code in {0, 2}
    workbook = load_workbook(root / "catalogue.xlsx", data_only=False)
    catalogue_headers = {cell.value: cell.column for cell in workbook["Catalogue"][1]}
    assert workbook["Catalogue"].cell(2, catalogue_headers["manual_tags"]).value == "keep-tag"
    assert workbook["Catalogue"].cell(2, catalogue_headers["topic_folder"]).value == "Keep topic"
    assert workbook["Catalogue"].cell(2, catalogue_headers["notes"]).value == "keep note"
    assert "USER_CONFIRMED" in workbook["Catalogue"].cell(2, catalogue_headers["uncertainty"]).value
    document_headers = {cell.value: cell.column for cell in workbook["Documents"][1]}
    uncertainty = workbook["Documents"].cell(2, document_headers["uncertainty"]).value
    assert uncertainty == "user free text"


def test_recover_provider_policy_is_forwarded_only_to_inbox(
    current_library_factory, capsys, monkeypatch
):
    root = current_library_factory()
    captured = {}
    journal = root / ".library_state" / "runs" / "inbox-interrupted" / "operation_journal.json"
    journal.parent.mkdir(parents=True)
    journal.write_text(
        json.dumps(
            {
                "run_id": "inbox-interrupted",
                "workflow": "inbox_register",
                "status": "planned",
                "operations": [{"operation_type": "move", "execution_state": "planned"}],
            }
        ),
        encoding="utf-8",
    )

    def fake_run(self, **kwargs):
        captured.update(kwargs)
        return WorkflowResult("inbox_register")

    monkeypatch.setattr("lam.workflows.recovery.InboxRegisterWorkflow.run", fake_run)
    code = main(
        [
            "--root",
            str(root),
            "--json",
            "recover",
            "--scope",
            "inbox",
            "--offline",
            "--no-cache-write",
            "--dry-run",
        ]
    )
    payload = _payload(capsys)
    assert code in {0, 3}
    assert captured["offline"] is True
    assert captured["cache_write"] is False
    assert payload["details"]["details"]["filed_documents_re_registered"] is False


def test_review_clears_stale_inbox_blocker_state(
    current_library_factory, capsys
):
    paper_uuid = str(uuid.uuid4())
    root = current_library_factory(rows=[{"paper_uuid": paper_uuid, "title": "Paper"}])
    blocker = root / ".library_state" / "inbox_blockers.json"
    blocker.parent.mkdir(parents=True)
    blocker.write_text(
        json.dumps(
            {
                "files": [
                    {
                        "paper_uuid": paper_uuid,
                        "source_path": "Inbox/deleted-duplicate.pdf",
                        "issue_keys": ["duplicate_inbox_file"],
                    }
                ]
            }
        ),
        encoding="utf-8",
    )
    code = main(
        ["--root", str(root), "--json", "review", "--all", "--apply"]
    )
    payload = _payload(capsys)
    assert code in {0, 2}
    assert not blocker.exists()
    assert payload["details"]["details"]["resolved_blockers"]


def test_recover_registered_orphan_from_unique_journal(
    current_library_factory, capsys
):
    paper_uuid = str(uuid.uuid4())
    document_id = f"{paper_uuid}:main"
    root = current_library_factory(
        rows=[{"paper_uuid": paper_uuid, "title": "Recovered paper"}],
        files={"Registered/recovered.pdf": b"recovered"},
    )
    journal = root / ".library_state" / "runs" / "interrupted" / "operation_journal.json"
    journal.parent.mkdir(parents=True)
    journal.write_text(
        json.dumps(
            {
                "run_id": "interrupted",
                "workflow": "inbox_register",
                "status": "file_moved",
                "operations": [
                    {
                        "operation_type": "move",
                        "paper_uuid": paper_uuid,
                        "document_id": document_id,
                        "source": "Inbox/original.pdf",
                        "target": "Registered/recovered.pdf",
                        "execution_state": "file_moved",
                        "planned_updates": {"document_type": "main"},
                    }
                ],
            }
        ),
        encoding="utf-8",
    )
    code = main(
        [
            "--root",
            str(root),
            "--json",
            "recover",
            "--scope",
            "registered",
            "--apply",
        ]
    )
    payload = _payload(capsys)
    assert code in {0, 2}
    workbook = load_workbook(root / "catalogue.xlsx", data_only=False)
    headers = {cell.value: cell.column for cell in workbook["Documents"][1]}
    assert workbook["Documents"].max_row == 2
    assert workbook["Documents"].cell(2, headers["paper_uuid"]).value == paper_uuid
    assert workbook["Documents"].cell(2, headers["relative_path"]).value == "Registered/recovered.pdf"
    assert payload["details"]["details"]["filed_documents_re_registered"] is False


def test_recover_publication_repair_is_detection_gated(
    current_library_factory, capsys, monkeypatch
):
    root = current_library_factory(rows=[{"publication_type": "Review"}])

    def forbidden(*args, **kwargs):
        raise AssertionError("repair must not run without a detected anomaly")

    monkeypatch.setattr("lam.workflows.recovery.PublicationTypeRepairWorkflow.run", forbidden)
    code = main(
        [
            "--root",
            str(root),
            "--json",
            "recover",
            "--scope",
            "publication-types",
            "--dry-run",
        ]
    )
    payload = _payload(capsys)
    assert code == 3
    assert payload["details"]["details"]["scope_results"][0]["status"] == "no_changes"


def test_unknown_schema_is_refused_by_public_migration(
    current_library_factory, capsys
):
    root = current_library_factory(rows=[{"future_field": "future"}])
    code = main(
        ["--root", str(root), "--json", "migrate", "identifiers", "--dry-run"]
    )
    payload = _payload(capsys)
    assert code == 20
    assert "unknown or future" in payload["errors"][0]["message"]
