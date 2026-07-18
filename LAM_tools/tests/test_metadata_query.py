from __future__ import annotations

import json
import socket
import uuid

import httpx
from openpyxl import load_workbook

from lam.config import Settings
from lam.models import (
    MetadataLookupRequest,
    MetadataLookupResult,
    MetadataLookupStatus,
    MetadataRecord,
    DownloadCandidate,
    ProviderResult,
    ProviderStatus,
    WorkflowStatus,
)
from lam.workflows.metadata_query import MetadataQueryWorkflow
from lam.services.download_service import DownloadService
from conftest import write_text_pdf


class FakeMetadataService:
    def __init__(self, results):
        self.results = list(results)
        self.requests = []

    def lookup(self, request):
        self.requests.append(request)
        return self.results.pop(0)


def found(record, confidence="exact_identifier"):
    provider = ProviderResult(
        "pubmed", ProviderStatus.FOUND, "pmid", record.pmid, records=[record]
    )
    return MetadataLookupResult(
        MetadataLookupStatus.FOUND,
        records=[record.to_dict()],
        best_record=record.to_dict(),
        confidence=confidence,
        providers_used=["pubmed"],
        provider_results=[provider],
        selection_reason="exact test evidence",
    )


def test_exact_pmid_adds_new_row_and_runs_one_final_check(library_factory):
    root = library_factory([])
    record = MetadataRecord(
        canonical_id="PMID:12345678",
        title="New Biomedical Paper",
        authors=["Alice Smith"],
        year="2025",
        journal="Biomedical Journal",
        journal_abbrev="Biomed J",
        doi="10.1000/new",
        pmid="12345678",
        publication_type=None,
        raw_publication_types=["Journal Article"],
        source=["pubmed"],
    )
    service = FakeMetadataService([found(record)])
    settings = Settings.from_root(root)
    settings.ensure_runtime_directories()
    result = MetadataQueryWorkflow(settings, service).run(
        MetadataLookupRequest(pmid="12345678")
    )
    assert result.status == WorkflowStatus.SUCCESS
    workbook = load_workbook(root / "catalogue.xlsx")
    sheet = workbook["Catalogue"]
    headers = {cell.value: cell.column for cell in sheet[1]}
    assert sheet.max_row == 2
    assert uuid.UUID(str(sheet.cell(2, headers["paper_uuid"]).value)).version == 4
    assert str(sheet.cell(2, headers["pmid"]).value) == "12345678"
    assert "id" not in headers
    assert "pdf_status" not in headers
    assert workbook["Documents"].max_row == 1
    assert result.details["final_check"]["status"] in {"success", "no_changes"}
    journal = next((root / ".library_state" / "runs").glob("*-metadata/operation_journal.json"))
    assert json.loads(journal.read_text(encoding="utf-8"))["status"] == "final_check_committed"


def test_existing_blank_fields_are_completed_without_overwriting_user_fields(library_factory):
    root = library_factory(
        [
            {
                "id": "P1",
                "title": "Existing Paper",
                "authors": "",
                "pmid": "12345678",
                "manual_tags": "keep tag",
                "topic_folder": "Topic_A",
                "notes": "keep note",
            }
        ]
    )
    record = MetadataRecord(
        canonical_id="PMID:12345678",
        title="Existing Paper",
        authors=["Alice Smith"],
        year="2025",
        journal="Biomedical Journal",
        pmid="12345678",
        source=["pubmed"],
    )
    settings = Settings.from_root(root)
    settings.ensure_runtime_directories()
    result = MetadataQueryWorkflow(settings, FakeMetadataService([found(record)])).run(
        MetadataLookupRequest(), catalogue_row=2
    )
    workbook = load_workbook(root / "catalogue.xlsx")
    sheet = workbook["Catalogue"]
    headers = {cell.value: cell.column for cell in sheet[1]}
    assert sheet.cell(2, headers["authors"]).value == "Alice Smith"
    assert sheet.cell(2, headers["manual_tags"]).value == "keep tag"
    assert sheet.cell(2, headers["topic_folder"]).value == "Topic_A"
    assert sheet.cell(2, headers["notes"]).value == "keep note"
    assert result.changed_rows == 1


def test_existing_bibliographic_conflict_is_preserved_and_reviewed(library_factory):
    root = library_factory(
        [{"id": "P1", "title": "Original Title", "pmid": "12345678"}]
    )
    record = MetadataRecord(
        canonical_id="PMID:12345678",
        title="Different Provider Title",
        pmid="12345678",
        source=["pubmed"],
    )
    settings = Settings.from_root(root)
    settings.ensure_runtime_directories()
    result = MetadataQueryWorkflow(settings, FakeMetadataService([found(record)])).run(
        MetadataLookupRequest(), catalogue_row=2
    )
    workbook = load_workbook(root / "catalogue.xlsx")
    sheet = workbook["Catalogue"]
    headers = {cell.value: cell.column for cell in sheet[1]}
    assert sheet.cell(2, headers["title"]).value == "Original Title"
    assert "catalogue_existing_value_conflict" in sheet.cell(
        2, headers["uncertainty"]
    ).value
    assert result.status == WorkflowStatus.NEEDS_REVIEW


def test_dry_run_plans_add_without_catalogue_backup_or_snapshot(library_factory):
    root = library_factory([])
    original = (root / "catalogue.xlsx").read_bytes()
    record = MetadataRecord(
        canonical_id="PMID:12345678",
        title="Dry Run Paper",
        authors=["Alice Smith"],
        year="2025",
        pmid="12345678",
        source=["pubmed"],
    )
    settings = Settings.from_root(root)
    settings.ensure_runtime_directories()
    result = MetadataQueryWorkflow(settings, FakeMetadataService([found(record)])).run(
        MetadataLookupRequest(pmid="12345678"), dry_run=True
    )
    assert result.completed[0]["action"] == "would_add"
    assert (root / "catalogue.xlsx").read_bytes() == original
    assert not list(root.glob("catalogue.backup.*.xlsx"))
    assert not (root / ".library_state" / "snapshot_commit.json").exists()
    assert not (root / ".library_state" / "runs").exists()


def test_not_found_creates_one_stable_metadata_blocker(library_factory):
    root = library_factory([{"id": "P1", "title": "Missing Paper"}])
    lookup = MetadataLookupResult(
        MetadataLookupStatus.NOT_FOUND,
        provider_results=[
            ProviderResult("pubmed", ProviderStatus.NOT_FOUND, "title", "Missing Paper")
        ],
        selection_reason="No result",
    )
    settings = Settings.from_root(root)
    settings.ensure_runtime_directories()
    MetadataQueryWorkflow(settings, FakeMetadataService([lookup])).run(
        MetadataLookupRequest(), catalogue_row=2
    )
    second = MetadataQueryWorkflow(settings, FakeMetadataService([lookup])).run(
        MetadataLookupRequest(), catalogue_row=2
    )
    workbook = load_workbook(root / "catalogue.xlsx")
    sheet = workbook["Catalogue"]
    headers = {cell.value: cell.column for cell in sheet[1]}
    value = sheet.cell(2, headers["uncertainty"]).value
    assert value.count("NEEDS_REVIEW:") == 1
    assert second.status == WorkflowStatus.NEEDS_REVIEW


def test_provider_unavailable_is_failure_not_not_found(library_factory):
    root = library_factory([])
    lookup = MetadataLookupResult(
        MetadataLookupStatus.UNAVAILABLE,
        provider_results=[
            ProviderResult("pubmed", ProviderStatus.UNAVAILABLE, "pmid", "12345678")
        ],
        selection_reason="network unavailable",
    )
    settings = Settings.from_root(root)
    settings.ensure_runtime_directories()
    result = MetadataQueryWorkflow(settings, FakeMetadataService([lookup])).run(
        MetadataLookupRequest(pmid="12345678"), dry_run=True
    )
    assert result.status == WorkflowStatus.FAILED
    assert result.details["network_failure"] is True


def test_empty_user_confirmation_resolves_not_found_blocker(library_factory):
    root = library_factory(
        [
            {
                "id": "P1",
                "title": "Known missing query",
                "uncertainty": "USER_CONFIRMED: field=metadata_query_not_found; value=",
            }
        ]
    )
    lookup = MetadataLookupResult(
        MetadataLookupStatus.NOT_FOUND,
        provider_results=[
            ProviderResult("pubmed", ProviderStatus.NOT_FOUND, "title", "Known missing query")
        ],
        selection_reason="No result",
    )
    settings = Settings.from_root(root)
    settings.ensure_runtime_directories()
    result = MetadataQueryWorkflow(settings, FakeMetadataService([lookup])).run(
        MetadataLookupRequest(), catalogue_row=2
    )
    workbook = load_workbook(root / "catalogue.xlsx")
    sheet = workbook["Catalogue"]
    headers = {cell.value: cell.column for cell in sheet[1]}
    uncertainty = sheet.cell(2, headers["uncertainty"]).value
    assert "USER_CONFIRMED:" in uncertainty
    assert "NEEDS_REVIEW:" not in uncertainty
    assert not result.needs_review


def test_user_confirmed_metadata_candidate_selects_ambiguous_result(library_factory):
    root = library_factory(
        [
            {
                "id": "P1",
                "title": "Confirmed Candidate Paper",
                "uncertainty": (
                    "USER_CONFIRMED: field=metadata_candidate; value=PMID:12345678"
                ),
            }
        ]
    )
    selected = MetadataRecord(
        canonical_id="PMID:12345678",
        title="Confirmed Candidate Paper",
        authors=["Alice Smith"],
        year="2025",
        pmid="12345678",
        source=["pubmed"],
    )
    other = MetadataRecord(
        canonical_id="PMID:23456789",
        title="Confirmed Candidate Paper",
        authors=["Bob Jones"],
        year="2024",
        pmid="23456789",
        source=["pubmed"],
    )
    lookup = MetadataLookupResult(
        MetadataLookupStatus.AMBIGUOUS,
        records=[selected.to_dict(), other.to_dict()],
        confidence="ambiguous",
        selection_reason="Two candidates",
    )
    settings = Settings.from_root(root)
    settings.ensure_runtime_directories()
    result = MetadataQueryWorkflow(settings, FakeMetadataService([lookup])).run(
        MetadataLookupRequest(), catalogue_row=2
    )
    workbook = load_workbook(root / "catalogue.xlsx")
    sheet = workbook["Catalogue"]
    headers = {cell.value: cell.column for cell in sheet[1]}
    assert sheet.cell(2, headers["authors"]).value == "Alice Smith"
    assert result.details["queries"][0]["selection_reason"].startswith("USER_CONFIRMED")


def test_normalized_doi_prevents_duplicate_row(library_factory):
    root = library_factory(
        [{"id": "P1", "title": "Existing DOI Paper", "doi": "https://doi.org/10.1000/ABC"}]
    )
    metadata = MetadataRecord(
        canonical_id="DOI:10.1000/abc",
        title="Existing DOI Paper",
        authors=["Alice Smith"],
        year="2025",
        doi="10.1000/abc",
        source=["pubmed"],
    )
    settings = Settings.from_root(root)
    settings.ensure_runtime_directories()
    MetadataQueryWorkflow(settings, FakeMetadataService([found(metadata)])).run(
        MetadataLookupRequest(doi="10.1000/abc")
    )
    workbook = load_workbook(root / "catalogue.xlsx")
    assert workbook["Catalogue"].max_row == 2
    sheet = workbook["Catalogue"]
    headers = {cell.value: cell.column for cell in sheet[1]}
    assert sheet.cell(2, headers["authors"]).value == "Alice Smith"


def test_explicit_download_commits_to_inbox_updates_catalogue_and_journal(
    library_factory, tmp_path
):
    root = library_factory([])
    fixture = tmp_path / "download.pdf"
    write_text_pdf(fixture, ["arXiv:2401.12345"])
    payload = fixture.read_bytes()
    calls = []

    def handler(request):
        calls.append(str(request.url))
        return httpx.Response(200, content=payload, headers={"content-type": "application/pdf"})

    settings = Settings.from_root(root)
    settings.ensure_runtime_directories()
    resolver = lambda *_args: [
        (socket.AF_INET, socket.SOCK_STREAM, 6, "", ("93.184.216.34", 443))
    ]
    downloader = DownloadService(
        settings,
        client=httpx.Client(transport=httpx.MockTransport(handler), follow_redirects=False),
        resolver=resolver,
    )
    record = MetadataRecord(
        canonical_id="ARXIV:2401.12345",
        title="Open Access Preprint",
        authors=["Alice Smith"],
        year="2024",
        arxiv_id="2401.12345",
        source=["arxiv"],
        download_candidates=[
            DownloadCandidate(
                provider="arxiv",
                source_url="https://arxiv.org/pdf/2401.12345",
                expected_arxiv_id="2401.12345",
                is_direct_pdf=True,
                priority=10,
            )
        ],
    )
    result = MetadataQueryWorkflow(
        settings, FakeMetadataService([found(record)]), downloader
    ).run(MetadataLookupRequest(arxiv_id="2401.12345"), download=True)
    target = root / "Inbox" / "download_arxiv_2401.12345.pdf"
    assert target.exists()
    assert not list((root / "Registered").glob("*.pdf"))
    workbook = load_workbook(root / "catalogue.xlsx")
    sheet = workbook["Documents"]
    headers = {cell.value: cell.column for cell in sheet[1]}
    assert sheet.cell(2, headers["file_status"]).value == "inbox"
    assert sheet.cell(2, headers["relative_path"]).value == (
        "Inbox/download_arxiv_2401.12345.pdf"
    )
    assert result.changed_files == 1
    assert result.details["downloads"][0]["status"] == "downloaded"
    assert calls == ["https://arxiv.org/pdf/2401.12345"]
    journal = next((root / ".library_state" / "runs").glob("*-download/operation_journal.json"))
    operation = json.loads(journal.read_text(encoding="utf-8"))["operations"][0]
    assert operation["stages"] == [
        "candidate_selected",
        "download_started",
        "temporary_file_written",
        "validation_passed",
        "committed_to_inbox",
        "catalogue_committed",
        "final_check_committed",
    ]


def test_download_dry_run_only_reports_plan(library_factory):
    root = library_factory([])
    settings = Settings.from_root(root)
    settings.ensure_runtime_directories()
    calls = []
    downloader = DownloadService(
        settings,
        client=httpx.Client(
            transport=httpx.MockTransport(
                lambda request: calls.append(request) or httpx.Response(500)
            )
        ),
        resolver=lambda *_args: (_ for _ in ()).throw(AssertionError("no DNS in dry run")),
    )
    record = MetadataRecord(
        canonical_id="DOI:10.1000/dry",
        title="Dry Download",
        authors=["Alice Smith"],
        year="2025",
        doi="10.1000/dry",
        source=["unpaywall"],
        download_candidates=[
            DownloadCandidate(
                provider="unpaywall",
                source_url="https://repository.example/dry.pdf?token=secret",
                expected_doi="10.1000/dry",
                is_direct_pdf=True,
            )
        ],
    )
    result = MetadataQueryWorkflow(
        settings, FakeMetadataService([found(record)]), downloader
    ).run(MetadataLookupRequest(doi="10.1000/dry"), download=True, dry_run=True)
    assert result.details["downloads"][0]["status"] == "planned"
    assert "token" not in result.details["downloads"][0]["selected_url"]
    assert calls == []
    assert not (root / ".library_state" / "tmp").exists()
    assert not (root / ".library_state" / "runs").exists()
    assert not list((root / "Inbox").glob("*.pdf"))
