from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook

from lam.config import Settings
from lam.schema import CATALOGUE_FIELDS, DOCUMENT_FIELDS
from lam.workflows.identifier_migration import IdentifierMigrationWorkflow


LEGACY_HEADERS = (
    "id",
    "record_uid",
    "paper_uuid",
    "title",
    "topic_folder",
    "pdf_status",
    "pdf_filename",
    "pdf_relative_path",
    "uncertainty",
    "source",
)


def _library(tmp_path: Path, row: dict[str, object] | None) -> Path:
    root = tmp_path / "library"
    for name in ("Inbox", "Registered", "Topics"):
        (root / name).mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    catalogue = workbook.active
    catalogue.title = "Catalogue"
    catalogue.append(LEGACY_HEADERS)
    if row is not None:
        catalogue.append([row.get(field) for field in LEGACY_HEADERS])
    other = workbook.create_sheet("User Notes")
    other["A1"] = "preserve"
    workbook.save(root / "catalogue.xlsx")
    return root


def test_identifier_migration_dry_run_does_not_change_workbook(tmp_path: Path):
    uid = "12345678-1234-4234-9234-123456789abc"
    root = _library(
        tmp_path,
        {
            "id": f"LOCAL:{uid}",
            "title": "Legacy paper",
            "topic_folder": "Topic_A",
            "uncertainty": "USER_CONFIRMED: field=title; value=Legacy paper",
        },
    )
    before = (root / "catalogue.xlsx").read_bytes()

    result = IdentifierMigrationWorkflow(Settings.from_root(root)).run(dry_run=True)

    assert result.status.value == "success"
    assert result.counts["local_id_recovered"] == 1
    assert (root / "catalogue.xlsx").read_bytes() == before
    assert not list(root.glob("catalogue.backup.*.xlsx"))


def test_identifier_migration_apply_reorders_and_removes_legacy_columns(tmp_path: Path):
    uid = "12345678-1234-4234-9234-123456789abc"
    root = _library(
        tmp_path,
        {
            "record_uid": uid,
            "title": "Legacy paper",
            "topic_folder": "Topic_A",
            "pdf_status": "registered",
            "pdf_filename": "legacy.pdf",
            "pdf_relative_path": "Registered/legacy.pdf",
            "uncertainty": "USER_CONFIRMED: field=title; value=Legacy paper",
            "source": "local_pdf",
        },
    )
    (root / "Registered" / "legacy.pdf").write_bytes(b"%PDF-1.4\nlegacy\n%%EOF\n")

    result = IdentifierMigrationWorkflow(Settings.from_root(root)).run(dry_run=False)

    assert result.status.value in {"success", "needs_review"}
    workbook = load_workbook(root / "catalogue.xlsx", read_only=True)
    catalogue = workbook["Catalogue"]
    documents = workbook["Documents"]
    assert tuple(cell.value for cell in catalogue[1]) == CATALOGUE_FIELDS
    assert tuple(cell.value for cell in documents[1]) == DOCUMENT_FIELDS
    values = {cell.value: catalogue.cell(2, index + 1).value for index, cell in enumerate(catalogue[1])}
    assert values["paper_uuid"] == uid
    assert values["uncertainty"].startswith("USER_CONFIRMED:")
    document = {cell.value: documents.cell(2, index + 1).value for index, cell in enumerate(documents[1])}
    assert document["paper_uuid"] == uid
    assert document["relative_path"] == "Registered/legacy.pdf"
    assert document["sha256"]
    assert workbook["User Notes"]["A1"].value == "preserve"
    assert list(root.glob("catalogue.backup.*.xlsx"))


def test_identifier_migration_blocks_disagreeing_row_keys(tmp_path: Path):
    root = _library(
        tmp_path,
        {
            "paper_uuid": "12345678-1234-4234-9234-123456789abc",
            "record_uid": "87654321-4321-4321-8321-cba987654321",
            "title": "Conflict",
            "topic_folder": "Topic_A",
        },
    )
    before = (root / "catalogue.xlsx").read_bytes()

    result = IdentifierMigrationWorkflow(Settings.from_root(root)).run(dry_run=False)

    assert result.status.value == "needs_review"
    assert result.needs_review[0]["issue_key"] == "identity_mismatch"
    assert (root / "catalogue.xlsx").read_bytes() == before
    assert not list(root.glob("catalogue.backup.*.xlsx"))


def test_empty_workbook_schema_migration_is_saved(tmp_path: Path):
    root = _library(tmp_path, None)
    result = IdentifierMigrationWorkflow(Settings.from_root(root)).run(dry_run=False)
    assert result.status.value in {"success", "no_changes"}
    workbook = load_workbook(root / "catalogue.xlsx", read_only=True)
    assert tuple(cell.value for cell in workbook["Catalogue"][1]) == CATALOGUE_FIELDS
    assert tuple(cell.value for cell in workbook["Documents"][1]) == DOCUMENT_FIELDS
    assert list(root.glob("catalogue.backup.*.xlsx"))
