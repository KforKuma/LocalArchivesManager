import json
from copy import deepcopy
from dataclasses import replace

import pytest

from openpyxl import load_workbook

from lam.config import Settings
from lam.exceptions import CatalogueError
from lam.models import (
    IdentifierCandidate,
    MetadataLookupResult,
    MetadataLookupStatus,
    MetadataRecord,
    OcrInspection,
    TitleCandidate,
    WorkflowStatus,
)
from lam.providers.unavailable import UnavailableMetadataService
from lam.services.catalogue_service import CatalogueService
from lam.services.file_service import FileService
from lam.workflows.inbox_register import InboxRegisterWorkflow

from conftest import write_text_pdf


class FixedOcrService:
    def __init__(self, results):
        self.results = results
        self.calls = []

    def inspect_first_page(self, path, **kwargs):
        self.calls.append(path.name)
        value = self.results[path.name] if isinstance(self.results, dict) else self.results
        result = deepcopy(value)
        result.trigger_reason = kwargs["trigger_reason"]
        return result


def with_ocr(settings):
    return replace(
        settings,
        ocr=replace(settings.ocr, enabled=True, gpu="false", min_text_chars=80),
    )


def ocr_success(*, title, doi="", year=""):
    return OcrInspection(
        status="success",
        title_candidates=[TitleCandidate(title, "high", "ocr_page_top", 1)],
        doi_candidates=(
            [IdentifierCandidate(doi, 1, doi, "high", "ocr")] if doi else []
        ),
        year_candidates=[year] if year else [],
        combined_text="\n".join(item for item in (title, doi, year) if item),
        gpu_mode="cpu",
        dpi=250,
    )


def test_filename_match_registers_updates_catalogue_and_final_checks(library_factory):
    root = library_factory(
        [
            {
                "id": "P1",
                "title": "A Registered Biomedical Paper",
                "year": "2025",
                "journal_abbrev": "Test J",
                "pdf_status": "inbox",
                "pdf_filename": "download.pdf",
                "pdf_relative_path": "Inbox/download.pdf",
            }
        ]
    )
    write_text_pdf(root / "Inbox" / "download.pdf", ["A Registered Biomedical Paper\nAuthors\n2025"])
    settings = Settings.from_root(root)
    settings.ensure_runtime_directories()
    result = InboxRegisterWorkflow(settings, UnavailableMetadataService()).run()
    target = root / "Registered" / "Test J, 2025 - A Registered Biomedical Paper.pdf"
    assert result.status == WorkflowStatus.SUCCESS
    assert target.exists()
    assert not (root / "Inbox" / "download.pdf").exists()
    workbook = load_workbook(root / "catalogue.xlsx")
    sheet = workbook["Catalogue"]
    headers = {cell.value: cell.column for cell in sheet[1]}
    assert sheet.cell(2, headers["pdf_status"]).value == "registered"
    assert sheet.cell(2, headers["pdf_relative_path"]).value == (
        "Registered/Test J, 2025 - A Registered Biomedical Paper.pdf"
    )
    assert result.details["manual_checkpoint_required"] is True
    assert result.details["final_check"]["status"] in {"success", "no_changes"}
    journal = next((root / ".library_state" / "runs").glob("*/operation_journal.json"))
    assert json.loads(journal.read_text(encoding="utf-8"))["status"] == "final_check_committed"
    second = InboxRegisterWorkflow(settings).run()
    assert second.status == WorkflowStatus.NO_CHANGES


def test_doi_match_registers_when_filename_is_unknown(library_factory):
    root = library_factory(
        [
            {
                "id": "P1",
                "title": "Identifier Matched Paper",
                "doi": "10.1000/matched.1",
                "year": "2024",
                "journal": "Identifier Journal",
            }
        ]
    )
    write_text_pdf(
        root / "Inbox" / "random.pdf",
        ["Identifier Matched Paper\nAuthors\ndoi:10.1000/matched.1\n2024"],
    )
    settings = Settings.from_root(root)
    settings.ensure_runtime_directories()
    result = InboxRegisterWorkflow(settings, UnavailableMetadataService()).run()
    assert result.status == WorkflowStatus.SUCCESS
    assert (root / "Registered" / "Identifier Journal, 2024 - Identifier Matched Paper.pdf").exists()
    assert result.details["files"][0]["match_method"] == "doi"


def test_unavailable_metadata_creates_provisional_catalogue_row(library_factory):
    root = library_factory([])
    write_text_pdf(root / "Inbox" / "unknown.pdf", ["Unknown Local Document\nNo identifiers"])
    settings = Settings.from_root(root)
    settings.ensure_runtime_directories()
    result = InboxRegisterWorkflow(settings, UnavailableMetadataService()).run()
    assert result.status == WorkflowStatus.NEEDS_REVIEW
    assert (root / "Inbox" / "unknown.pdf").exists()
    assert result.details["metadata_lookup_requests"] == 1
    assert result.details["files"][0]["result_status"] == "provisional"
    assert result.details["files"][0]["issue_keys"] == [
        "metadata_provider_unavailable",
        "metadata_identity_unconfirmed",
    ]
    workbook = load_workbook(root / "catalogue.xlsx")
    sheet = workbook["Catalogue"]
    headers = {cell.value: cell.column for cell in sheet[1]}
    assert sheet.max_row == 2
    assert str(sheet.cell(2, headers["id"]).value).startswith("LOCAL:")
    assert sheet.cell(2, headers["pdf_status"]).value == "inbox"
    assert "issue_key=metadata_identity_unconfirmed" in sheet.cell(
        2, headers["uncertainty"]
    ).value


def test_existing_exact_filename_record_skips_pdf_content_and_registers(library_factory):
    root = library_factory(
        [
            {
                "id": "P1",
                "title": "Blank PDF",
                "year": "2025",
                "journal": "Test Journal",
                "pdf_filename": "blank.pdf",
            }
        ]
    )
    write_text_pdf(root / "Inbox" / "blank.pdf", [""])
    settings = Settings.from_root(root)
    settings.ensure_runtime_directories()
    result = InboxRegisterWorkflow(settings).run()
    assert result.status == WorkflowStatus.SUCCESS
    assert not (root / "Inbox" / "blank.pdf").exists()
    assert (root / "Registered" / "Test Journal, 2025 - Blank PDF.pdf").exists()
    assert result.details["files"][0]["inspection_level_used"] == "skip"


def test_one_blocked_file_does_not_prevent_other_registration(library_factory):
    root = library_factory(
        [
            {
                "id": "P1",
                "title": "Ready Paper",
                "year": "2025",
                "journal": "Ready Journal",
                "pdf_filename": "ready.pdf",
            }
        ]
    )
    write_text_pdf(root / "Inbox" / "ready.pdf", ["Ready Paper\nAuthors\n2025"])
    write_text_pdf(root / "Inbox" / "unknown.pdf", ["Unknown Paper"])
    settings = Settings.from_root(root)
    settings.ensure_runtime_directories()
    result = InboxRegisterWorkflow(settings, UnavailableMetadataService()).run()
    assert result.status == WorkflowStatus.NEEDS_REVIEW
    assert (root / "Registered" / "Ready Journal, 2025 - Ready Paper.pdf").exists()
    assert (root / "Inbox" / "unknown.pdf").exists()
    assert result.changed_files == 1


def test_register_dry_run_writes_no_catalogue_snapshot_or_journal(library_factory):
    root = library_factory(
        [
            {
                "id": "P1",
                "title": "Dry Run Paper",
                "year": "2025",
                "journal": "Dry Journal",
                "pdf_filename": "dry.pdf",
            }
        ]
    )
    write_text_pdf(root / "Inbox" / "dry.pdf", ["Dry Run Paper\nAuthors"])
    original = (root / "catalogue.xlsx").read_bytes()
    settings = Settings.from_root(root)
    settings.ensure_runtime_directories()
    result = InboxRegisterWorkflow(settings).run(dry_run=True)
    assert result.completed[0]["action"] == "would_register"
    assert (root / "Inbox" / "dry.pdf").exists()
    assert (root / "catalogue.xlsx").read_bytes() == original
    assert not (root / ".library_state" / "snapshot_commit.json").exists()
    assert not (root / ".library_state" / "runs").exists()


def test_filename_only_parses_standard_filename_without_page_text(library_factory):
    root = library_factory(
        [
            {
                "id": "P1",
                "title": "Standard Filename Paper",
                "year": "2025",
                "journal_abbrev": "Std J",
            }
        ]
    )
    name = "Std J, 2025 - Standard Filename Paper.pdf"
    write_text_pdf(root / "Inbox" / name, [""])
    settings = Settings.from_root(root)
    settings.ensure_runtime_directories()
    result = InboxRegisterWorkflow(settings).run(dry_run=True, filename_only=True)
    assert result.completed[0]["action"] == "would_register"
    assert result.details["files"][0]["match_method"] == "standard_filename_title"


def test_supplement_is_kept_in_inbox_under_single_pdf_schema(library_factory):
    root = library_factory(
        [
            {
                "id": "P1",
                "title": "Main Paper",
                "doi": "10.1000/main",
                "year": "2025",
                "journal": "Test Journal",
            }
        ]
    )
    write_text_pdf(
        root / "Inbox" / "Main Paper - Supporting Information.pdf",
        ["Supporting Information\ndoi:10.1000/main"],
    )
    settings = Settings.from_root(root)
    settings.ensure_runtime_directories()
    result = InboxRegisterWorkflow(settings, UnavailableMetadataService()).run()
    assert result.status == WorkflowStatus.NEEDS_REVIEW
    assert (root / "Inbox" / "Main Paper - Supporting Information.pdf").exists()
    assert any(item.get("issue") == "supplement_parent_unknown" for item in result.needs_review)


def test_unknown_file_blocker_state_is_stable_across_runs(library_factory):
    root = library_factory([])
    write_text_pdf(root / "Inbox" / "unknown.pdf", ["Unknown Document"])
    settings = Settings.from_root(root)
    settings.ensure_runtime_directories()
    InboxRegisterWorkflow(settings, UnavailableMetadataService()).run()
    blocker_path = root / ".library_state" / "inbox_blockers.json"
    first = blocker_path.read_bytes()
    workbook = load_workbook(root / "catalogue.xlsx")
    first_local_id = workbook["Catalogue"]["A2"].value
    InboxRegisterWorkflow(settings, UnavailableMetadataService()).run()
    assert blocker_path.read_bytes() == first
    workbook = load_workbook(root / "catalogue.xlsx")
    assert workbook["Catalogue"].max_row == 2
    assert workbook["Catalogue"]["A2"].value == first_local_id
    payload = json.loads(first)
    assert len(payload["files"]) == 1
    assert payload["files"][0]["issue_keys"] == [
        "metadata_identity_unconfirmed",
        "ocr_unavailable",
    ]


def test_renamed_unchanged_provisional_reuses_saved_local_identity(library_factory):
    root = library_factory([])
    original = root / "Inbox" / "unknown.pdf"
    renamed = root / "Inbox" / "renamed_unknown.pdf"
    write_text_pdf(original, ["Unknown Local Document"])
    settings = Settings.from_root(root)
    settings.ensure_runtime_directories()
    InboxRegisterWorkflow(settings, UnavailableMetadataService()).run()
    workbook = load_workbook(root / "catalogue.xlsx")
    local_id = workbook["Catalogue"]["A2"].value
    original.rename(renamed)

    InboxRegisterWorkflow(settings, UnavailableMetadataService()).run()
    workbook = load_workbook(root / "catalogue.xlsx")
    sheet = workbook["Catalogue"]
    headers = {cell.value: cell.column for cell in sheet[1]}
    assert sheet.max_row == 2
    assert sheet.cell(2, headers["id"]).value == local_id
    assert sheet.cell(2, headers["pdf_relative_path"]).value == "Inbox/renamed_unknown.pdf"


def test_source_change_before_move_is_blocked_and_other_state_is_preserved(
    library_factory, monkeypatch
):
    root = library_factory(
        [
            {
                "id": "P1",
                "title": "Changing Paper",
                "year": "2025",
                "journal": "Change Journal",
                "pdf_filename": "change.pdf",
            }
        ]
    )
    source = root / "Inbox" / "change.pdf"
    write_text_pdf(source, ["Changing Paper\nAuthors"])
    original_apply = FileService.apply_registration_move

    def change_then_apply(self, operation):
        operation.source.write_bytes(operation.source.read_bytes() + b"changed")
        return original_apply(self, operation)

    monkeypatch.setattr(FileService, "apply_registration_move", change_then_apply)
    settings = Settings.from_root(root)
    settings.ensure_runtime_directories()
    result = InboxRegisterWorkflow(settings).run()
    assert result.status == WorkflowStatus.NEEDS_REVIEW
    assert source.exists()
    assert any(item.get("issue") == "source_changed_during_run" for item in result.needs_review)


def test_catalogue_failure_leaves_recoverable_file_moved_journal(
    library_factory, monkeypatch
):
    root = library_factory(
        [
            {
                "id": "P1",
                "title": "Journal Recovery Paper",
                "year": "2025",
                "journal": "Recovery Journal",
                "pdf_filename": "recovery.pdf",
            }
        ]
    )
    write_text_pdf(root / "Inbox" / "recovery.pdf", ["Journal Recovery Paper\nAuthors"])

    def fail_save(self):
        raise CatalogueError("simulated catalogue failure")

    monkeypatch.setattr(CatalogueService, "save_atomic", fail_save)
    settings = Settings.from_root(root)
    settings.ensure_runtime_directories()
    with pytest.raises(CatalogueError, match="simulated catalogue failure"):
        InboxRegisterWorkflow(settings).run()
    target = root / "Registered" / "Recovery Journal, 2025 - Journal Recovery Paper.pdf"
    assert target.exists()
    journal = next((root / ".library_state" / "runs").glob("*/operation_journal.json"))
    payload = json.loads(journal.read_text(encoding="utf-8"))
    assert payload["status"] == "file_moved"
    assert payload["operations"][0]["execution_state"] == "file_moved"


def test_user_confirmed_year_is_used_and_fills_blank_metadata(library_factory):
    root = library_factory(
        [
            {
                "id": "P1",
                "title": "Confirmed Year Paper",
                "journal": "Confirm Journal",
                "pdf_filename": "confirm.pdf",
                "uncertainty": (
                    "USER_CONFIRMED: field=publication_year; value=2025"
                ),
            }
        ]
    )
    write_text_pdf(root / "Inbox" / "confirm.pdf", ["Confirmed Year Paper\nAuthors"])
    settings = Settings.from_root(root)
    settings.ensure_runtime_directories()
    result = InboxRegisterWorkflow(settings).run()
    assert result.status == WorkflowStatus.SUCCESS
    assert (root / "Registered" / "Confirm Journal, 2025 - Confirmed Year Paper.pdf").exists()
    workbook = load_workbook(root / "catalogue.xlsx")
    sheet = workbook["Catalogue"]
    headers = {cell.value: cell.column for cell in sheet[1]}
    assert str(sheet.cell(2, headers["year"]).value) == "2025"


def test_only_direct_visible_inbox_pdfs_are_candidates(library_factory):
    root = library_factory([])
    write_text_pdf(root / "Inbox" / ".hidden.pdf", ["Hidden"])
    write_text_pdf(root / "Inbox" / "subfolder" / "nested.pdf", ["Nested"])
    (root / "Inbox" / "notes.txt").write_text("notes", encoding="utf-8")
    settings = Settings.from_root(root)
    settings.ensure_runtime_directories()
    result = InboxRegisterWorkflow(settings).run(dry_run=True)
    assert result.counts["files_discovered"] == 0
    reasons = {item["reason"] for item in result.skipped}
    assert reasons == {"hidden_or_temporary", "inbox_subdirectory", "non_pdf"}


def test_registered_filename_collision_preserves_both_files(library_factory):
    root = library_factory(
        [
            {
                "id": "P1",
                "title": "Collision Paper",
                "year": "2025",
                "journal": "Collision Journal",
                "pdf_filename": "incoming.pdf",
            }
        ]
    )
    source = root / "Inbox" / "incoming.pdf"
    target = root / "Registered" / "Collision Journal, 2025 - Collision Paper.pdf"
    write_text_pdf(source, ["Collision Paper\nSource"])
    write_text_pdf(target, ["Different existing content"])
    settings = Settings.from_root(root)
    settings.ensure_runtime_directories()
    target_before = target.read_bytes()
    result = InboxRegisterWorkflow(settings).run()
    assert result.status == WorkflowStatus.NEEDS_REVIEW
    assert source.exists()
    assert target.read_bytes() == target_before
    assert any(item.get("issue") == "registered_filename_collision" for item in result.needs_review)


class FixedMetadataService:
    def __init__(self, lookup):
        self.lookup_result = lookup
        self.calls = 0

    def lookup(self, request):
        self.calls += 1
        return self.lookup_result


def test_workflow3_provider_result_creates_row_and_registers(library_factory):
    root = library_factory([])
    write_text_pdf(
        root / "Inbox" / "provider.pdf",
        ["Provider Identified Paper\nAlice Smith\ndoi:10.1000/provider\n2025"],
    )
    metadata = MetadataRecord(
        canonical_id="PMID:12345678",
        title="Provider Identified Paper",
        authors=["Alice Smith"],
        year="2025",
        journal="Biomedical Journal",
        journal_abbrev="Biomed J",
        doi="10.1000/provider",
        pmid="12345678",
        source=["pubmed"],
    )
    lookup = MetadataLookupResult(
        MetadataLookupStatus.FOUND,
        records=[metadata.to_dict()],
        best_record=metadata.to_dict(),
        confidence="exact_identifier",
        providers_used=["pubmed"],
        selection_reason="DOI matched exactly.",
    )
    service = FixedMetadataService(lookup)
    settings = Settings.from_root(root)
    settings.ensure_runtime_directories()
    result = InboxRegisterWorkflow(settings, service).run()
    target = root / "Registered" / "Biomed J, 2025 - Provider Identified Paper.pdf"
    assert target.exists()
    assert service.calls == 1
    assert result.details["files"][0]["match_method"] == "workflow2_provider"
    workbook = load_workbook(root / "catalogue.xlsx")
    assert workbook["Catalogue"].max_row == 2
    assert workbook["Catalogue"]["A2"].value == "PMID:12345678"


def test_workflow3_ambiguous_provider_keeps_file_and_adds_provisional_row(library_factory):
    root = library_factory([])
    write_text_pdf(root / "Inbox" / "ambiguous.pdf", ["Ambiguous Provider Paper"])
    lookup = MetadataLookupResult(
        MetadataLookupStatus.AMBIGUOUS,
        confidence="ambiguous",
        selection_reason="Two candidates remain.",
    )
    settings = Settings.from_root(root)
    settings.ensure_runtime_directories()
    result = InboxRegisterWorkflow(settings, FixedMetadataService(lookup)).run()
    assert result.status == WorkflowStatus.NEEDS_REVIEW
    assert (root / "Inbox" / "ambiguous.pdf").exists()
    sheet = load_workbook(root / "catalogue.xlsx")["Catalogue"]
    assert sheet.max_row == 2
    assert str(sheet["A2"].value).startswith("LOCAL:")


def test_scanned_pdf_ocr_doi_registers_to_existing_row(library_factory):
    root = library_factory(
        [
            {
                "id": "P-OCR",
                "title": "Scanned Biomedical Registration",
                "doi": "10.1000/ocr.register",
                "year": "2025",
                "journal": "OCR Journal",
            }
        ]
    )
    source = root / "Inbox" / "scan.pdf"
    write_text_pdf(source, [""])
    settings = with_ocr(Settings.from_root(root))
    settings.ensure_runtime_directories()
    ocr = FixedOcrService(
        ocr_success(
            title="Scanned Biomedical Registration",
            doi="10.1000/ocr.register",
            year="2025",
        )
    )
    result = InboxRegisterWorkflow(
        settings, UnavailableMetadataService(), ocr
    ).run()
    target = root / "Registered" / "OCR Journal, 2025 - Scanned Biomedical Registration.pdf"
    assert target.exists()
    file_report = result.details["files"][0]
    assert file_report["ocr_triggered"] is True
    assert file_report["ocr_status"] == "success"
    assert file_report["ocr_doi_candidates"] == ["10.1000/ocr.register"]
    assert "combined_text" not in json.dumps(file_report)
    assert result.details["final_check"]["status"] in {"success", "no_changes"}


def test_ocr_fuzzy_title_does_not_auto_register(library_factory):
    root = library_factory(
        [
            {
                "id": "P1",
                "title": "A Highly Specific Biomedical Paper Title",
                "year": "2025",
                "journal": "Test Journal",
            }
        ]
    )
    source = root / "Inbox" / "scan.pdf"
    write_text_pdf(source, [""])
    settings = with_ocr(Settings.from_root(root))
    settings.ensure_runtime_directories()
    ocr = FixedOcrService(
        ocr_success(title="A Highly Specific Biomedical Paper Tltle")
    )
    result = InboxRegisterWorkflow(
        settings, UnavailableMetadataService(), ocr
    ).run()
    assert source.exists()
    assert result.status == WorkflowStatus.NEEDS_REVIEW
    assert "paper_identity_ambiguous" in result.details["files"][0]["issue_keys"]


def test_ocr_exact_title_without_support_still_requires_workflow2(library_factory):
    root = library_factory(
        [
            {
                "id": "P1",
                "title": "Exact OCR Title Needs Support",
                "year": "2025",
                "journal": "Test Journal",
            }
        ]
    )
    source = root / "Inbox" / "scan.pdf"
    write_text_pdf(source, [""])
    settings = with_ocr(Settings.from_root(root))
    settings.ensure_runtime_directories()
    result = InboxRegisterWorkflow(
        settings,
        UnavailableMetadataService(),
        FixedOcrService(ocr_success(title="Exact OCR Title Needs Support")),
    ).run()
    assert source.exists()
    assert result.details["metadata_lookup_requests"] == 1
    assert "metadata_provider_unavailable" in result.details["files"][0]["issue_keys"]


def test_ocr_unavailable_for_one_file_does_not_block_other_file(library_factory):
    root = library_factory(
        [
            {
                "id": "P1",
                "title": "Ready Text Paper",
                "year": "2025",
                "journal": "Ready Journal",
                "pdf_filename": "ready.pdf",
            }
        ]
    )
    write_text_pdf(root / "Inbox" / "blank.pdf", [""])
    write_text_pdf(root / "Inbox" / "ready.pdf", ["Ready Text Paper\nAuthor"])
    unavailable = OcrInspection(
        status="ocr_unavailable_model_missing",
        errors=["ocr_unavailable_model_missing"],
    )
    settings = with_ocr(Settings.from_root(root))
    settings.ensure_runtime_directories()
    ocr = FixedOcrService({"blank.pdf": unavailable, "ready.pdf": unavailable})
    result = InboxRegisterWorkflow(
        settings, UnavailableMetadataService(), ocr
    ).run()
    assert (root / "Inbox" / "blank.pdf").exists()
    assert (root / "Registered" / "Ready Journal, 2025 - Ready Text Paper.pdf").exists()
    assert result.changed_files == 1
    assert any(
        item.get("issue") == "ocr_unavailable_model_missing"
        for item in result.needs_review
    )


def test_corrected_ocr_doi_alone_does_not_match_catalogue(library_factory):
    root = library_factory(
        [
            {
                "id": "P1",
                "title": "Different Catalogue Title",
                "doi": "10.1000/corrected",
                "year": "2025",
                "journal": "Test Journal",
            }
        ]
    )
    source = root / "Inbox" / "scan.pdf"
    write_text_pdf(source, [""])
    ocr = OcrInspection(
        status="success",
        title_candidates=[TitleCandidate("Unconfirmed OCR Title", "medium", "ocr_page_top", 1)],
        doi_candidates=[
            IdentifierCandidate(
                "10.1000/corrected", 1, "1O.1000/corrected", "medium", "ocr_corrected"
            )
        ],
    )
    settings = with_ocr(Settings.from_root(root))
    settings.ensure_runtime_directories()
    result = InboxRegisterWorkflow(
        settings, UnavailableMetadataService(), FixedOcrService(ocr)
    ).run()
    assert source.exists()
    assert result.details["files"][0]["match_method"] == "none"
    assert "metadata_provider_unavailable" in result.details["files"][0]["issue_keys"]


def test_ocr_doi_workflow2_metadata_is_revalidated_before_registration(library_factory):
    root = library_factory([])
    source = root / "Inbox" / "scan.pdf"
    write_text_pdf(source, [""])
    metadata = MetadataRecord(
        canonical_id="DOI:10.1000/newocr",
        title="Workflow Two Confirmed OCR Paper",
        authors=["Alice Smith"],
        year="2025",
        journal="Metadata Journal",
        doi="10.1000/newocr",
        source=["pubmed"],
    )
    lookup = MetadataLookupResult(
        MetadataLookupStatus.FOUND,
        records=[metadata.to_dict()],
        best_record=metadata.to_dict(),
        confidence="exact_identifier",
        providers_used=["pubmed"],
    )
    settings = with_ocr(Settings.from_root(root))
    settings.ensure_runtime_directories()
    ocr = FixedOcrService(
        ocr_success(
            title="Workflow Two Confirmed OCR Paper",
            doi="10.1000/newocr",
            year="2025",
        )
    )
    result = InboxRegisterWorkflow(settings, FixedMetadataService(lookup), ocr).run()
    assert (root / "Registered" / "Metadata Journal, 2025 - Workflow Two Confirmed OCR Paper.pdf").exists()
    assert result.details["files"][0]["match_method"] == "workflow2_provider"


def test_filename_provider_success_skips_pdf_inspection(library_factory, monkeypatch):
    root = library_factory([])
    name = "Biomed J, 2025 - Provider Canonical Paper.pdf"
    write_text_pdf(root / "Inbox" / name, [""])
    metadata = MetadataRecord(
        canonical_id="PMID:11112222",
        title="Provider Canonical Paper",
        authors=["Alice Smith"],
        year="2025",
        journal="Biomedical Journal",
        journal_abbrev="Biomed J",
        doi="10.1000/provider.canonical",
        pmid="11112222",
        source=["pubmed"],
    )
    lookup = MetadataLookupResult(
        MetadataLookupStatus.FOUND,
        records=[metadata.to_dict()],
        best_record=metadata.to_dict(),
        confidence="exact_title_supported",
        providers_used=["pubmed"],
        selection_reason="Filename title and year matched uniquely.",
    )

    def fail_inspection(*args, **kwargs):
        raise AssertionError("PDF content should not be opened after filename lookup succeeds")

    monkeypatch.setattr("lam.services.pdf_service.PdfService.inspect", fail_inspection)
    settings = Settings.from_root(root)
    settings.ensure_runtime_directories()
    result = InboxRegisterWorkflow(settings, FixedMetadataService(lookup)).run()
    assert result.status == WorkflowStatus.SUCCESS
    assert result.details["files"][0]["inspection_level_used"] == "skip"
    assert result.details["files"][0]["canonical_title_selected"] == metadata.title
    assert (root / "Registered" / name).exists()


def test_filename_lookup_failure_then_pypdf_creates_provisional(library_factory):
    root = library_factory([])
    name = "Test J, 2025 - Filename Search Candidate.pdf"
    write_text_pdf(root / "Inbox" / name, ["PDF Layer Candidate\nAlice Smith\n2025"])
    settings = Settings.from_root(root)
    settings.ensure_runtime_directories()
    service = FixedMetadataService(
        MetadataLookupResult(MetadataLookupStatus.NOT_FOUND, selection_reason="No result")
    )
    result = InboxRegisterWorkflow(settings, service).run()
    file_result = result.details["files"][0]
    assert service.calls == 2
    assert file_result["inspection_level_used"] == "pypdf_text"
    assert file_result["result_status"] == "provisional"
    assert (root / "Inbox" / name).exists()


def test_pypdf_success_does_not_trigger_ocr(library_factory):
    root = library_factory([])
    write_text_pdf(
        root / "Inbox" / "local.pdf",
        ["A High Quality Local PDF Title\nAlice Smith\n2025"],
    )
    settings = with_ocr(Settings.from_root(root))
    settings.ensure_runtime_directories()
    ocr = FixedOcrService(ocr_success(title="Unused OCR Title"))
    result = InboxRegisterWorkflow(
        settings, UnavailableMetadataService(), ocr
    ).run()
    assert ocr.calls == []
    assert result.details["files"][0]["inspection_level_used"] == "pypdf_text"


def test_pypdf_insufficient_triggers_ocr_and_keeps_provisional(library_factory):
    root = library_factory([])
    write_text_pdf(root / "Inbox" / "scan.pdf", [""])
    settings = with_ocr(Settings.from_root(root))
    settings.ensure_runtime_directories()
    ocr = FixedOcrService(ocr_success(title="OCR Identified Local Candidate"))
    result = InboxRegisterWorkflow(
        settings, UnavailableMetadataService(), ocr
    ).run()
    assert ocr.calls == ["scan.pdf"]
    assert result.details["files"][0]["inspection_level_used"] == "ocr"
    assert result.details["files"][0]["result_status"] in {"provisional", "blocked"}
    assert (root / "Inbox" / "scan.pdf").exists()


def test_provisional_record_uid_is_stable_when_provider_upgrades_id(library_factory):
    root = library_factory([])
    name = "Biomed J, 2025 - Temporary Local Title.pdf"
    write_text_pdf(root / "Inbox" / name, ["Temporary Local Title\nAlice Smith\n2025"])
    settings = Settings.from_root(root)
    settings.ensure_runtime_directories()
    first = InboxRegisterWorkflow(
        settings,
        FixedMetadataService(
            MetadataLookupResult(MetadataLookupStatus.AMBIGUOUS, selection_reason="Two results")
        ),
    ).run()
    assert first.status == WorkflowStatus.NEEDS_REVIEW
    workbook = load_workbook(root / "catalogue.xlsx")
    sheet = workbook["Catalogue"]
    headers = {cell.value: cell.column for cell in sheet[1]}
    local_id = sheet.cell(2, headers["id"]).value
    record_uid = sheet.cell(2, headers["record_uid"]).value
    assert str(local_id).startswith("LOCAL:")

    metadata = MetadataRecord(
        canonical_id="PMID:22223333",
        title="Provider Canonical Replacement Title",
        authors=["Alice Smith"],
        year="2025",
        journal="Biomedical Journal",
        journal_abbrev="Biomed J",
        doi="10.1000/replacement",
        pmid="22223333",
        source=["pubmed"],
    )
    second = InboxRegisterWorkflow(
        settings,
        FixedMetadataService(
            MetadataLookupResult(
                MetadataLookupStatus.FOUND,
                records=[metadata.to_dict()],
                best_record=metadata.to_dict(),
                confidence="exact_title_supported",
                providers_used=["pubmed"],
                selection_reason="Provider identity confirmed.",
            )
        ),
    ).run()
    assert second.status == WorkflowStatus.SUCCESS
    workbook = load_workbook(root / "catalogue.xlsx")
    sheet = workbook["Catalogue"]
    assert sheet.max_row == 2
    assert sheet.cell(2, headers["id"]).value == "PMID:22223333"
    assert sheet.cell(2, headers["record_uid"]).value == record_uid
    assert sheet.cell(2, headers["title"]).value == metadata.title
    assert not sheet.cell(2, headers["uncertainty"]).value


def test_provisional_uses_pypdf_title_before_filename(library_factory):
    root = library_factory([])
    write_text_pdf(
        root / "Inbox" / "Misleading Filename Candidate.pdf",
        ["Reliable PDF Layer Paper Title\nAlice Smith"],
    )
    settings = Settings.from_root(root)
    settings.ensure_runtime_directories()
    InboxRegisterWorkflow(settings, UnavailableMetadataService()).run()
    workbook = load_workbook(root / "catalogue.xlsx")
    sheet = workbook["Catalogue"]
    headers = {cell.value: cell.column for cell in sheet[1]}
    assert sheet.cell(2, headers["title"]).value == "Reliable PDF Layer Paper Title"
    assert "title_provisional_pypdf" in sheet.cell(2, headers["uncertainty"]).value


def test_standard_filename_beats_administrative_pypdf_heading(library_factory):
    root = library_factory([])
    name = "Test J, 2025 - Reliable Filename Paper Title.pdf"
    write_text_pdf(
        root / "Inbox" / name,
        ["Accepted: 7 July 2025 / Published online: 31 July 2025"],
    )
    settings = Settings.from_root(root)
    settings.ensure_runtime_directories()
    InboxRegisterWorkflow(settings, UnavailableMetadataService()).run()
    workbook = load_workbook(root / "catalogue.xlsx")
    sheet = workbook["Catalogue"]
    headers = {cell.value: cell.column for cell in sheet[1]}
    assert sheet.cell(2, headers["title"]).value == "Reliable Filename Paper Title"
    assert "title_provisional_filename" in sheet.cell(2, headers["uncertainty"]).value


def test_filename_title_is_preferred_to_ocr_for_provisional(library_factory):
    root = library_factory([])
    name = "Meaningful Filename Candidate.pdf"
    write_text_pdf(root / "Inbox" / name, [""])
    settings = with_ocr(Settings.from_root(root))
    settings.ensure_runtime_directories()
    InboxRegisterWorkflow(
        settings,
        UnavailableMetadataService(),
        FixedOcrService(ocr_success(title="Different OCR Candidate")),
    ).run()
    workbook = load_workbook(root / "catalogue.xlsx")
    sheet = workbook["Catalogue"]
    headers = {cell.value: cell.column for cell in sheet[1]}
    assert sheet.cell(2, headers["title"]).value == "Meaningful Filename Candidate"
    assert "title_provisional_filename" in sheet.cell(2, headers["uncertainty"]).value


def test_user_confirmed_provisional_title_is_not_overwritten(library_factory):
    root = library_factory(
        [
            {
                "id": "LOCAL:fixed-id",
                "title": "User Chosen Local Title",
                "pdf_status": "inbox",
                "pdf_filename": "candidate.pdf",
                "pdf_relative_path": "Inbox/candidate.pdf",
                "uncertainty": (
                    "USER_CONFIRMED: field=title; value=User Chosen Local Title\n"
                    "NEEDS_REVIEW: field=paper_identity; "
                    "issue_key=metadata_identity_unconfirmed; issue=Pending identity."
                ),
            }
        ]
    )
    write_text_pdf(root / "Inbox" / "candidate.pdf", ["User Chosen Local Title\n2025"])
    metadata = MetadataRecord(
        canonical_id="PMID:77778888",
        title="Different Provider Canonical Title",
        authors=["Alice Smith"],
        year="2025",
        journal="Provider Journal",
        pmid="77778888",
        source=["pubmed"],
    )
    lookup = MetadataLookupResult(
        MetadataLookupStatus.FOUND,
        records=[metadata.to_dict()],
        best_record=metadata.to_dict(),
        confidence="exact_title_supported",
        providers_used=["pubmed"],
    )
    settings = Settings.from_root(root)
    settings.ensure_runtime_directories()
    result = InboxRegisterWorkflow(settings, FixedMetadataService(lookup)).run()
    assert result.status == WorkflowStatus.NEEDS_REVIEW
    workbook = load_workbook(root / "catalogue.xlsx")
    sheet = workbook["Catalogue"]
    headers = {cell.value: cell.column for cell in sheet[1]}
    assert sheet.cell(2, headers["title"]).value == "User Chosen Local Title"
    assert "USER_CONFIRMED: field=title" in sheet.cell(2, headers["uncertainty"]).value


def test_provisional_possible_formal_duplicate_is_blocked_without_merge(library_factory):
    root = library_factory(
        [
            {
                "id": "PMID:99990000",
                "title": "Formal Paper Identity",
                "doi": "10.1000/formal.duplicate",
                "year": "2025",
                "journal": "Formal Journal",
            },
            {
                "id": "LOCAL:provisional-id",
                "title": "Local Copy",
                "pdf_status": "inbox",
                "pdf_filename": "copy.pdf",
                "pdf_relative_path": "Inbox/copy.pdf",
                "uncertainty": (
                    "NEEDS_REVIEW: field=paper_identity; "
                    "issue_key=metadata_identity_unconfirmed; issue=Pending identity."
                ),
            },
        ]
    )
    write_text_pdf(
        root / "Inbox" / "copy.pdf",
        ["Formal Paper Identity\ndoi:10.1000/formal.duplicate\n2025"],
    )
    settings = Settings.from_root(root)
    settings.ensure_runtime_directories()
    result = InboxRegisterWorkflow(settings, UnavailableMetadataService()).run()
    assert result.status == WorkflowStatus.NEEDS_REVIEW
    assert "possible_duplicate_provisional_record" in result.details["files"][0]["issue_keys"]
    assert load_workbook(root / "catalogue.xlsx")["Catalogue"].max_row == 3
    assert (root / "Inbox" / "copy.pdf").exists()
