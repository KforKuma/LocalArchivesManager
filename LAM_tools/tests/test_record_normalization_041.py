from __future__ import annotations

import uuid
from pathlib import Path

from openpyxl import load_workbook

from lam.config import Settings
from lam.models import (
    MetadataLookupRequest,
    MetadataLookupResult,
    MetadataLookupStatus,
    MetadataRecord,
    WorkflowResult,
)
from lam.services.catalogue_service import CatalogueService
from lam.services.record_canonicalization_service import RegisteredRecordCanonicalizer
from lam.services.snapshot_service import SnapshotService
from lam.workflows.inbox_register import InboxRegisterWorkflow
from lam.workflows.metadata_query import MetadataQueryWorkflow
from lam.workflows.record_normalization import RecordNormalizationWorkflow

from conftest import write_text_pdf


class FixedMetadataService:
    def __init__(self, metadata: MetadataRecord):
        self.metadata = metadata
        self.requests = []

    def lookup(self, request):
        self.requests.append(request)
        return found(self.metadata)

    def lookup_many(self, requests):
        self.requests.extend(requests)
        return [found(self.metadata) for _ in requests]


class NotFoundMetadataService:
    def lookup(self, request):
        return MetadataLookupResult(
            MetadataLookupStatus.NOT_FOUND,
            selection_reason="No result",
        )


def found(metadata: MetadataRecord) -> MetadataLookupResult:
    return MetadataLookupResult(
        MetadataLookupStatus.FOUND,
        records=[metadata.to_dict()],
        best_record=metadata.to_dict(),
        confidence="exact_identifier",
        providers_used=list(metadata.source),
        selection_reason="Identifier matched exactly.",
    )


def add_fields(path: Path, row_number: int, **values) -> None:
    workbook = load_workbook(path)
    sheet = workbook["Catalogue"]
    headers = {cell.value: cell.column for cell in sheet[1] if cell.value}
    for field_name, value in values.items():
        if field_name not in headers:
            column = sheet.max_column + 1
            sheet.cell(1, column).value = field_name
            headers[field_name] = column
        sheet.cell(row_number, headers[field_name]).value = value
    workbook.save(path)


def row_values(path: Path, row_number: int = 2) -> dict[str, object]:
    workbook = load_workbook(path)
    sheet = workbook["Catalogue"]
    return {
        str(cell.value): sheet.cell(row_number, cell.column).value
        for cell in sheet[1]
        if cell.value
    }


def pubmed_record(**overrides) -> MetadataRecord:
    values = {
        "canonical_id": "PMID:12345678",
        "title": "Canonical Paper Title",
        "authors": ["Alice Smith", "Bob Jones"],
        "year": "2002",
        "journal": "Structure (London, England : 1993)",
        "journal_abbrev": "Structure",
        "doi": "10.1000/canonical",
        "pmid": "12345678",
        "abstract": "Canonical abstract.",
        "source": ["pubmed", "unpaywall"],
    }
    values.update(overrides)
    return MetadataRecord(**values)


def test_01_exact_match_preserves_paper_uuid_and_pmid_field(library_factory):
    root = library_factory([{"id": "LOCAL:x", "title": "Canonical Paper Title", "pmid": "12345678"}])
    service = CatalogueService(root / "catalogue.xlsx")
    record = service.load()[0]
    paper_uuid = record.get("paper_uuid")
    result = RegisteredRecordCanonicalizer().canonicalize(service, record, pubmed_record())
    assert result.accepted
    assert record.get("paper_uuid") == paper_uuid
    assert record.get("pmid") == "12345678"
    assert "id" not in service.headers


def test_02_paper_uuid_is_immutable_during_canonicalization(library_factory):
    root = library_factory([{"id": "LOCAL:x", "title": "Canonical Paper Title", "pmid": "12345678"}])
    service = CatalogueService(root / "catalogue.xlsx")
    record = service.load()[0]
    stable_uid = record.get("paper_uuid")
    RegisteredRecordCanonicalizer().canonicalize(service, record, pubmed_record())
    assert record.get("paper_uuid") == stable_uid


def test_03_snapshot_matches_row_by_paper_uuid_after_external_id_update(library_factory):
    root = library_factory([{"id": "LOCAL:x", "title": "Canonical Paper Title"}])
    service = CatalogueService(root / "catalogue.xlsx")
    record = service.load()[0]
    before = service.snapshot_payload()
    service.update_canonical_fields(record, {"pmid": "12345678"})
    after = service.snapshot_payload()
    assert before["rows"][0]["paper_uuid"] == after["rows"][0]["paper_uuid"]
    assert service.changes[0].field_name == "pmid"


def test_04_issue_key_changes_when_external_identity_evidence_changes(library_factory):
    root = library_factory([{"id": "LOCAL:x", "title": "Canonical Paper Title"}])
    service = CatalogueService(root / "catalogue.xlsx")
    record = service.load()[0]
    before = service._review_decision_key(record, "paper_identity", "pending", "Pending")
    service.update_canonical_fields(record, {"pmid": "12345678"})
    after = service._review_decision_key(record, "paper_identity", "pending", "Pending")
    assert before != after
    assert record.get("paper_uuid")


def test_05_pubmed_becomes_canonical_source(library_factory):
    root = library_factory([{"id": "LOCAL:x", "title": "Canonical Paper Title", "pmid": "12345678"}])
    add_fields(root / "catalogue.xlsx", 2, source="local_pdf; pubmed; unpaywall")
    service = CatalogueService(root / "catalogue.xlsx")
    record = service.load()[0]
    RegisteredRecordCanonicalizer().canonicalize(service, record, pubmed_record())
    assert record.get("source") == "pubmed"


def test_06_local_only_source_remains_local_pdf(library_factory):
    root = library_factory([{"id": "LOCAL:x", "title": "Local Paper"}])
    add_fields(root / "catalogue.xlsx", 2, source="local_pdf")
    record = CatalogueService(root / "catalogue.xlsx").load()[0]
    assert RegisteredRecordCanonicalizer.canonical_source(MetadataRecord(), record) == "local_pdf"


def test_07_structure_moves_to_journal_abbrev(library_factory):
    root = library_factory([{"id": "LOCAL:x", "title": "Canonical Paper Title", "pmid": "12345678", "journal": "Structure"}])
    service = CatalogueService(root / "catalogue.xlsx")
    record = service.load()[0]
    RegisteredRecordCanonicalizer().canonicalize(service, record, pubmed_record())
    assert record.get("journal_abbrev") == "Structure"


def test_08_pubmed_full_journal_is_written(library_factory):
    root = library_factory([{"id": "LOCAL:x", "title": "Canonical Paper Title", "pmid": "12345678", "journal": "Structure"}])
    service = CatalogueService(root / "catalogue.xlsx")
    record = service.load()[0]
    RegisteredRecordCanonicalizer().canonicalize(service, record, pubmed_record())
    assert record.get("journal") == "Structure (London, England : 1993)"


def test_09_equivalent_journal_removes_variant_note(library_factory):
    root = library_factory([{
        "id": "LOCAL:x",
        "title": "Canonical Paper Title",
        "pmid": "12345678",
        "journal": "Structure",
        "uncertainty": "USER_CONFIRMED:\nMACHINE_NOTE: field=journal; issue_key=journal_name_variant; issue=Variant.",
    }])
    service = CatalogueService(root / "catalogue.xlsx")
    record = service.load()[0]
    result = RegisteredRecordCanonicalizer().canonicalize(service, record, pubmed_record())
    assert not result.conflicts
    assert record.get("uncertainty") == "USER_CONFIRMED:"


def test_10_different_journal_is_hard_conflict(library_factory):
    root = library_factory([{"id": "LOCAL:x", "title": "Canonical Paper Title", "pmid": "12345678", "journal": "Nature"}])
    service = CatalogueService(root / "catalogue.xlsx")
    record = service.load()[0]
    result = RegisteredRecordCanonicalizer().canonicalize(service, record, pubmed_record())
    assert result.conflicts == ["metadata_journal_conflict"]
    assert record.get("journal") == "Nature"


def test_11_workflow2_completes_registered_record_without_moving_pdf(library_factory):
    name = "Existing.pdf"
    root = library_factory([{
        "id": "LOCAL:x",
        "title": "Canonical Paper Title",
        "pmid": "12345678",
        "pdf_status": "registered",
        "pdf_filename": name,
        "pdf_relative_path": f"Registered/{name}",
    }], {f"Registered/{name}": b"registered"})
    add_fields(root / "catalogue.xlsx", 2, source="local_pdf", abstract=None)
    settings = Settings.from_root(root)
    settings.ensure_runtime_directories()
    result = MetadataQueryWorkflow(settings, FixedMetadataService(pubmed_record())).run(
        MetadataLookupRequest(), incomplete_records=True
    )
    row = row_values(root / "catalogue.xlsx")
    assert result.changed_files == 0
    assert (root / "Registered" / name).is_file()
    assert row["pmid"] == "12345678"
    assert "id" not in row
    assert row["authors"] == "Alice Smith; Bob Jones"
    assert row["abstract"] == "Canonical abstract."
    assert row["source"] == "pubmed"


def test_12_workflow3_does_not_scan_registered(library_factory):
    name = "Registered Only.pdf"
    root = library_factory([], {f"Registered/{name}": b"registered"})
    settings = Settings.from_root(root)
    settings.ensure_runtime_directories()
    result = InboxRegisterWorkflow(settings, NotFoundMetadataService()).run()
    assert result.counts["files_discovered"] == 0
    assert (root / "Registered" / name).is_file()


def test_13_normalize_records_dry_run_does_not_modify_catalogue(library_factory):
    root = library_factory([{"id": "LOCAL:x", "title": "Canonical Paper Title", "pmid": "12345678"}])
    add_fields(root / "catalogue.xlsx", 2, source="local_pdf; pubmed; unpaywall")
    before = (root / "catalogue.xlsx").read_bytes()
    settings = Settings.from_root(root)
    settings.ensure_runtime_directories()
    result = RecordNormalizationWorkflow(settings, FixedMetadataService(pubmed_record())).run(dry_run=True)
    assert result.dry_run
    assert (root / "catalogue.xlsx").read_bytes() == before
    assert not list(root.glob("catalogue.backup.*.xlsx"))


def test_14_user_confirmed_and_free_text_are_preserved(library_factory):
    uncertainty = "USER_CONFIRMED:\nKeep this free-form note.\nNEEDS_REVIEW: field=paper_identity; issue_key=metadata_identity_unconfirmed; issue=Pending."
    root = library_factory([{"id": "LOCAL:x", "title": "Canonical Paper Title", "pmid": "12345678", "uncertainty": uncertainty}])
    service = CatalogueService(root / "catalogue.xlsx")
    record = service.load()[0]
    RegisteredRecordCanonicalizer().canonicalize(service, record, pubmed_record())
    assert record.get("uncertainty") == "USER_CONFIRMED:\nKeep this free-form note."


def test_15_metadata_normalization_runs_final_check_once(library_factory, monkeypatch):
    root = library_factory([{"id": "LOCAL:x", "title": "Canonical Paper Title", "pmid": "12345678"}])
    add_fields(root / "catalogue.xlsx", 2, source="local_pdf")
    settings = Settings.from_root(root)
    settings.ensure_runtime_directories()
    calls = []

    def fake_final_check(self, *, dry_run=False, final_check=False):
        calls.append(final_check)
        result = WorkflowResult("daily_check")
        result.state_committed = True
        return result

    monkeypatch.setattr("lam.workflows.metadata_query.DailyCheckWorkflow.run", fake_final_check)
    RecordNormalizationWorkflow(settings, FixedMetadataService(pubmed_record())).run()
    assert calls == [True]


def test_16_normalize_existing_uses_pmid_without_title_query(library_factory):
    root = library_factory([{"id": "LOCAL:x", "title": "Untrusted title", "pmid": "12345678"}])
    add_fields(root / "catalogue.xlsx", 2, source="local_pdf")
    metadata_service = FixedMetadataService(pubmed_record(title="Untrusted title"))
    settings = Settings.from_root(root)
    settings.ensure_runtime_directories()
    MetadataQueryWorkflow(settings, metadata_service).run(
        MetadataLookupRequest(), normalize_existing=True, dry_run=True
    )
    assert metadata_service.requests[0].pmid == "12345678"
    assert metadata_service.requests[0].title is None


def test_17_doi_remains_dedicated_external_identifier(library_factory):
    root = library_factory([{"id": "LOCAL:x", "title": "Canonical Paper Title", "doi": "10.1000/canonical"}])
    add_fields(root / "catalogue.xlsx", 2, source="local_pdf")
    metadata = pubmed_record(pmid="", source=["unpaywall"])
    service = CatalogueService(root / "catalogue.xlsx")
    record = service.load()[0]
    RegisteredRecordCanonicalizer().canonicalize(service, record, metadata)
    assert record.get("doi") == "10.1000/canonical"
    assert "id" not in service.headers
    assert record.get("source") == "unpaywall"


def test_18_operation_journal_contains_paper_uuid(library_factory):
    root = library_factory([{"id": "LOCAL:x", "title": "Canonical Paper Title", "pmid": "12345678"}])
    add_fields(root / "catalogue.xlsx", 2, source="local_pdf")
    settings = Settings.from_root(root)
    settings.ensure_runtime_directories()
    RecordNormalizationWorkflow(settings, FixedMetadataService(pubmed_record())).run()
    journal = next((root / ".library_state" / "runs").glob("*/operation_journal.json"))
    import json

    payload = json.loads(journal.read_text(encoding="utf-8"))
    assert payload["operations"][0]["paper_uuid"]


def test_19_normalize_records_preserves_paper_uuid_without_identifiers(library_factory):
    root = library_factory([
        {"id": "LOCAL:no-id", "title": "Local Only"},
        {"id": "LOCAL:online", "title": "Canonical Paper Title", "pmid": "12345678"},
    ])
    add_fields(root / "catalogue.xlsx", 2, source="local_pdf")
    add_fields(root / "catalogue.xlsx", 3, source="local_pdf")
    settings = Settings.from_root(root)
    settings.ensure_runtime_directories()
    RecordNormalizationWorkflow(settings, FixedMetadataService(pubmed_record())).run()
    workbook = load_workbook(root / "catalogue.xlsx")
    sheet = workbook["Catalogue"]
    headers = {cell.value: cell.column for cell in sheet[1]}
    assert uuid.UUID(str(sheet.cell(2, headers["paper_uuid"]).value)).version == 4
    assert uuid.UUID(str(sheet.cell(3, headers["paper_uuid"]).value)).version == 4
    assert "record_uid" not in headers


def test_20_normalization_reports_filename_change_without_moving_pdf(library_factory):
    name = "Old Name.pdf"
    root = library_factory([{
        "id": "LOCAL:x",
        "title": "Temporary title",
        "pmid": "12345678",
        "year": "2002",
        "journal": "Structure",
        "pdf_status": "registered",
        "pdf_filename": name,
        "pdf_relative_path": f"Registered/{name}",
    }], {f"Registered/{name}": b"registered"})
    add_fields(root / "catalogue.xlsx", 2, source="local_pdf")
    settings = Settings.from_root(root)
    settings.ensure_runtime_directories()
    result = RecordNormalizationWorkflow(settings, FixedMetadataService(pubmed_record())).run(dry_run=True)
    implication = result.details["filename_implications"][0]
    assert implication["location_scope"] == "registered"
    assert implication["action"] == "report_only"
    assert implication["proposed_filename"].startswith("Structure, 2002 - Canonical Paper Title")
    assert (root / "Registered" / name).is_file()
