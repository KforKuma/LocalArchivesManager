from __future__ import annotations

import logging
from pathlib import Path

from openpyxl import Workbook, load_workbook

from lam.config import Settings
from lam.models import (
    IdentifierCandidate,
    MetadataLookupResult,
    MetadataLookupStatus,
    OcrInspection,
    PdfInspection,
    TitleCandidate,
)
from lam.schema import CATALOGUE_FIELDS, DOCUMENT_FIELDS
from lam.services.pdf_service import PdfService, _PypdfWarningHandler
from lam.utils.local_pdf_metadata import extract_local_pdf_metadata
from lam.utils.text import title_candidates_from_page
from lam.workflows.inbox_register import InboxRegisterWorkflow


ENGLISH_ABSTRACT = (
    "This study presents a bilingual biomedical method with reliable experimental "
    "validation and discusses the resulting implications for clinical research."
)
CHINESE_ABSTRACT = (
    "本研究提出一种面向生物医学文献的双语分析方法，并通过可靠实验验证其有效性，"
    "同时讨论该方法在临床研究和知识管理中的潜在应用价值。"
)


def _bilingual_text() -> str:
    return "\n".join(
        (
            "Vol. 42 No. 8 Aug. 2022",
            "中国生物医学人工智能研究",
            "Artificial Intelligence Methods for Biomedical Research",
            "张三，李四",
            "Zhang San; Li Si",
            "中国医学科学院学报 Vol. 42 No. 8 Aug. 2022",
            "DOI: 10.1234/bilingual.2022.8",
            "摘要：" + CHINESE_ABSTRACT,
            "关键词：人工智能；生物医学",
            "Abstract: " + ENGLISH_ABSTRACT,
            "Keywords: artificial intelligence; biomedicine",
            "Introduction",
        )
    )


def _empty_library(tmp_path: Path) -> Path:
    root = tmp_path / "library"
    for name in ("Inbox", "Registered", "Topics"):
        (root / name).mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    catalogue = workbook.active
    catalogue.title = "Catalogue"
    catalogue.append(CATALOGUE_FIELDS)
    documents = workbook.create_sheet("Documents")
    documents.append(DOCUMENT_FIELDS)
    workbook.save(root / "catalogue.xlsx")
    return root


def _append_catalogue_row(root: Path, values: dict[str, object]) -> None:
    workbook = load_workbook(root / "catalogue.xlsx")
    sheet = workbook["Catalogue"]
    sheet.append([values.get(field) for field in CATALOGUE_FIELDS])
    workbook.save(root / "catalogue.xlsx")
    workbook.close()


class _NotFoundMetadata:
    def __init__(self):
        self.requests = []

    def lookup(self, request):
        self.requests.append(request)
        return MetadataLookupResult(
            MetadataLookupStatus.NOT_FOUND,
            selection_reason="No online record was found.",
        )


def test_bilingual_first_page_extracts_titles_abstract_authors_journal_and_doi():
    metadata = extract_local_pdf_metadata(
        first_page_text=_bilingual_text(),
        filename="中文论文.pdf",
        title_candidates=title_candidates_from_page(_bilingual_text(), page=1),
        doi_candidates=["10.1234/bilingual.2022.8"],
        pmid_candidates=[],
    )

    assert metadata.primary_title == "中国生物医学人工智能研究"
    assert metadata.local_language_title == "中国生物医学人工智能研究"
    assert metadata.english_title == "Artificial Intelligence Methods for Biomedical Research"
    assert metadata.translated_title == metadata.english_title
    assert metadata.title == metadata.primary_title
    assert metadata.authors == ("Zhang San", "Li Si")
    assert metadata.journal == "中国医学科学院学报"
    assert metadata.year == "2022"
    assert metadata.doi == "10.1234/bilingual.2022.8"
    assert metadata.abstract == ENGLISH_ABSTRACT


def test_chinese_abstract_is_used_only_when_english_abstract_is_absent():
    text = _bilingual_text().split("Abstract:", 1)[0] + "Introduction"
    metadata = extract_local_pdf_metadata(
        first_page_text=text,
        filename="中文论文.pdf",
        title_candidates=[],
        doi_candidates=[],
        pmid_candidates=[],
    )
    assert metadata.abstract == CHINESE_ABSTRACT


def test_volume_header_is_not_a_title_candidate():
    candidates = title_candidates_from_page(
        "Vol. 42 No. 8 Aug. 2022\n"
        "V ol . 42 No . 8 Aug . 2022\n"
        "Chinese Medical Journal Vol. 42 No. 8\n"
        "Artificial Intelligence Methods for Biomedical Research",
        page=1,
    )
    assert [item.value for item in candidates] == [
        "Artificial Intelligence Methods for Biomedical Research"
    ]


def test_provider_not_found_regates_ocr_and_first_provisional_has_local_fields(
    tmp_path: Path, monkeypatch
):
    root = _empty_library(tmp_path)
    source = root / "Inbox" / "中文论文.pdf"
    source.write_bytes(b"%PDF-1.4\nlocal\n%%EOF\n")
    stat = source.stat()
    pypdf_local = {
        "title": "中国生物医学人工智能研究",
        "primary_title": "中国生物医学人工智能研究",
        "translated_title": "Artificial Intelligence Methods for Biomedical Research",
        "local_language_title": "中国生物医学人工智能研究",
        "english_title": "Artificial Intelligence Methods for Biomedical Research",
        "authors": (),
        "year": "2022",
        "journal": "",
        "doi": "10.1234/bilingual.2022.8",
        "abstract": "",
        "field_sources": {"title": "pypdf_first_page", "doi": "pypdf_first_page"},
        "field_confidence": {"title": "high", "doi": "high"},
        "warnings": (),
    }
    complete_local = {
        **pypdf_local,
        "authors": ("Zhang San", "Li Si"),
        "journal": "中国医学科学院学报",
        "abstract": ENGLISH_ABSTRACT,
        "field_sources": {
            **pypdf_local["field_sources"],
            "authors": "ocr_first_page",
            "journal": "ocr_first_page",
            "abstract": "ocr_first_page",
        },
        "field_confidence": {
            **pypdf_local["field_confidence"],
            "authors": "medium",
            "journal": "medium",
            "abstract": "high",
        },
    }
    calls: list[tuple[str, str | None]] = []

    def fake_inspect(self, path, *, ocr_mode="auto", ocr_trigger_reason=None, **kwargs):
        calls.append((ocr_mode, ocr_trigger_reason))
        local = complete_local if ocr_mode == "always" else pypdf_local
        return PdfInspection(
            relative_path="Inbox/中文论文.pdf",
            filename=source.name,
            size=stat.st_size,
            mtime_ns=stat.st_mtime_ns,
            is_readable=True,
            page_count=1,
            first_page_text=_bilingual_text(),
            sampled_text=_bilingual_text(),
            doi_candidates=[
                IdentifierCandidate(
                    "10.1234/bilingual.2022.8", "high", "first_page", 1
                )
            ],
            title_candidates=[
                TitleCandidate(
                    "中国生物医学人工智能研究",
                    "high",
                    "page_top",
                    1,
                    "first_page_title",
                ),
                TitleCandidate(
                    "Artificial Intelligence Methods for Biomedical Research",
                    "high",
                    "ocr_page_top" if ocr_mode == "always" else "page_top",
                    1,
                    "bilingual_title",
                ),
            ],
            year_candidates=["2022"],
            local_metadata=local,
            pypdf_text_available=True,
            text_extraction_method="pypdf+ocr" if ocr_mode == "always" else "pypdf",
            ocr_result=(
                OcrInspection(
                    status="success",
                    combined_text=_bilingual_text(),
                    trigger_reason=ocr_trigger_reason or "",
                )
                if ocr_mode == "always"
                else None
            ),
        )

    monkeypatch.setattr(PdfService, "inspect", fake_inspect)
    final_checks: list[bool] = []
    from lam.workflows import progressive_register

    original_check = progressive_register.DailyCheckWorkflow.run

    def counted_check(self, *, dry_run=False, final_check=False):
        final_checks.append(final_check)
        return original_check(self, dry_run=dry_run, final_check=final_check)

    monkeypatch.setattr(progressive_register.DailyCheckWorkflow, "run", counted_check)
    metadata_service = _NotFoundMetadata()

    result = InboxRegisterWorkflow(
        Settings.from_root(root), metadata_service=metadata_service
    ).run(ocr_mode="auto")

    assert result.status.value == "needs_review"
    assert calls == [
        ("never", None),
        ("always", "provider_not_found_local_metadata_incomplete"),
    ]
    assert final_checks == [True]
    assert source.is_file()
    workbook = load_workbook(root / "catalogue.xlsx", read_only=True)
    catalogue = workbook["Catalogue"]
    documents = workbook["Documents"]
    headers = {cell.value: cell.column for cell in catalogue[1]}
    assert catalogue.max_row == 2
    assert catalogue.cell(2, headers["title"]).value == "中国生物医学人工智能研究"
    assert catalogue.cell(2, headers["authors"]).value == "Zhang San; Li Si"
    assert catalogue.cell(2, headers["year"]).value == "2022"
    assert catalogue.cell(2, headers["journal"]).value == "中国医学科学院学报"
    assert catalogue.cell(2, headers["doi"]).value == "10.1234/bilingual.2022.8"
    assert catalogue.cell(2, headers["abstract"]).value == ENGLISH_ABSTRACT
    uncertainty = str(catalogue.cell(2, headers["uncertainty"]).value)
    assert "metadata_identity_unconfirmed" in uncertainty
    assert "local_english_title_candidate" in uncertainty
    document_headers = {cell.value: cell.column for cell in documents[1]}
    assert documents.max_row == 2
    assert documents.cell(2, document_headers["relative_path"]).value == source.relative_to(root).as_posix()
    assert documents.cell(2, document_headers["file_status"]).value == "inbox"


def test_complete_pypdf_local_fields_skip_ocr():
    inspection = PdfInspection(
        relative_path="Inbox/a.pdf",
        filename="a.pdf",
        size=1,
        mtime_ns=1,
        local_metadata={
            "authors": ("A", "B"),
            "journal": "Journal",
            "abstract": ENGLISH_ABSTRACT,
            "english_title": "A Complete English Research Title",
        },
    )
    from lam.workflows.progressive_register import ProgressiveInboxRegisterWorkflow

    assert ProgressiveInboxRegisterWorkflow._local_metadata_complete(inspection)


def test_local_fallback_does_not_overwrite_existing_user_value(
    tmp_path: Path, monkeypatch
):
    root = _empty_library(tmp_path)
    paper_uuid = "12345678-1234-4234-9234-123456789abc"
    _append_catalogue_row(
        root,
        {
            "paper_uuid": paper_uuid,
            "title": "用户保留标题",
            "source": "local_pdf",
            "uncertainty": (
                "USER_CONFIRMED: field=title; value=用户保留标题\n"
                "NEEDS_REVIEW: field=paper_identity; "
                "issue_key=metadata_identity_unconfirmed; issue=Pending"
            ),
        },
    )
    source = root / "Inbox" / "existing.pdf"
    source.write_bytes(b"%PDF-1.4\nlocal\n%%EOF\n")
    stat = source.stat()
    settings = Settings.from_root(root)
    settings.ensure_runtime_directories()
    blocker = settings.state_dir / "inbox_blockers.json"
    blocker.write_text(
        '{"version":1,"files":[{"source_path":"Inbox/existing.pdf",'
        f'"paper_uuid":"{paper_uuid}","size":{stat.st_size},'
        f'"mtime_ns":{stat.st_mtime_ns},"issue_keys":["metadata_identity_unconfirmed"]}}]}}',
        encoding="utf-8",
    )

    local = {
        "title": "机器候选标题",
        "authors": ("Zhang San", "Li Si"),
        "year": "2022",
        "journal": "中国医学科学院学报",
        "doi": "10.1234/local",
        "abstract": ENGLISH_ABSTRACT,
        "english_title": "Machine Candidate Title",
        "field_sources": {"title": "pypdf_first_page"},
        "field_confidence": {"title": "high"},
    }

    def fake_inspect(self, path, **kwargs):
        return PdfInspection(
            relative_path="Inbox/existing.pdf",
            filename=source.name,
            size=stat.st_size,
            mtime_ns=stat.st_mtime_ns,
            is_readable=True,
            page_count=1,
            title_candidates=[
                TitleCandidate("机器候选标题", "high", "page_top", 1, "local")
            ],
            local_metadata=local,
            pypdf_text_available=True,
        )

    monkeypatch.setattr(PdfService, "inspect", fake_inspect)
    InboxRegisterWorkflow(settings, metadata_service=_NotFoundMetadata()).run(
        ocr_mode="never"
    )
    workbook = load_workbook(root / "catalogue.xlsx", read_only=True)
    sheet = workbook["Catalogue"]
    headers = {cell.value: cell.column for cell in sheet[1]}
    assert sheet.cell(2, headers["title"]).value == "用户保留标题"
    assert sheet.cell(2, headers["authors"]).value == "Zhang San; Li Si"


def test_repeated_pypdf_font_warning_is_nonblocking_and_collapsed():
    warnings: list[str] = []
    handler = _PypdfWarningHandler(warnings)
    record = logging.LogRecord(
        "pypdf", logging.WARNING, "", 1, "Multiple definitions of /Ascent", (), None
    )
    handler.emit(record)
    handler.emit(record)
    assert warnings == ["pypdf_font_dictionary_warning"]
