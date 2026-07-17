from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import load_workbook

from lam.exceptions import CatalogueError
from lam.services.catalogue_service import CatalogueService


def test_targeted_write_preserves_sheet_style_and_extra_content(current_library_factory):
    root = current_library_factory(
        [
            {
                "id": "P1",
                "title": "Example",
                "manual_tags": "keep",
                "topic_folder": "Topic_A",
                "pdf_status": "registered",
                "pdf_filename": "paper.pdf",
                "pdf_relative_path": "Registered/paper.pdf",
                "notes": "user note",
                "custom_column": "custom",
            }
        ]
    )
    service = CatalogueService(root / "catalogue.xlsx")
    record = service.load()[0]
    service.update_fields(record, {"authors": "Example Author"})
    backup = service.save_atomic()

    assert backup and backup.exists()
    workbook = load_workbook(root / "catalogue.xlsx")
    sheet = workbook["Catalogue"]
    headers = {cell.value: index + 1 for index, cell in enumerate(sheet[1])}
    assert sheet.cell(2, headers["authors"]).value == "Example Author"
    assert sheet.cell(2, headers["manual_tags"]).value == "keep"
    assert sheet.cell(2, headers["notes"]).value == "user note"
    assert sheet.cell(2, headers["custom_column"]).value == "custom"
    assert sheet["A1"].font.bold is True
    assert workbook["Other sheet"]["A1"].value == "preserve me"


def test_user_controlled_field_cannot_be_overwritten(current_library_factory):
    root = current_library_factory([{"title": "Example", "topic_folder": "Topic_A"}])
    service = CatalogueService(root / "catalogue.xlsx")
    record = service.load()[0]
    with pytest.raises(CatalogueError):
        service.update_fields(record, {"topic_folder": "Topic_B"})


def test_uncertainty_is_deduplicated_and_confirmation_is_preserved(current_library_factory):
    root = current_library_factory(
        [
            {
                "id": "P1",
                "title": "Example",
                "topic_folder": "Topic_A",
                "uncertainty": (
                    "free user text\n"
                    "USER_CONFIRMED: field=topic_folder; value=Topic_A; note=keep"
                ),
            }
        ]
    )
    service = CatalogueService(root / "catalogue.xlsx")
    record = service.load()[0]
    assert service.add_uncertainty(
        record,
        "NEEDS_REVIEW:",
        "topic_folder",
        "classification questionable",
    ) is False
    assert service.add_uncertainty(
        record,
        "NEEDS_REVIEW:",
        "pdf_file",
        "PDF missing",
    ) is True
    assert service.add_uncertainty(
        record,
        "NEEDS_REVIEW:",
        "pdf_file",
        "PDF missing",
    ) is False
    value = record.get("uncertainty")
    assert "free user text" in value
    assert "USER_CONFIRMED:" in value
    assert value.count("PDF missing") == 1


def test_rapid_successive_writes_never_overwrite_a_backup(current_library_factory):
    root = current_library_factory(
        [{"title": "Example", "topic_folder": "Topic_A"}]
    )
    first = CatalogueService(root / "catalogue.xlsx")
    first_record = first.load()[0]
    first.update_fields(first_record, {"auto_tags": "first"})
    first_backup = first.save_atomic()

    second = CatalogueService(root / "catalogue.xlsx")
    second_record = second.load()[0]
    second.update_fields(second_record, {"auto_tags": "second"})
    second_backup = second.save_atomic()

    assert first_backup != second_backup
    assert first_backup.exists()
    assert second_backup.exists()
    assert len(list(root.glob("catalogue.backup.*.xlsx"))) == 2


def test_one_active_review_per_field_and_empty_confirmation_resolves_it(current_library_factory):
    root = current_library_factory([{"title": "Example", "topic_folder": "Topic_A"}])
    service = CatalogueService(root / "catalogue.xlsx")
    record = service.load()[0]
    first = service.ensure_review_blocker(
        record, "pdf_file", "PDF missing", issue_key="missing"
    )
    second = service.ensure_review_blocker(
        record, "pdf_file", "A newer missing message", issue_key="missing_new"
    )
    assert first == "added"
    assert second == "existing"
    assert str(record.get("uncertainty")).count("NEEDS_REVIEW:") == 1

    service.update_fields(
        record,
        {
            "uncertainty": (
                str(record.get("uncertainty"))
                + "\nUSER_CONFIRMED: field=pdf_file; value="
            )
        },
    )
    outcome = service.ensure_review_blocker(
        record, "pdf_file", "PDF missing", issue_key="missing"
    )
    assert outcome == "confirmed"
    assert "NEEDS_REVIEW:" not in str(record.get("uncertainty"))
    assert "USER_CONFIRMED:" in str(record.get("uncertainty"))


def test_user_cleared_review_is_remembered_by_snapshot(current_library_factory):
    root = current_library_factory([{"title": "Example", "topic_folder": "Topic_A"}])
    first = CatalogueService(root / "catalogue.xlsx")
    record = first.load()[0]
    first.ensure_review_blocker(record, "pdf_file", "PDF missing", issue_key="missing")
    previous = first.snapshot_payload()
    first.save_atomic()

    workbook = load_workbook(root / "catalogue.xlsx")
    headers = {cell.value: index + 1 for index, cell in enumerate(workbook["Catalogue"][1])}
    workbook["Catalogue"].cell(2, headers["uncertainty"]).value = None
    workbook.save(root / "catalogue.xlsx")

    second = CatalogueService(root / "catalogue.xlsx")
    record = second.load()[0]
    second.configure_review_state(previous)
    outcome = second.ensure_review_blocker(
        record, "pdf_file", "PDF missing", issue_key="missing"
    )
    assert outcome == "cleared"
    assert not record.get("uncertainty")


def test_blank_bibliographic_field_can_be_filled_but_not_overwritten(current_library_factory):
    root = current_library_factory([{"title": "", "topic_folder": "Topic_A"}])
    service = CatalogueService(root / "catalogue.xlsx")
    record = service.load()[0]
    service.update_fields(record, {"title": "Confirmed Title"})
    with pytest.raises(CatalogueError, match="non-empty bibliographic"):
        service.update_fields(record, {"title": "Conflicting Title"})


def test_normal_save_never_prunes_existing_backups(current_library_factory):
    root = current_library_factory([{"title": "", "topic_folder": "Topic_A"}])
    existing = []
    for index in range(12):
        path = root / f"catalogue.backup.202501{index + 1:02d}-010101.xlsx"
        path.write_bytes((root / "catalogue.xlsx").read_bytes())
        existing.append(path)
    service = CatalogueService(root / "catalogue.xlsx")
    record = service.load()[0]
    service.update_fields(record, {"title": "Example"})

    backup = service.save_atomic()

    assert backup is not None and backup.exists()
    reloaded = CatalogueService(root / "catalogue.xlsx")
    assert reloaded.load()[0].get("title") == "Example"
    assert all(path.exists() for path in existing)
    assert service.maintenance_actions == []
