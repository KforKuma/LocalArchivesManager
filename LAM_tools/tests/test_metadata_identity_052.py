from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook, load_workbook

from lam.config import Settings
from lam.models import (
    MetadataLookupRequest,
    MetadataLookupResult,
    MetadataLookupStatus,
    MetadataRecord,
    CatalogueRecord,
    DocumentRecord,
    ProviderResult,
    ProviderStatus,
)
from lam.schema import CATALOGUE_FIELDS, DOCUMENT_FIELDS
from lam.workflows.metadata_query import MetadataQueryWorkflow
from lam.workflows.inbox_register import InboxRegisterWorkflow
from lam.services.matching_service import MatchingService
from conftest import write_text_pdf


class _MetadataService:
    def __init__(self, record: MetadataRecord):
        self.record = record

    def lookup(self, request):
        provider = ProviderResult(
            "pubmed",
            ProviderStatus.FOUND,
            "pmid",
            self.record.pmid,
            records=[self.record],
        )
        return MetadataLookupResult(
            MetadataLookupStatus.FOUND,
            records=[self.record.to_dict()],
            best_record=self.record.to_dict(),
            confidence="exact_identifier",
            providers_used=["pubmed"],
            provider_results=[provider],
            selection_reason="exact test identity",
        )


class _NotFoundService:
    def lookup(self, request):
        return MetadataLookupResult(
            MetadataLookupStatus.NOT_FOUND,
            selection_reason="no provider match",
        )


def _empty_library(tmp_path: Path) -> Path:
    root = tmp_path / "library"
    for name in ("Inbox", "Registered", "Topics"):
        (root / name).mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    catalogue = workbook.active
    catalogue.title = "Catalogue"
    catalogue.append(CATALOGUE_FIELDS)
    documents = workbook.create_sheet("Documents")
    documents.append(DOCUMENT_FIELDS)
    workbook.save(root / "catalogue.xlsx")
    return root


def test_workflow2_new_record_uses_uuid_and_external_identifier_columns(tmp_path: Path):
    root = _empty_library(tmp_path)
    metadata = MetadataRecord(
        canonical_id="PMID:12345678",
        title="New Biomedical Paper",
        authors=["Alice Smith"],
        year="2025",
        journal="Biomedical Journal",
        journal_abbrev="Biomed J",
        doi="10.1000/new",
        pmid="12345678",
        source=["pubmed"],
    )

    result = MetadataQueryWorkflow(
        Settings.from_root(root), _MetadataService(metadata)
    ).run(MetadataLookupRequest(pmid="12345678"))

    assert result.status.value == "success"
    sheet = load_workbook(root / "catalogue.xlsx", read_only=True)["Catalogue"]
    headers = tuple(cell.value for cell in sheet[1])
    values = {field: sheet.cell(2, index + 1).value for index, field in enumerate(headers)}
    assert headers == CATALOGUE_FIELDS
    assert len(str(values["paper_uuid"])) == 36
    assert values["pmid"] == "12345678"
    assert values["doi"] == "10.1000/new"
    assert "id" not in headers
    journal = next(
        (root / ".library_state" / "runs").glob("*-metadata/operation_journal.json")
    )
    operation = json.loads(journal.read_text(encoding="utf-8"))["operations"][0]
    assert operation["paper_uuid"] == values["paper_uuid"]
    assert "record_uid" not in operation


def test_workflow3_provisional_identity_tracks_inbox_document(tmp_path: Path):
    root = _empty_library(tmp_path)
    source = root / "Inbox" / "unknown.pdf"
    write_text_pdf(source, ["Untitled local document"])

    result = InboxRegisterWorkflow(
        Settings.from_root(root), metadata_service=_NotFoundService()
    ).run(ocr_mode="never")

    assert result.status.value == "needs_review"
    workbook = load_workbook(root / "catalogue.xlsx", read_only=True)
    catalogue = workbook["Catalogue"]
    documents = workbook["Documents"]
    catalogue_headers = tuple(cell.value for cell in catalogue[1])
    document_headers = tuple(cell.value for cell in documents[1])
    assert catalogue_headers == CATALOGUE_FIELDS
    assert document_headers == DOCUMENT_FIELDS
    assert catalogue.max_row == 2
    paper_uuid = catalogue.cell(2, catalogue_headers.index("paper_uuid") + 1).value
    assert paper_uuid
    assert documents.max_row == 2
    assert documents.cell(2, document_headers.index("paper_uuid") + 1).value == paper_uuid
    assert documents.cell(2, document_headers.index("relative_path") + 1).value == "Inbox/unknown.pdf"
    assert documents.cell(2, document_headers.index("file_status") + 1).value == "inbox"
    assert source.is_file()


def test_matching_uses_documents_path_and_returns_paper_uuid():
    paper_uuid = "12345678-1234-4234-9234-123456789abc"
    record = CatalogueRecord(2, {"paper_uuid": paper_uuid, "title": "Example"})
    document = DocumentRecord(
        2,
        {
            "document_id": f"{paper_uuid}:main",
            "paper_uuid": paper_uuid,
            "document_type": "main",
            "filename": "paper.pdf",
            "relative_path": "Inbox/paper.pdf",
        },
    )

    match = MatchingService().match(
        [record],
        documents=[document],
        relative_path="Inbox/paper.pdf",
        filename="paper.pdf",
    )

    assert match.matched_row_id == 2
    assert match.matched_paper_uuid == paper_uuid
    assert match.method == "document_relative_path"
