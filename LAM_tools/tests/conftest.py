from __future__ import annotations

from pathlib import Path
import ctypes
import os
import socket
import uuid
import hashlib
import re

import pytest
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from pypdf import PdfWriter
from pypdf.generic import DictionaryObject, NameObject, DecodedStreamObject
from lam.schema import CATALOGUE_FIELDS, DOCUMENT_FIELDS
from fixtures.legacy.factory import create_legacy_library


PROJECT_ROOT = Path(__file__).resolve().parents[2]
REAL_LIBRARY_ROOT = PROJECT_ROOT.resolve()


def _is_within(path: Path, parent: Path) -> bool:
    try:
        path.resolve().relative_to(parent.resolve())
        return True
    except ValueError:
        return False


def _windows_elevated() -> bool:
    if os.name != "nt":
        return False
    try:
        return bool(ctypes.windll.shell32.IsUserAnAdmin())
    except Exception:
        return True


def pytest_sessionstart(session):
    """Fail before collection if pytest infrastructure can touch the real library."""
    os.environ["LAM_TESTING"] = "1"
    os.environ["LAM_REAL_LIBRARY_ROOT"] = str(REAL_LIBRARY_ROOT)
    os.environ.pop("LIBRARY_ROOT", None)
    basetemp = session.config._tmp_path_factory.getbasetemp().resolve()
    if (
        _is_within(basetemp, REAL_LIBRARY_ROOT)
        or _is_within(basetemp, PROJECT_ROOT)
    ):
        pytest.exit(
            f"Unsafe pytest basetemp inside project/real library: {basetemp}",
            returncode=10,
        )
    if _windows_elevated():
        pytest.exit(
            "LAM tests refuse an elevated Windows token by default",
            returncode=10,
        )


@pytest.fixture(autouse=True)
def isolate_default_test_environment(monkeypatch, request):
    """Default tests never use real roots, dotenv, network, or OCR downloads."""
    monkeypatch.setenv("OCR_ENABLED", "false")
    monkeypatch.setenv("OCR_DOWNLOAD_ENABLED", "false")
    monkeypatch.delenv("LIBRARY_ROOT", raising=False)
    network_markers = ("live", "live_provider", "live_download", "ocr_live")
    if any(request.node.get_closest_marker(name) for name in network_markers):
        return

    def blocked(*_args, **_kwargs):
        raise RuntimeError("Default LAM tests may not open network sockets")

    monkeypatch.setattr(socket, "create_connection", blocked)
    monkeypatch.setattr(socket.socket, "connect", blocked)
    monkeypatch.setattr(socket.socket, "connect_ex", blocked)


def write_pdf(path: Path, marker: bytes = b"paper") -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_bytes(b"%PDF-1.4\n" + marker + b"\n%%EOF\n")


def write_text_pdf(
    path: Path,
    pages: list[str],
    *,
    metadata: dict[str, str] | None = None,
    password: str | None = None,
) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    writer = PdfWriter()
    font = DictionaryObject(
        {
            NameObject("/Type"): NameObject("/Font"),
            NameObject("/Subtype"): NameObject("/Type1"),
            NameObject("/BaseFont"): NameObject("/Helvetica"),
        }
    )
    font_ref = writer._add_object(font)
    for text in pages:
        page = writer.add_blank_page(width=612, height=792)
        page[NameObject("/Resources")] = DictionaryObject(
            {
                NameObject("/Font"): DictionaryObject(
                    {NameObject("/F1"): font_ref}
                )
            }
        )
        commands = [b"BT /F1 12 Tf 72 720 Td 16 TL"]
        for index, line in enumerate(text.splitlines() or [""]):
            escaped = (
                line.replace("\\", "\\\\")
                .replace("(", "\\(")
                .replace(")", "\\)")
                .encode("latin-1", errors="replace")
            )
            if index:
                commands.append(b"T*")
            commands.append(b"(" + escaped + b") Tj")
        commands.append(b"ET")
        stream = DecodedStreamObject()
        stream.set_data(b"\n".join(commands))
        page[NameObject("/Contents")] = writer._add_object(stream)
    if metadata:
        writer.add_metadata(metadata)
    if password:
        writer.encrypt(password)
    with path.open("wb") as handle:
        writer.write(handle)


@pytest.fixture
def legacy_library_factory(tmp_path: Path):
    def factory(rows: list[dict[str, object]], files: dict[str, bytes] | None = None) -> Path:
        return create_legacy_library(tmp_path / "library", rows, files)

    return factory


@pytest.fixture
def current_library_factory(tmp_path: Path):
    """Create a strict current-schema library for ordinary workflow/CLI tests."""
    def factory(
        rows: list[dict[str, object]] | None = None,
        documents: list[dict[str, object]] | None = None,
        files: dict[str, bytes] | None = None,
    ) -> Path:
        root = tmp_path / "current-library"
        root.mkdir()
        for name in ("Inbox", "Registered", "Topics"):
            (root / name).mkdir()
        workbook = Workbook()
        catalogue = workbook.active
        catalogue.title = "Catalogue"
        supplied_rows = rows or []
        legacy_only = {"id", "record_uid", "pdf_status", "pdf_filename", "pdf_relative_path"}
        extra_fields = sorted(
            {
                key
                for row in supplied_rows
                for key in row
                if key not in CATALOGUE_FIELDS and key not in legacy_only
            }
        )
        catalogue_headers = (*CATALOGUE_FIELDS, *extra_fields)
        catalogue.append(catalogue_headers)
        for cell in catalogue[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E78")
        normalized_rows: list[dict[str, object]] = []
        for index, row in enumerate(supplied_rows, start=1):
            values = dict(row)
            values.setdefault(
                "paper_uuid",
                str(uuid.UUID(f"00000000-0000-4000-8000-{index:012x}")),
            )
            normalized_rows.append(values)
            catalogue.append([values.get(field) for field in catalogue_headers])
        document_sheet = workbook.create_sheet("Documents")
        document_sheet.append(DOCUMENT_FIELDS)
        for document in documents or []:
            document_sheet.append([document.get(field) for field in DOCUMENT_FIELDS])
        notes = workbook.create_sheet("Other sheet")
        notes["A1"] = "preserve me"
        workbook.save(root / "catalogue.xlsx")
        for relative, marker in (files or {}).items():
            write_pdf(root / relative, marker)
        return root

    return factory


@pytest.fixture
def library_factory(current_library_factory):
    """Backward-friendly input adapter that always emits the current schema."""
    local_id = re.compile(r"^LOCAL:([0-9a-f-]{36})$", re.I)

    def factory(
        rows: list[dict[str, object]], files: dict[str, bytes] | None = None
    ) -> Path:
        normalized_rows: list[dict[str, object]] = []
        documents: list[dict[str, object]] = []
        for index, raw in enumerate(rows, start=1):
            values = dict(raw)
            candidate = str(values.get("paper_uuid") or values.get("record_uid") or "")
            legacy_id = str(values.get("id") or "")
            match = local_id.fullmatch(legacy_id)
            if not candidate and match:
                candidate = match.group(1)
            try:
                parsed = uuid.UUID(candidate)
                if parsed.version != 4:
                    raise ValueError
                paper_uuid = str(parsed)
            except ValueError:
                paper_uuid = str(uuid.UUID(f"00000000-0000-4000-8000-{index:012x}"))
            values["paper_uuid"] = paper_uuid
            if legacy_id.upper().startswith("PMID:") and not values.get("pmid"):
                values["pmid"] = legacy_id.split(":", 1)[1]
            if legacy_id.upper().startswith("DOI:") and not values.get("doi"):
                values["doi"] = legacy_id.split(":", 1)[1]
            normalized_rows.append(values)

            relative = str(values.get("pdf_relative_path") or "").replace("\\", "/")
            filename = str(values.get("pdf_filename") or "")
            # Inbox files are not registered documents yet. Workflow 3 creates
            # the Documents row only after identity is accepted and the file
            # is moved to Registered/.
            legacy_status = str(values.get("pdf_status") or "").casefold()
            has_managed_location = bool(relative) and not relative.casefold().startswith(
                "inbox/"
            )
            status_is_registered = legacy_status in {
                "registered",
                "filed",
                "missing",
                "unclear",
            }
            if (relative or filename) and (
                has_managed_location or status_is_registered
            ):
                relative = relative or f"Registered/{filename}"
                filename = filename or Path(relative).name
                marker = (files or {}).get(relative)
                digest = ""
                if marker is not None:
                    content = b"%PDF-1.4\n" + marker + b"\n%%EOF\n"
                    digest = hashlib.sha256(content).hexdigest()
                documents.append(
                    {
                        "document_id": f"{paper_uuid}:main",
                        "paper_uuid": paper_uuid,
                        "document_type": "main",
                        "filename": filename,
                        "relative_path": relative,
                        "extension": Path(filename).suffix,
                        "sha256": digest,
                        "file_status": values.get("pdf_status") or "unclear",
                        "source": values.get("source") or "",
                        "date_added": values.get("date_added") or "",
                        "date_updated": values.get("date_updated") or "",
                    }
                )
        return current_library_factory(normalized_rows, documents, files)

    return factory
