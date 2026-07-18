from __future__ import annotations

import json
import uuid
from pathlib import Path

from openpyxl import Workbook, load_workbook

from lam.config import Settings
from lam.schema import CATALOGUE_060_FIELDS, CATALOGUE_FIELDS, DOCUMENT_FIELDS
from lam.workflows.migration import MigrationWorkflow


def _old_library(tmp_path: Path) -> tuple[Path, dict[str, str]]:
    root = tmp_path / "library"
    for name in ("Inbox", "Registered", "Topics"):
        (root / name).mkdir(parents=True, exist_ok=True)
    ids = {
        name: str(uuid.uuid4())
        for name in ("pdf", "reference", "legacy")
    }
    workbook = Workbook()
    catalogue = workbook.active
    catalogue.title = "Catalogue"
    catalogue.append(CATALOGUE_060_FIELDS)
    rows = (
        {
            "paper_uuid": ids["pdf"],
            "title": "Paper with PDF",
            "uncertainty": "USER_CONFIRMED: field=title; value=Paper with PDF",
        },
        {"paper_uuid": ids["reference"], "title": "Reference-only paper"},
        {"paper_uuid": ids["legacy"], "title": "Historical metadata"},
    )
    for row in rows:
        catalogue.append([row.get(field) for field in CATALOGUE_060_FIELDS])
    documents = workbook.create_sheet("Documents")
    documents.append(DOCUMENT_FIELDS)
    documents.append(
        [
            {
                "document_id": f"{ids['pdf']}:main",
                "paper_uuid": ids["pdf"],
                "document_type": "main",
                "filename": "paper.pdf",
                "relative_path": "Registered/paper.pdf",
                "extension": ".pdf",
                "file_status": "registered",
            }.get(field)
            for field in DOCUMENT_FIELDS
        ]
    )
    notes = workbook.create_sheet("User Notes")
    notes["A1"] = "preserve"
    workbook.save(root / "catalogue.xlsx")
    (root / "Registered" / "paper.pdf").write_bytes(b"%PDF-1.4\n%%EOF\n")

    receipts = root / ".library_state" / "imports" / "reference_text"
    receipts.mkdir(parents=True)
    (receipts / "receipt.json").write_text(
        json.dumps(
            {
                "schema_version": 1,
                "resolutions": [
                    {
                        "status": "registered_new",
                        "paper_uuid": ids["reference"],
                    }
                ],
            }
        ),
        encoding="utf-8",
    )
    return root, ids


def test_schema_migration_dry_run_is_read_only(tmp_path: Path):
    root, _ = _old_library(tmp_path)
    before = (root / "catalogue.xlsx").read_bytes()

    result = MigrationWorkflow(Settings.from_root(root)).schema(dry_run=True)

    assert result.status.value == "needs_review"
    assert result.counts["document_required"] == 1
    assert result.counts["document_optional"] == 1
    assert result.counts["document_unknown"] == 1
    assert (root / "catalogue.xlsx").read_bytes() == before
    assert not list(root.glob("catalogue.backup.*.xlsx"))


def test_schema_migration_apply_preserves_entities_and_infers_semantics(tmp_path: Path):
    root, ids = _old_library(tmp_path)

    result = MigrationWorkflow(Settings.from_root(root)).schema(dry_run=False)

    assert result.status.value in {"success", "needs_review"}
    workbook = load_workbook(root / "catalogue.xlsx", read_only=True)
    catalogue = workbook["Catalogue"]
    assert tuple(cell.value for cell in catalogue[1]) == CATALOGUE_FIELDS
    headers = {cell.value: index + 1 for index, cell in enumerate(catalogue[1])}
    rows = {
        catalogue.cell(row, headers["paper_uuid"]).value: {
            field: catalogue.cell(row, column).value
            for field, column in headers.items()
        }
        for row in range(2, catalogue.max_row + 1)
    }
    assert rows[ids["pdf"]]["record_origin"] == "pdf"
    assert rows[ids["pdf"]]["document_expectation"] == "required"
    assert rows[ids["pdf"]]["uncertainty"].startswith("USER_CONFIRMED:")
    assert rows[ids["reference"]]["record_origin"] == "reference_text"
    assert rows[ids["reference"]]["document_expectation"] == "optional"
    assert rows[ids["legacy"]]["record_origin"] == "legacy"
    assert rows[ids["legacy"]]["document_expectation"] == "unknown"
    assert "legacy_document_expectation_unknown" in rows[ids["legacy"]]["uncertainty"]
    assert workbook["User Notes"]["A1"].value == "preserve"
    assert list(root.glob("catalogue.backup.*.xlsx"))


def test_schema_migration_skips_current_schema(current_library_factory):
    root = current_library_factory()

    result = MigrationWorkflow(Settings.from_root(root)).schema(dry_run=True)

    assert result.status.value == "no_changes"
    assert result.skipped == [{"reason": "already_current_schema"}]
