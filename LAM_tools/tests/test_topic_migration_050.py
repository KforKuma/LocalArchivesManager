from __future__ import annotations

import builtins
import json
from pathlib import Path

import pytest
from openpyxl import load_workbook

from lam.config import Settings
from lam.exceptions import CatalogueError
from lam.models import WorkflowStatus
from lam.services.catalogue_service import CatalogueService
from lam.workflows.topic_migration import TopicMigrationWorkflow


def _legacy_library(library_factory):
    return library_factory(
        [
            {
                "id": "P1",
                "title": "Legacy",
                "topic_folder": "Topic_A",
                "pdf_status": "filed",
                "pdf_filename": "paper.pdf",
                "pdf_relative_path": "Topic_A/paper.pdf",
            }
        ],
        {"Topic_A/paper.pdf": b"paper"},
    )


def test_migrate_topics_dry_run_changes_nothing(library_factory):
    root = _legacy_library(library_factory)
    settings = Settings.from_root(root)
    settings.ensure_runtime_directories()
    result = TopicMigrationWorkflow(settings).run(dry_run=True)
    assert result.status == WorkflowStatus.SUCCESS
    assert (root / "Topic_A" / "paper.pdf").exists()
    assert not (root / "Topics" / "Topic_A").exists()
    assert not list(root.glob("catalogue.backup.*.xlsx"))
    assert not list((root / ".library_state" / "runs").glob("*"))
    workbook = load_workbook(root / "catalogue.xlsx")
    documents = workbook["Documents"]
    headers = {cell.value: cell.column for cell in documents[1]}
    assert documents.cell(2, headers["relative_path"]).value == "Topic_A/paper.pdf"


def test_migrate_topics_moves_directory_and_updates_catalogue(library_factory):
    root = _legacy_library(library_factory)
    settings = Settings.from_root(root)
    settings.ensure_runtime_directories()
    result = TopicMigrationWorkflow(settings).run(dry_run=False)
    assert result.status == WorkflowStatus.SUCCESS
    assert not (root / "Topic_A").exists()
    assert (root / "Topics" / "Topic_A" / "paper.pdf").exists()
    workbook = load_workbook(root / "catalogue.xlsx")
    sheet = workbook["Catalogue"]
    catalogue_headers = {cell.value: cell.column for cell in sheet[1]}
    assert sheet.cell(2, catalogue_headers["topic_folder"]).value == "Topic_A"
    documents = workbook["Documents"]
    document_headers = {cell.value: cell.column for cell in documents[1]}
    assert documents.cell(2, document_headers["relative_path"]).value == (
        "Topics/Topic_A/paper.pdf"
    )
    journal = next(
        (root / ".library_state" / "runs").glob(
            "*-migrate-topics/operation_journal.json"
        )
    )
    assert json.loads(journal.read_text(encoding="utf-8"))["status"] == "final_check_committed"


def test_migrate_topics_excludes_management_and_reports_unknown(library_factory):
    root = _legacy_library(library_factory)
    (root / "scripts" / "nested").mkdir(parents=True)
    (root / "UnknownDirectory").mkdir()
    settings = Settings.from_root(root)
    settings.ensure_runtime_directories()
    result = TopicMigrationWorkflow(settings).run(dry_run=True)
    unknown = {item.get("directory") for item in result.skipped}
    assert "UnknownDirectory" in unknown
    assert "scripts" not in unknown
    assert (root / "scripts").exists()


def test_migrate_topics_target_collision_never_overwrites(library_factory):
    root = _legacy_library(library_factory)
    target = root / "Topics" / "Topic_A"
    target.mkdir()
    (target / "paper.pdf").write_bytes(b"different")
    settings = Settings.from_root(root)
    settings.ensure_runtime_directories()
    result = TopicMigrationWorkflow(settings).run(dry_run=False)
    assert result.status == WorkflowStatus.NEEDS_REVIEW
    assert (root / "Topic_A" / "paper.pdf").exists()
    assert (target / "paper.pdf").read_bytes() == b"different"


def test_summary_moves_with_topic_without_being_read(library_factory, monkeypatch):
    root = _legacy_library(library_factory)
    summary = root / "Topic_A" / "summary.md"
    summary.write_text("user content", encoding="utf-8")
    real_open = builtins.open
    real_read_text = Path.read_text

    def guarded_open(file, mode="r", *args, **kwargs):
        if str(file).casefold().endswith("summary.md") and "r" in mode:
            raise AssertionError("summary.md must not be read")
        return real_open(file, mode, *args, **kwargs)

    monkeypatch.setattr(builtins, "open", guarded_open)

    def guarded_read_text(self, *args, **kwargs):
        if self.name.casefold() == "summary.md":
            raise AssertionError("summary.md must not be read")
        return real_read_text(self, *args, **kwargs)

    monkeypatch.setattr(Path, "read_text", guarded_read_text)
    settings = Settings.from_root(root)
    settings.ensure_runtime_directories()
    result = TopicMigrationWorkflow(settings).run(dry_run=False)
    assert result.status == WorkflowStatus.SUCCESS
    assert (root / "Topics" / "Topic_A" / "summary.md").exists()


def test_migration_recovers_catalogue_after_directory_already_moved(library_factory):
    root = _legacy_library(library_factory)
    source = root / "Topic_A"
    target = root / "Topics" / "Topic_A"
    source.rename(target)
    settings = Settings.from_root(root)
    settings.ensure_runtime_directories()
    result = TopicMigrationWorkflow(settings).run(dry_run=False)
    assert result.status == WorkflowStatus.SUCCESS
    workbook = load_workbook(root / "catalogue.xlsx")
    documents = workbook["Documents"]
    headers = {cell.value: cell.column for cell in documents[1]}
    assert documents.cell(2, headers["relative_path"]).value == (
        "Topics/Topic_A/paper.pdf"
    )
    assert "Topic_A" in result.details["recovery_roots"]


def test_catalogue_failure_rolls_back_topic_directory(library_factory, monkeypatch):
    root = _legacy_library(library_factory)
    settings = Settings.from_root(root)
    settings.ensure_runtime_directories()

    def fail_save(self):
        raise CatalogueError("simulated catalogue failure")

    monkeypatch.setattr(CatalogueService, "save_atomic", fail_save)
    with pytest.raises(CatalogueError, match="simulated"):
        TopicMigrationWorkflow(settings).run(dry_run=False)
    assert (root / "Topic_A" / "paper.pdf").exists()
    assert not (root / "Topics" / "Topic_A").exists()


def test_topics_prefix_in_topic_folder_is_normalized(library_factory):
    root = library_factory(
        [
            {
                "id": "P1",
                "title": "Legacy",
                "topic_folder": "Topics/Topic_A",
                "pdf_status": "filed",
                "pdf_filename": "paper.pdf",
                "pdf_relative_path": "Topic_A/paper.pdf",
            }
        ],
        {"Topic_A/paper.pdf": b"paper"},
    )
    settings = Settings.from_root(root)
    settings.ensure_runtime_directories()
    TopicMigrationWorkflow(settings).run(dry_run=False)
    workbook = load_workbook(root / "catalogue.xlsx")
    sheet = workbook["Catalogue"]
    headers = {cell.value: cell.column for cell in sheet[1]}
    assert sheet.cell(2, headers["topic_folder"]).value == "Topic_A"


def test_registered_filename_inside_root_directory_is_a_migration_candidate(
    library_factory,
):
    root = library_factory(
        [
            {
                "id": "P1",
                "title": "Located by registered filename",
                "pdf_status": "unclear",
                "pdf_filename": "paper.pdf",
                "pdf_relative_path": "",
            }
        ],
        {"LegacyByFilename/nested/paper.pdf": b"paper"},
    )
    settings = Settings.from_root(root)
    settings.ensure_runtime_directories()
    result = TopicMigrationWorkflow(settings).run(dry_run=False)
    assert result.status in {WorkflowStatus.SUCCESS, WorkflowStatus.NEEDS_REVIEW}
    assert (root / "Topics" / "LegacyByFilename" / "nested" / "paper.pdf").exists()
    workbook = load_workbook(root / "catalogue.xlsx")
    documents = workbook["Documents"]
    headers = {cell.value: cell.column for cell in documents[1]}
    assert documents.cell(2, headers["relative_path"]).value == (
        "Topics/LegacyByFilename/nested/paper.pdf"
    )
