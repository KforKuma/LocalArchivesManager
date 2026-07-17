from __future__ import annotations

import json

from openpyxl import load_workbook

from lam.config import Settings
from lam.models import WorkflowStatus
from lam.workflows.catalogue_filing import CatalogueFilingWorkflow
from lam.workflows.daily_check import DailyCheckWorkflow


def test_registered_pdf_is_filed_and_final_checked(library_factory):
    root = library_factory(
        [
            {
                "id": "P1",
                "title": "Example",
                "topic_folder": "Topic_A",
                "pdf_status": "registered",
                "pdf_filename": "paper.pdf",
                "pdf_relative_path": "Registered/paper.pdf",
            }
        ],
        {"Registered/paper.pdf": b"paper"},
    )
    settings = Settings.from_root(root)
    settings.ensure_runtime_directories()
    DailyCheckWorkflow(settings).run()
    result = CatalogueFilingWorkflow(settings).run()
    workbook = load_workbook(root / "catalogue.xlsx")
    sheet = workbook["Documents"]
    headers = {cell.value: cell.column for cell in sheet[1]}
    assert result.status == WorkflowStatus.SUCCESS
    assert not (root / "Registered" / "paper.pdf").exists()
    assert (root / "Topics" / "Topic_A" / "paper.pdf").exists()
    assert sheet.cell(2, headers["file_status"]).value == "filed"
    assert sheet.cell(2, headers["relative_path"]).value == "Topics/Topic_A/paper.pdf"
    assert result.details["final_check"]["status"] in {"success", "no_changes"}


def test_dry_run_plans_move_without_changes(library_factory):
    root = library_factory(
        [
            {
                "id": "P1",
                "title": "Example",
                "topic_folder": "Topic_A",
                "pdf_status": "registered",
                "pdf_filename": "paper.pdf",
                "pdf_relative_path": "Registered/paper.pdf",
            }
        ],
        {"Registered/paper.pdf": b"paper"},
    )
    settings = Settings.from_root(root)
    settings.ensure_runtime_directories()
    result = CatalogueFilingWorkflow(settings).run(dry_run=True)
    assert result.status == WorkflowStatus.SUCCESS
    assert result.completed[0]["action"] == "would_file_from_registered"
    assert (root / "Registered" / "paper.pdf").exists()
    assert not (root / "Topics" / "Topic_A").exists()
    assert not list(root.glob("catalogue.backup.*.xlsx"))
    assert not (root / ".library_state" / "file_manifest.json").exists()


def test_unclassified_and_collision_are_left_in_place(library_factory):
    root = library_factory(
        [
            {
                "id": "P1",
                "title": "Unclassified",
                "topic_folder": "Unclassified",
                "pdf_status": "registered",
                "pdf_filename": "one.pdf",
                "pdf_relative_path": "Registered/one.pdf",
            },
            {
                "id": "P2",
                "title": "Collision",
                "topic_folder": "Topic_A",
                "pdf_status": "registered",
                "pdf_filename": "two.pdf",
                "pdf_relative_path": "Registered/two.pdf",
            },
        ],
        {
            "Registered/one.pdf": b"one",
            "Registered/two.pdf": b"source",
            "Topics/Topic_A/two.pdf": b"different",
        },
    )
    settings = Settings.from_root(root)
    settings.ensure_runtime_directories()
    result = CatalogueFilingWorkflow(settings).run()
    assert result.status == WorkflowStatus.NEEDS_REVIEW
    assert (root / "Registered" / "one.pdf").exists()
    assert (root / "Registered" / "two.pdf").exists()
    assert (root / "Topics" / "Topic_A" / "two.pdf").read_bytes().find(b"different") >= 0


def test_inbox_pdf_is_never_filed_by_workflow4(library_factory):
    root = library_factory(
        [
            {
                "id": "P1",
                "title": "Inbox paper",
                "topic_folder": "Topic_A",
                "pdf_status": "inbox",
                "pdf_filename": "paper.pdf",
                "pdf_relative_path": "Inbox/paper.pdf",
            }
        ],
        {"Inbox/paper.pdf": b"paper"},
    )
    settings = Settings.from_root(root)
    settings.ensure_runtime_directories()
    result = CatalogueFilingWorkflow(settings).run(dry_run=True)
    assert result.status == WorkflowStatus.NEEDS_REVIEW
    assert not result.completed
    assert (root / "Inbox" / "paper.pdf").exists()
    assert not (root / "Topics" / "Topic_A").exists()


def test_final_check_needs_review_overrides_no_changes(library_factory):
    root = library_factory(
        [
            {
                "id": "P1",
                "title": "Missing filed paper",
                "topic_folder": "Topic_A",
                "pdf_status": "filed",
                "pdf_filename": "missing.pdf",
                "pdf_relative_path": "Topics/Topic_A/missing.pdf",
            }
        ]
    )
    (root / "Topics" / "Topic_A").mkdir()
    settings = Settings.from_root(root)
    settings.ensure_runtime_directories()
    result = CatalogueFilingWorkflow(settings).run()
    assert result.details["final_check"]["status"] == "needs_review"
    assert result.status == WorkflowStatus.NEEDS_REVIEW
    assert any(
        item.get("issue")
        in {"expected_pdf_missing", "source_missing", "document_file_missing"}
        for item in result.needs_review
    )


def test_filed_pdf_is_refiled_after_topic_folder_change(library_factory):
    root = library_factory(
        [
            {
                "id": "P1",
                "title": "Reclassified",
                "topic_folder": "Topic_B",
                "pdf_status": "filed",
                "pdf_filename": "paper.pdf",
                "pdf_relative_path": "Topics/Topic_A/paper.pdf",
            }
        ],
        {"Topics/Topic_A/paper.pdf": b"same"},
    )
    settings = Settings.from_root(root)
    settings.ensure_runtime_directories()
    DailyCheckWorkflow(settings).run()
    result = CatalogueFilingWorkflow(settings).run()
    assert result.status == WorkflowStatus.SUCCESS
    assert not (root / "Topics" / "Topic_A" / "paper.pdf").exists()
    assert (root / "Topics" / "Topic_B" / "paper.pdf").exists()
    assert not (root / "Topics" / "Topic_A").exists()
    assert any(item.get("action") == "refiled_from_topic" for item in result.completed)
    assert result.details["removed_empty_directories"] == ["Topics/Topic_A"]
    workbook = load_workbook(root / "catalogue.xlsx")
    documents = workbook["Documents"]
    headers = {cell.value: cell.column for cell in documents[1]}
    assert (
        documents.cell(2, headers["relative_path"]).value
        == "Topics/Topic_B/paper.pdf"
    )
    journal = next((root / ".library_state" / "runs").glob("*-filing/operation_journal.json"))
    payload = json.loads(journal.read_text(encoding="utf-8"))
    assert payload["status"] == "final_check_committed"
    assert payload["operations"][0]["old_directory_removed"] is True
    final_report = json.loads(
        open(result.details["final_check"]["report"], encoding="utf-8").read()
    )
    diff_types = [item["diff_type"] for item in final_report["details"]["file_diffs"]]
    assert "expected_move_or_rename" in diff_types
    assert "possible_collision" not in diff_types
    assert "quick_hash_candidate" not in diff_types


def test_refile_keeps_nonempty_old_topic_directory(library_factory):
    root = library_factory(
        [
            {
                "id": "P1",
                "title": "Reclassified",
                "topic_folder": "Topic_B",
                "pdf_status": "filed",
                "pdf_filename": "paper.pdf",
                "pdf_relative_path": "Topics/Topic_A/paper.pdf",
            }
        ],
        {"Topics/Topic_A/paper.pdf": b"same"},
    )
    summary = root / "Topics" / "Topic_A" / "summary.md"
    summary.write_text("preserve", encoding="utf-8")
    settings = Settings.from_root(root)
    settings.ensure_runtime_directories()
    result = CatalogueFilingWorkflow(settings).run()
    assert result.status == WorkflowStatus.SUCCESS
    assert summary.exists()
    assert (root / "Topics" / "Topic_A").exists()
    assert result.details["removed_empty_directories"] == []


def test_refile_target_collision_does_not_overwrite(library_factory):
    root = library_factory(
        [
            {
                "id": "P1",
                "title": "Reclassified",
                "topic_folder": "Topic_B",
                "pdf_status": "filed",
                "pdf_filename": "paper.pdf",
                "pdf_relative_path": "Topics/Topic_A/paper.pdf",
            }
        ],
        {
            "Topics/Topic_A/paper.pdf": b"source",
            "Topics/Topic_B/paper.pdf": b"different",
        },
    )
    settings = Settings.from_root(root)
    settings.ensure_runtime_directories()
    result = CatalogueFilingWorkflow(settings).run()
    assert result.status == WorkflowStatus.NEEDS_REVIEW
    assert (root / "Topics" / "Topic_A" / "paper.pdf").exists()
    assert (root / "Topics" / "Topic_B" / "paper.pdf").exists()
    assert any(item.get("issue") == "target_collision" for item in result.needs_review)


def test_workflow4_runs_final_check_once(library_factory, monkeypatch):
    root = library_factory(
        [
            {
                "id": "P1",
                "title": "Example",
                "topic_folder": "Topic_A",
                "pdf_status": "registered",
                "pdf_filename": "paper.pdf",
                "pdf_relative_path": "Registered/paper.pdf",
            }
        ],
        {"Registered/paper.pdf": b"paper"},
    )
    from lam.workflows import catalogue_filing

    calls = 0
    original = catalogue_filing.DailyCheckWorkflow.run

    def counted(self, *args, **kwargs):
        nonlocal calls
        calls += 1
        return original(self, *args, **kwargs)

    monkeypatch.setattr(catalogue_filing.DailyCheckWorkflow, "run", counted)
    settings = Settings.from_root(root)
    settings.ensure_runtime_directories()
    CatalogueFilingWorkflow(settings).run()
    assert calls == 1


def test_workflow4_does_not_inspect_pdf_or_call_metadata_workflow(
    library_factory, monkeypatch
):
    root = library_factory(
        [
            {
                "id": "P1",
                "title": "Example",
                "topic_folder": "Topic_A",
                "pdf_status": "registered",
                "pdf_filename": "paper.pdf",
                "pdf_relative_path": "Registered/paper.pdf",
            }
        ],
        {"Registered/paper.pdf": b"paper"},
    )
    from lam.services.pdf_service import PdfService
    from lam.workflows.metadata_query import MetadataQueryWorkflow

    def forbidden(*args, **kwargs):
        raise AssertionError("Workflow 4 must not inspect PDFs or query metadata")

    monkeypatch.setattr(PdfService, "inspect", forbidden)
    monkeypatch.setattr(MetadataQueryWorkflow, "run", forbidden)
    settings = Settings.from_root(root)
    settings.ensure_runtime_directories()
    result = CatalogueFilingWorkflow(settings).run()
    assert result.status == WorkflowStatus.SUCCESS
    assert (root / "Topics" / "Topic_A" / "paper.pdf").exists()


def test_workflow4_supports_nested_topics(library_factory):
    root = library_factory(
        [
            {
                "id": "P1",
                "title": "Nested",
                "topic_folder": "IBD/Epithelial",
                "pdf_status": "registered",
                "pdf_filename": "paper.pdf",
                "pdf_relative_path": "Registered/paper.pdf",
            }
        ],
        {"Registered/paper.pdf": b"paper"},
    )
    settings = Settings.from_root(root)
    settings.ensure_runtime_directories()
    result = CatalogueFilingWorkflow(settings).run()
    assert result.status == WorkflowStatus.SUCCESS
    assert (root / "Topics" / "IBD" / "Epithelial" / "paper.pdf").exists()
