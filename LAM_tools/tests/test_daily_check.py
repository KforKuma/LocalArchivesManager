from __future__ import annotations

from openpyxl import load_workbook

from lam.config import Settings
from lam.models import WorkflowStatus
from lam.workflows.daily_check import DailyCheckWorkflow


def test_initial_reconciliation_then_idempotent_no_changes(library_factory):
    root = library_factory(
        [
            {
                "id": "P1",
                "title": "Example",
                "topic_folder": "Topic_A",
                "pdf_status": "registered",
                "pdf_filename": "paper.pdf",
                "pdf_relative_path": "Registered/paper.pdf",
            }
        ],
        {"Registered/paper.pdf": b"paper"},
    )
    settings = Settings.from_root(root)
    settings.ensure_runtime_directories()
    first = DailyCheckWorkflow(settings).run()
    second = DailyCheckWorkflow(settings).run()
    assert first.mode == "initial"
    assert first.status == WorkflowStatus.SUCCESS
    assert first.state_committed is True
    assert second.status == WorkflowStatus.NO_CHANGES
    assert second.changed_rows == 0
    assert second.state_committed is False


def test_missing_pdf_updates_machine_fields_only(library_factory):
    root = library_factory(
        [
            {
                "id": "P1",
                "title": "Example",
                "manual_tags": "keep tag",
                "topic_folder": "Topic_A",
                "pdf_status": "registered",
                "pdf_filename": "paper.pdf",
                "pdf_relative_path": "Registered/paper.pdf",
                "notes": "keep note",
            }
        ],
        {"Registered/paper.pdf": b"paper"},
    )
    settings = Settings.from_root(root)
    settings.ensure_runtime_directories()
    DailyCheckWorkflow(settings).run()
    (root / "Registered" / "paper.pdf").unlink()
    result = DailyCheckWorkflow(settings).run()
    workbook = load_workbook(root / "catalogue.xlsx", data_only=False)
    sheet = workbook["Catalogue"]
    assert result.status == WorkflowStatus.NEEDS_REVIEW
    assert sheet["H2"].value == "missing"
    assert sheet["F2"].value == "keep tag"
    assert sheet["G2"].value == "Topic_A"
    assert sheet["L2"].value == "keep note"
    assert "NEEDS_REVIEW: field=pdf_file" in sheet["M2"].value


def test_dry_run_does_not_create_official_snapshots_or_modify_catalogue(library_factory):
    root = library_factory(
        [
            {
                "id": "P1",
                "title": "Example",
                "topic_folder": "Topic_A",
                "pdf_status": "registered",
                "pdf_filename": "missing.pdf",
                "pdf_relative_path": "Registered/missing.pdf",
            }
        ]
    )
    settings = Settings.from_root(root)
    settings.ensure_runtime_directories()
    result = DailyCheckWorkflow(settings).run(dry_run=True)
    workbook = load_workbook(root / "catalogue.xlsx")
    assert result.needs_review
    assert workbook["Catalogue"]["H2"].value == "registered"
    assert not (root / ".library_state" / "file_manifest.json").exists()
    assert not list(root.glob("catalogue.backup.*.xlsx"))


def test_daily_check_reports_legacy_topic_without_scanning_unknown_roots(library_factory):
    root = library_factory(
        [
            {
                "id": "P1",
                "title": "Legacy",
                "topic_folder": "Topic_A",
                "pdf_status": "filed",
                "pdf_filename": "paper.pdf",
                "pdf_relative_path": "Topic_A/paper.pdf",
            }
        ],
        {
            "Topic_A/paper.pdf": b"legacy",
            "Unknown/other.pdf": b"unmanaged",
            "Topics/Managed/managed.pdf": b"managed",
        },
    )
    settings = Settings.from_root(root)
    settings.ensure_runtime_directories()
    result = DailyCheckWorkflow(settings).run(dry_run=True)
    assert any(item.get("issue") == "legacy_topic_location" for item in result.needs_review)
    assert result.counts["managed_pdfs"] == 1
    unmanaged = result.details["unmanaged_items"]
    assert {item["path"] for item in unmanaged} >= {"Topic_A", "Unknown"}


def test_filed_status_requires_topics_relative_topic_match(library_factory):
    root = library_factory(
        [
            {
                "id": "P1",
                "title": "Nested",
                "topic_folder": "IBD/Epithelial",
                "pdf_status": "filed",
                "pdf_filename": "paper.pdf",
                "pdf_relative_path": "Topics/IBD/Epithelial/paper.pdf",
            }
        ],
        {"Topics/IBD/Epithelial/paper.pdf": b"paper"},
    )
    settings = Settings.from_root(root)
    settings.ensure_runtime_directories()
    result = DailyCheckWorkflow(settings).run()
    assert result.status == WorkflowStatus.SUCCESS
    workbook = load_workbook(root / "catalogue.xlsx")
    assert workbook["Catalogue"]["H2"].value == "filed"
