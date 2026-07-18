from __future__ import annotations

import os
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from ..config import Settings
from ..exceptions import ConfigurationError, FileOperationError
from ..models import WorkflowResult
from ..schema import CATALOGUE_FIELDS, DOCUMENT_FIELDS
from ..services.report_service import ReportService, append_change_log
from .daily_check import DailyCheckWorkflow


ENV_EXAMPLE = """# LAM provider identity
NCBI_EMAIL=
NCBI_TOOL=LAM
NCBI_API_KEY=
UNPAYWALL_EMAIL=
CROSSREF_ENABLED=true
CROSSREF_EMAIL=
CROSSREF_MIN_INTERVAL_SECONDS=1.0
CROSSREF_MAX_RESULTS=10

# Document analysis orchestration
DOCUMENT_ANALYSIS_BACKEND=auto
DOCUMENT_ANALYSIS_FALLBACKS=native,easyocr
DOI_MIN_SUFFIX_ALNUM=3
DOI_MAX_LENGTH=200

# Optional OCR configuration
POPPLER_PATH=
OCR_MODEL_STORAGE_DIR=

# Production temporary workspace lifecycle
LAM_KEEP_FAILED_TEMP=false
LAM_TEMP_RETENTION_HOURS=24
"""

ENV_FIELDS = (
    "NCBI_EMAIL",
    "NCBI_TOOL",
    "NCBI_API_KEY (optional)",
    "UNPAYWALL_EMAIL",
    "CROSSREF_EMAIL (recommended)",
    "DOCUMENT_ANALYSIS_BACKEND",
    "POPPLER_PATH (when OCR is used)",
    "OCR_MODEL_STORAGE_DIR (optional)",
)


class LibraryInitWorkflow:
    """Create a strict empty LAM library without provider or PDF activity."""

    _PERMITTED_RUNTIME_CHILDREN = {
        "invocations",
        "logs",
        "reports",
        "lam.lock",
    }

    def __init__(self, settings: Settings):
        self.settings = settings

    def run(self, *, dry_run: bool) -> WorkflowResult:
        result = WorkflowResult(
            "library_init",
            dry_run=dry_run,
            mode="dry_run" if dry_run else "apply",
        )
        root = self.settings.library_root
        blockers = self._blocking_entries(root)
        if blockers:
            raise ConfigurationError(
                "init requires a missing or demonstrably empty target; found: "
                + ", ".join(blockers)
            )
        if self.settings.catalogue_path.exists():
            raise ConfigurationError(
                f"init refuses to overwrite an existing workbook: {self.settings.catalogue_path}"
            )

        planned = [
            "Inbox/",
            "Registered/",
            "Topics/",
            "Imports/ReferenceText/Processed/",
            ".library_state/",
            "catalogue.xlsx",
            "library_changes.md",
            ".env.example",
        ]
        result.details = {
            "planned_paths": planned,
            "required_configuration": list(ENV_FIELDS),
            "uses_network": False,
            "uses_ocr": False,
            "reads_pdf": False,
        }
        if dry_run:
            result.completed.extend(
                {"action": "would_create", "path": item} for item in planned
            )
            result.finalize_status()
            return result

        root.mkdir(parents=True, exist_ok=True)
        created: list[Path] = []
        try:
            for directory in (
                self.settings.inbox_dir,
                self.settings.registered_dir,
                self.settings.topics_dir,
                root / "Imports" / "ReferenceText" / "Processed",
                self.settings.state_dir,
            ):
                if not directory.exists():
                    directory.mkdir(parents=True, exist_ok=False)
                    created.append(directory)
            self._write_workbook()
            created.append(self.settings.catalogue_path)
            if not self.settings.changes_log_path.exists():
                self.settings.changes_log_path.write_text(
                    "# Library changes\n",
                    encoding="utf-8",
                    newline="\n",
                )
                created.append(self.settings.changes_log_path)
            env_example = root / ".env.example"
            if not env_example.exists():
                env_example.write_text(ENV_EXAMPLE, encoding="utf-8", newline="\n")
                created.append(env_example)
        except Exception as exc:
            # Only remove files/directories created by this invocation, and
            # only while they remain empty. Existing runtime audit material is
            # never touched.
            for path in reversed(created):
                try:
                    if path.is_file():
                        path.unlink()
                    elif path.is_dir():
                        path.rmdir()
                except OSError:
                    pass
            raise FileOperationError(f"Could not initialize library: {exc}") from exc

        baseline_settings = Settings.from_root(root)
        baseline = DailyCheckWorkflow(baseline_settings).run()
        append_change_log(
            baseline_settings.changes_log_path,
            workflow="Initialization",
            action="Create empty LAM library and initial Workflow 1 baseline",
            files_changed=len(created),
            catalogue_rows_changed=0,
            reason="Explicit lam init --apply",
            uncertainty="None",
        )
        result.changed_files = len(created)
        result.state_committed = baseline.state_committed
        result.details["baseline"] = {
            "status": baseline.status.value,
            "report": baseline.report_path,
        }
        result.completed.append(
            {
                "action": "initialized_library",
                "paths_created": [path.relative_to(root).as_posix() for path in created],
            }
        )
        result.finalize_status()
        ReportService(baseline_settings.reports_dir).write(result)
        return result

    def _write_workbook(self) -> None:
        workbook = Workbook()
        catalogue = workbook.active
        catalogue.title = "Catalogue"
        catalogue.append(CATALOGUE_FIELDS)
        documents = workbook.create_sheet("Documents")
        documents.append(DOCUMENT_FIELDS)
        for sheet in (catalogue, documents):
            for cell in sheet[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="1F4E78")
            sheet.freeze_panes = "A2"
        temporary = self.settings.catalogue_path.with_name(".catalogue.init.tmp.xlsx")
        try:
            workbook.save(temporary)
            os.replace(temporary, self.settings.catalogue_path)
        finally:
            temporary.unlink(missing_ok=True)

    @classmethod
    def _blocking_entries(cls, root: Path) -> list[str]:
        if not root.exists():
            return []
        blockers: list[str] = []
        for path in root.iterdir():
            if path.name != ".library_state":
                blockers.append(path.name)
                continue
            if not path.is_dir():
                blockers.append(path.name)
                continue
            for child in path.iterdir():
                if child.name not in cls._PERMITTED_RUNTIME_CHILDREN:
                    blockers.append(f".library_state/{child.name}")
        return sorted(blockers, key=str.casefold)
