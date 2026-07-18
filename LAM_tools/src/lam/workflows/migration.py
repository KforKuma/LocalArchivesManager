from __future__ import annotations

from openpyxl import load_workbook

from ..config import Settings
from ..exceptions import CatalogueError
from ..models import WorkflowResult
from ..schema import (
    CATALOGUE_FIELDS,
    CATALOGUE_060_FIELDS,
    DOCUMENT_FIELDS,
    LEGACY_CATALOGUE_REQUIRED_FIELDS,
    LEGACY_IDENTITY_FIELDS,
    LEGACY_PDF_FIELDS,
)
from ..services.report_service import ReportService
from .identifier_migration import IdentifierMigrationWorkflow
from .schema_migration import SchemaMigrationWorkflow
from .topic_migration import TopicMigrationWorkflow


class MigrationWorkflow:
    """Public migration dispatcher with strict schema boundary detection."""

    def __init__(self, settings: Settings):
        self.settings = settings

    def identifiers(self, *, dry_run: bool) -> WorkflowResult:
        state = self._schema_state()
        if state["classification"] == "current":
            result = WorkflowResult(
                "identifier_migration",
                dry_run=dry_run,
                mode="dry_run" if dry_run else "apply",
            )
            result.details["schema_detection"] = state
            result.skipped.append({"reason": "already_current_schema"})
            result.finalize_status()
            ReportService(self.settings.reports_dir).write(result)
            return result
        if state["classification"] != "legacy":
            raise CatalogueError(
                "migrate identifiers refuses an unknown or future workbook schema: "
                + state.get("reason", "unrecognized columns")
            )
        result = IdentifierMigrationWorkflow(self.settings).run(dry_run=dry_run)
        result.details["schema_detection"] = state
        result.details["internal_document_stage"] = True
        return result

    def schema(self, *, dry_run: bool) -> WorkflowResult:
        state = self._schema_state()
        if state["classification"] == "current":
            result = WorkflowResult(
                "schema_migration",
                dry_run=dry_run,
                mode="dry_run" if dry_run else "apply",
            )
            result.details["schema_detection"] = state
            result.skipped.append({"reason": "already_current_schema"})
            result.finalize_status()
            ReportService(self.settings.reports_dir).write(result)
            return result
        if state["classification"] != "schema_060":
            raise CatalogueError(
                "migrate schema requires a strict LAM 0.6.0 workbook; "
                + state.get("reason", f"detected {state['classification']}")
            )
        result = SchemaMigrationWorkflow(self.settings).run(dry_run=dry_run)
        result.details["schema_detection"] = state
        return result

    def topics(self, *, dry_run: bool, include_topics: tuple[str, ...] = ()):
        return TopicMigrationWorkflow(self.settings).run(
            dry_run=dry_run,
            include_topics=include_topics,
        )

    def _schema_state(self) -> dict[str, object]:
        try:
            workbook = load_workbook(
                self.settings.catalogue_path, read_only=True, data_only=False
            )
        except Exception as exc:
            raise CatalogueError(
                f"Cannot inspect workbook schema: {self.settings.catalogue_path}"
            ) from exc
        if "Catalogue" not in workbook.sheetnames:
            workbook.close()
            return {
                "classification": "unknown",
                "reason": "Catalogue sheet is missing",
            }
        catalogue = workbook["Catalogue"]
        catalogue_headers = tuple(
            cell.value
            for cell in next(catalogue.iter_rows(min_row=1, max_row=1))
            if cell.value
        )
        document_headers: tuple[object, ...] = ()
        if "Documents" in workbook.sheetnames:
            documents = workbook["Documents"]
            document_headers = tuple(
                cell.value
                for cell in next(documents.iter_rows(min_row=1, max_row=1))
                if cell.value
            )
        workbook.close()
        if catalogue_headers == CATALOGUE_FIELDS and document_headers == DOCUMENT_FIELDS:
            return {
                "classification": "current",
                "catalogue_columns": list(catalogue_headers),
                "document_columns": list(document_headers),
            }
        if (
            catalogue_headers == CATALOGUE_060_FIELDS
            and document_headers == DOCUMENT_FIELDS
        ):
            return {
                "classification": "schema_060",
                "catalogue_columns": list(catalogue_headers),
                "document_columns": list(document_headers),
            }

        known = set(CATALOGUE_FIELDS) | set(LEGACY_IDENTITY_FIELDS) | set(
            LEGACY_PDF_FIELDS
        )
        unknown = sorted(set(catalogue_headers) - known)
        has_legacy_signature = LEGACY_CATALOGUE_REQUIRED_FIELDS.issubset(
            set(catalogue_headers)
        ) and (
            bool(set(catalogue_headers) & (set(LEGACY_IDENTITY_FIELDS) | set(LEGACY_PDF_FIELDS)))
            or "Documents" not in workbook.sheetnames
            or tuple(catalogue_headers) != CATALOGUE_FIELDS
        )
        documents_known = not document_headers or set(document_headers).issubset(
            set(DOCUMENT_FIELDS)
        )
        if has_legacy_signature and not unknown and documents_known:
            return {
                "classification": "legacy",
                "catalogue_columns": list(catalogue_headers),
                "document_columns": list(document_headers),
                "document_migration_required": document_headers != DOCUMENT_FIELDS,
            }
        return {
            "classification": "unknown",
            "catalogue_columns": list(catalogue_headers),
            "document_columns": list(document_headers),
            "reason": (
                f"unknown Catalogue columns: {unknown}"
                if unknown
                else "schema does not match a supported legacy or current layout"
            ),
        }
