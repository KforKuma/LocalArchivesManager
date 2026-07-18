from __future__ import annotations

import re
import uuid
from copy import copy
from datetime import date
from pathlib import Path
from typing import Any

from openpyxl.utils import get_column_letter

from ..config import Settings
from ..exceptions import CatalogueError
from ..models import CatalogueChange, WorkflowResult
from ..schema import CATALOGUE_FIELDS, DOCUMENT_FIELDS
from ..services.catalogue_service import CatalogueService
from ..services.journal_service import OperationJournal
from ..services.report_service import ReportService, append_change_log
from ..utils.hashing import full_hash
from ..utils.normalize import normalized_text
from .daily_check import DailyCheckWorkflow


_LOCAL_UUID = re.compile(
    r"^LOCAL:([0-9a-f]{8}-[0-9a-f]{4}-4[0-9a-f]{3}-[89ab][0-9a-f]{3}-[0-9a-f]{12})$",
    re.IGNORECASE,
)
_UNKNOWN_EXPECTATION_REVIEW = (
    "NEEDS_REVIEW: field=document_expectation; "
    "issue_key=legacy_document_expectation_unknown; "
    "issue=No Documents row establishes whether a managed document is expected."
)


class IdentifierMigrationWorkflow:
    """Migrate a legacy workbook to the strict current paper/document schema."""

    def __init__(self, settings: Settings):
        self.settings = settings

    def run(self, *, dry_run: bool) -> WorkflowResult:
        result = WorkflowResult(
            "identifier_migration",
            dry_run=dry_run,
            mode="dry_run" if dry_run else "apply",
        )
        catalogue = CatalogueService(
            self.settings.catalogue_path, allow_legacy_schema=True
        )
        records = catalogue.load()
        plan, blockers = self._plan(catalogue, records)
        result.details["plan"] = [self._public_plan(item) for item in plan]
        result.counts = {
            "catalogue_rows": len(records),
            "paper_uuid_preserved": sum(
                item["uuid_source"] == "paper_uuid" for item in plan
            ),
            "record_uid_recovered": sum(
                item["uuid_source"] == "record_uid" for item in plan
            ),
            "local_id_recovered": sum(
                item["uuid_source"] == "legacy_local_id" for item in plan
            ),
            "paper_uuid_generated": sum(
                item["uuid_source"] == "generated" for item in plan
            ),
            "documents_created": sum(item["document"] is not None for item in plan),
            "stale_legacy_file_fields_discarded": sum(
                bool(item.get("legacy_document_resolution")) for item in plan
            ),
            "blockers": len(blockers),
        }
        result.needs_review.extend(blockers)
        if blockers:
            result.skipped.append(
                {
                    "reason": "migration_blocked",
                    "message": "Catalogue was not modified because identity or document checks failed.",
                }
            )
            result.finalize_status()
            ReportService(self.settings.reports_dir).write(result)
            return result

        schema_changes = self._schema_changes(catalogue)
        result.details["schema_changes"] = schema_changes
        if dry_run:
            result.completed.extend(
                {
                    "action": "would_migrate_identifier",
                    "row": item["row_number"],
                    "paper_uuid_source": item["uuid_source"],
                    "would_create_document": item["document"] is not None,
                }
                for item in plan
            )
            result.changed_rows = len(plan)
            result.finalize_status()
            ReportService(self.settings.reports_dir).write(result)
            return result

        journal = OperationJournal.create(
            self.settings.state_dir,
            [
                {
                    "operation_type": "identifier_schema_migration",
                    "operation_id": f"paper:{item['paper_uuid']}",
                    "catalogue_row": item["row_number"],
                    "paper_uuid": item["paper_uuid"],
                    "document_id": (
                        item["document"]["document_id"] if item["document"] else None
                    ),
                    "execution_state": "planned",
                }
                for item in plan
            ],
            workflow="identifier_migration",
            suffix="migrate-identifiers",
        )
        self._apply(catalogue, records, plan)
        backup = catalogue.save_atomic()
        if backup:
            result.catalogue_backup = str(backup)
            journal.payload["catalogue_backup"] = str(backup)
            journal.write()
        for item in plan:
            journal.set_operation_state(
                item["row_number"],
                "catalogue_committed",
                operation_id=f"paper:{item['paper_uuid']}",
                document_id=(
                    item["document"]["document_id"] if item["document"] else None
                ),
            )
        result.changed_rows = len(plan)
        result.details["backup_cleanup"] = list(catalogue.maintenance_actions)
        result.completed.append(
            {
                "action": "migrated_identifier_schema",
                "catalogue_columns": list(CATALOGUE_FIELDS),
                "document_columns": list(DOCUMENT_FIELDS),
            }
        )
        append_change_log(
            self.settings.changes_log_path,
            workflow="Identifier migration",
            action="Replace legacy row identifiers and Catalogue PDF columns",
            files_changed=0,
            catalogue_rows_changed=len(plan),
            reason="Adopt paper_uuid as the sole paper key and Documents as the file-state table",
            uncertainty="None",
        )
        final_check = DailyCheckWorkflow(self.settings).run(
            dry_run=False, final_check=True
        )
        result.details["final_check"] = {
            "status": final_check.status.value,
            "report": final_check.report_path,
        }
        result.state_committed = final_check.state_committed
        result.needs_review.extend(
            item for item in final_check.needs_review if item not in result.needs_review
        )
        result.failures.extend(
            item for item in final_check.failures if item not in result.failures
        )
        journal.finish("final_check_committed")
        result.finalize_status()
        ReportService(self.settings.reports_dir).write(result)
        return result

    def _plan(
        self, catalogue: CatalogueService, records: list[Any]
    ) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
        blockers: list[dict[str, Any]] = []
        plan: list[dict[str, Any]] = []
        targets: dict[str, int] = {}
        by_current_uuid: dict[str, str] = {}

        for record in records:
            paper = self._uuid4(record.get("paper_uuid"))
            record_uid = self._uuid4(record.get("record_uid"))
            raw_paper = str(record.get("paper_uuid") or "").strip()
            raw_record_uid = str(record.get("record_uid") or "").strip()
            if raw_paper and not paper:
                blockers.append(self._blocker(record.row_number, "invalid_paper_uuid", raw_paper))
                continue
            if raw_record_uid and not record_uid:
                blockers.append(
                    self._blocker(record.row_number, "invalid_record_uid", raw_record_uid)
                )
                continue
            if paper and record_uid and paper != record_uid:
                blockers.append(
                    self._blocker(
                        record.row_number,
                        "identity_mismatch",
                        f"paper_uuid={paper}; record_uid={record_uid}",
                    )
                )
                continue
            local_uuid = self._local_uuid(record.get("id"))
            if paper:
                target, source = paper, "paper_uuid"
            elif record_uid:
                target, source = record_uid, "record_uid"
            elif local_uuid:
                target, source = local_uuid, "legacy_local_id"
            else:
                target, source = str(uuid.uuid4()), "generated"
            key = normalized_text(target)
            if key in targets:
                blockers.append(
                    self._blocker(
                        record.row_number,
                        "duplicate_paper_uuid",
                        f"also used by Catalogue row {targets[key]}",
                    )
                )
                continue
            targets[key] = record.row_number
            if paper:
                by_current_uuid[normalized_text(paper)] = target
            if record_uid:
                by_current_uuid[normalized_text(record_uid)] = target
            plan.append(
                {
                    "row_number": record.row_number,
                    "paper_uuid": target,
                    "uuid_source": source,
                    "record": record,
                    "document": None,
                    "legacy_document_resolution": "",
                }
            )

        if blockers:
            return plan, blockers

        target_set = set(targets)
        for document in catalogue.documents:
            raw_fk = str(document.get("paper_uuid") or "").strip()
            parsed_fk = self._uuid4(raw_fk)
            mapped = by_current_uuid.get(normalized_text(parsed_fk)) if parsed_fk else None
            if not parsed_fk or normalized_text(mapped or parsed_fk) not in target_set:
                blockers.append(
                    {
                        "sheet": "Documents",
                        "row": document.row_number,
                        "issue_key": "orphan_document",
                        "value": raw_fk,
                    }
                )

        today = date.today().isoformat()
        seen_legacy_paths: dict[str, int] = {}
        for item in plan:
            record = item["record"]
            paper_uuid = item["paper_uuid"]
            mains = [
                document
                for document in catalogue.documents
                if normalized_text(document.get("paper_uuid")) == normalized_text(paper_uuid)
                and normalized_text(document.get("document_type")) == "main"
            ]
            if len(mains) > 1:
                blockers.append(
                    self._blocker(record.row_number, "multiple_main_documents", paper_uuid)
                )
                continue
            relative = str(record.get("pdf_relative_path") or "").strip().replace("\\", "/")
            filename = str(record.get("pdf_filename") or "").strip()
            status = str(record.get("pdf_status") or "").strip()
            if not any((relative, filename, status)):
                continue
            if not relative:
                blockers.append(
                    self._blocker(record.row_number, "legacy_document_path_missing", filename)
                )
                continue
            path_key = normalized_text(relative)
            if path_key in seen_legacy_paths and seen_legacy_paths[path_key] != record.row_number:
                blockers.append(
                    self._blocker(
                        record.row_number,
                        "duplicate_legacy_document_path",
                        f"also used by Catalogue row {seen_legacy_paths[path_key]}",
                    )
                )
                continue
            seen_legacy_paths[path_key] = record.row_number
            if mains:
                conflicts = self._document_conflicts(mains[0], relative, filename, status)
                if conflicts:
                    documents_path = (
                        self.settings.library_root
                        / str(mains[0].get("relative_path") or "")
                    )
                    legacy_path = self.settings.library_root / relative
                    if documents_path.is_file() and not legacy_path.is_file():
                        item["legacy_document_resolution"] = (
                            "discard_stale_legacy_fields_use_observed_documents"
                        )
                    else:
                        blockers.append(
                            self._blocker(
                                record.row_number,
                                "legacy_document_mismatch",
                                "; ".join(conflicts),
                            )
                        )
                continue
            path = (self.settings.library_root / relative).resolve()
            try:
                path.relative_to(self.settings.library_root.resolve())
            except ValueError:
                blockers.append(
                    self._blocker(record.row_number, "unsafe_legacy_document_path", relative)
                )
                continue
            observed_name = path.name if path.is_file() else filename or Path(relative).name
            item["document"] = {
                "document_id": f"{paper_uuid}:main",
                "paper_uuid": paper_uuid,
                "uncertainty": "" if path.is_file() else "document_file_missing",
                "document_type": "main",
                "filename": observed_name,
                "relative_path": relative,
                "extension": Path(observed_name).suffix,
                "sha256": full_hash(path) if path.is_file() else "",
                "file_status": status or ("registered" if path.is_file() else "missing"),
                "source": str(record.get("source") or ""),
                "date_added": str(record.get("date_added") or today),
                "date_updated": today,
            }
        for item in plan:
            has_document = item["document"] is not None or any(
                normalized_text(document.get("paper_uuid"))
                == normalized_text(item["paper_uuid"])
                for document in catalogue.documents
            )
            item["record_origin"] = "pdf" if has_document else "legacy"
            item["document_expectation"] = "required" if has_document else "unknown"
        return plan, blockers

    def _apply(
        self,
        catalogue: CatalogueService,
        records: list[Any],
        plan: list[dict[str, Any]],
    ) -> None:
        for field_name in CATALOGUE_FIELDS:
            catalogue.ensure_header(field_name)
        catalogue.ensure_documents_sheet()
        for field_name in DOCUMENT_FIELDS:
            self._ensure_document_header(catalogue, field_name)
        by_row = {record.row_number: record for record in records}
        for item in plan:
            record = by_row[item["row_number"]]
            old = record.get("paper_uuid")
            if normalized_text(old) != normalized_text(item["paper_uuid"]):
                column = catalogue.headers["paper_uuid"]
                catalogue.worksheet.cell(row=record.row_number, column=column).value = item[
                    "paper_uuid"
                ]
                record.values["paper_uuid"] = item["paper_uuid"]
                catalogue.changes.append(
                    CatalogueChange(record.row_number, "paper_uuid", old, item["paper_uuid"])
                )
            for field_name in ("record_origin", "document_expectation"):
                old_value = record.get(field_name)
                new_value = item[field_name]
                column = catalogue.headers[field_name]
                catalogue.worksheet.cell(row=record.row_number, column=column).value = new_value
                record.values[field_name] = new_value
                if old_value != new_value:
                    catalogue.changes.append(
                        CatalogueChange(record.row_number, field_name, old_value, new_value)
                    )
            if item["document_expectation"] == "unknown":
                old_uncertainty = str(record.get("uncertainty") or "").strip()
                if "issue_key=legacy_document_expectation_unknown" not in old_uncertainty:
                    new_uncertainty = "\n".join(
                        value
                        for value in (old_uncertainty, _UNKNOWN_EXPECTATION_REVIEW)
                        if value
                    )
                    column = catalogue.headers["uncertainty"]
                    catalogue.worksheet.cell(row=record.row_number, column=column).value = (
                        new_uncertainty
                    )
                    record.values["uncertainty"] = new_uncertainty
                    catalogue.changes.append(
                        CatalogueChange(
                            record.row_number,
                            "uncertainty",
                            old_uncertainty,
                            new_uncertainty,
                        )
                    )
            if item["document"]:
                catalogue.add_document(item["document"])

        old_catalogue_headers = tuple(catalogue.headers)
        old_document_headers = tuple(catalogue.document_headers)
        self._reorder_sheet(catalogue.worksheet, catalogue.headers, CATALOGUE_FIELDS)
        self._reorder_sheet(
            catalogue.documents_worksheet, catalogue.document_headers, DOCUMENT_FIELDS
        )
        catalogue.mark_schema_dirty()
        catalogue.headers = {
            field: index for index, field in enumerate(CATALOGUE_FIELDS, start=1)
        }
        catalogue.document_headers = {
            field: index for index, field in enumerate(DOCUMENT_FIELDS, start=1)
        }
        for record in catalogue.records:
            record.values = {
                field: catalogue.worksheet.cell(
                    row=record.row_number, column=catalogue.headers[field]
                ).value
                for field in CATALOGUE_FIELDS
            }
        for document in catalogue.documents:
            document.values = {
                field: catalogue.documents_worksheet.cell(
                    row=document.row_number, column=catalogue.document_headers[field]
                ).value
                for field in DOCUMENT_FIELDS
            }
        catalogue.changes.append(
            CatalogueChange(1, "__schema__", old_catalogue_headers, CATALOGUE_FIELDS)
        )
        catalogue.document_changes.append(
            CatalogueChange(1, "__schema__", old_document_headers, DOCUMENT_FIELDS)
        )
        catalogue._validate_duplicate_values()
        catalogue._validate_documents()

    @staticmethod
    def _ensure_document_header(catalogue: CatalogueService, field_name: str) -> int:
        if field_name in catalogue.document_headers:
            return catalogue.document_headers[field_name]
        sheet = catalogue.documents_worksheet
        if sheet is None:
            raise CatalogueError("Documents sheet is unavailable")
        column = sheet.max_column + 1
        target = sheet.cell(row=1, column=column, value=field_name)
        if column > 1:
            target._style = copy(sheet.cell(row=1, column=column - 1)._style)
        catalogue.document_headers[field_name] = column
        for document in catalogue.documents:
            document.values.setdefault(field_name, None)
        return column

    @staticmethod
    def _reorder_sheet(sheet: Any, headers: dict[str, int], order: tuple[str, ...]) -> None:
        max_row = max(sheet.max_row, 1)
        cells: dict[str, list[dict[str, Any]]] = {}
        widths: dict[str, float | None] = {}
        for field in order:
            column = headers[field]
            letter = get_column_letter(column)
            widths[field] = sheet.column_dimensions[letter].width
            cells[field] = []
            for row in range(1, max_row + 1):
                cell = sheet.cell(row=row, column=column)
                cells[field].append(
                    {
                        "value": cell.value,
                        "style": copy(cell._style),
                        "comment": copy(cell.comment),
                        "hyperlink": copy(cell.hyperlink),
                    }
                )
        sheet.delete_cols(1, sheet.max_column)
        for column, field in enumerate(order, start=1):
            for row, payload in enumerate(cells[field], start=1):
                cell = sheet.cell(row=row, column=column, value=payload["value"])
                cell._style = payload["style"]
                cell.comment = payload["comment"]
                if payload["hyperlink"]:
                    cell._hyperlink = payload["hyperlink"]
            sheet.column_dimensions[get_column_letter(column)].width = widths[field]
        last = get_column_letter(len(order))
        sheet.auto_filter.ref = f"A1:{last}{max_row}"
        for table in sheet.tables.values():
            table.ref = f"A1:{last}{max_row}"

    @staticmethod
    def _uuid4(value: Any) -> str:
        raw = str(value or "").strip()
        if not raw:
            return ""
        try:
            parsed = uuid.UUID(raw)
        except ValueError:
            return ""
        return str(parsed) if parsed.version == 4 else ""

    @staticmethod
    def _local_uuid(value: Any) -> str:
        match = _LOCAL_UUID.fullmatch(str(value or "").strip())
        return str(uuid.UUID(match.group(1))) if match else ""

    @staticmethod
    def _document_conflicts(
        document: Any, relative: str, filename: str, status: str
    ) -> list[str]:
        conflicts: list[str] = []
        checks = (
            ("relative_path", relative),
            ("filename", filename),
            ("file_status", status),
        )
        for field, legacy in checks:
            current = str(document.get(field) or "").strip()
            if current and legacy and normalized_text(current) != normalized_text(legacy):
                conflicts.append(f"{field}: Catalogue={legacy!r}, Documents={current!r}")
        return conflicts

    @staticmethod
    def _blocker(row: int, issue_key: str, value: str) -> dict[str, Any]:
        return {
            "sheet": "Catalogue",
            "row": row,
            "issue_key": issue_key,
            "value": value,
        }

    @staticmethod
    def _public_plan(item: dict[str, Any]) -> dict[str, Any]:
        return {
            "row": item["row_number"],
            "paper_uuid": item["paper_uuid"],
            "paper_uuid_source": item["uuid_source"],
            "would_create_document": item["document"] is not None,
            "legacy_document_resolution": item.get("legacy_document_resolution") or None,
            "record_origin": item.get("record_origin"),
            "document_expectation": item.get("document_expectation"),
        }

    @staticmethod
    def _schema_changes(catalogue: CatalogueService) -> dict[str, Any]:
        catalogue_headers = tuple(catalogue.headers)
        document_headers = tuple(catalogue.document_headers)
        return {
            "catalogue_removed": [
                field for field in catalogue_headers if field not in CATALOGUE_FIELDS
            ],
            "catalogue_added": [
                field for field in CATALOGUE_FIELDS if field not in catalogue_headers
            ],
            "catalogue_reordered": catalogue_headers != CATALOGUE_FIELDS,
            "documents_removed": [
                field for field in document_headers if field not in DOCUMENT_FIELDS
            ],
            "documents_added": [
                field for field in DOCUMENT_FIELDS if field not in document_headers
            ],
            "documents_reordered": document_headers != DOCUMENT_FIELDS,
        }
