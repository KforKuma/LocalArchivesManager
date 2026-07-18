from __future__ import annotations

import json
import os
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from ..config import Settings
from ..models import WorkflowResult
from ..schema import MANAGED_DOCUMENT_EXTENSIONS
from ..services.catalogue_service import CatalogueService
from ..services.journal_service import OperationJournal, incomplete_journals
from ..services.report_service import ReportService, append_change_log
from ..services.snapshot_service import SnapshotService
from ..utils.hashing import full_hash
from ..utils.normalize import normalized_relative_path, normalized_text
from ..utils.publication_type import canonicalize_publication_type
from .daily_check import DailyCheckWorkflow
from .inbox_register import InboxRegisterWorkflow
from .publication_type_repair import PublicationTypeRepairWorkflow
from .trash_recovery import TrashRecoveryWorkflow


class RecoveryWorkflow:
    """Conservative recovery orchestration for interrupted LAM operations."""

    SCOPES = {"auto", "workbook", "inbox", "registered", "publication-types"}

    def __init__(self, settings: Settings):
        self.settings = settings

    def run(
        self,
        *,
        dry_run: bool,
        scope: str = "auto",
        offline: bool = False,
        refresh: bool = False,
        cache_write: bool = True,
        list_trash: bool = False,
        trash_id: str | None = None,
    ) -> WorkflowResult:
        if list_trash:
            return TrashRecoveryWorkflow(self.settings).list()
        if trash_id:
            return TrashRecoveryWorkflow(self.settings).run(
                deletion_id=trash_id,
                dry_run=dry_run,
            )
        result = WorkflowResult(
            "recover",
            dry_run=dry_run,
            mode="dry_run" if dry_run else "apply",
        )
        unfinished = incomplete_journals(self.settings.state_dir)
        selected = self._selected_scopes(scope, unfinished)
        result.details = {
            "requested_scope": scope,
            "selected_scopes": sorted(selected),
            "incomplete_journals": unfinished,
            "provider_policy": {
                "offline": offline,
                "refresh": refresh,
                "cache_write": cache_write,
                "inbox_only": True,
                "filed_network": False,
            },
            "filed_documents_re_registered": False,
        }

        nested: list[dict[str, Any]] = []
        changed_rows = 0
        changed_files = 0

        if "registered" in selected:
            registered = self._recover_registered(dry_run=dry_run)
            nested.append(registered)
            changed_rows += registered["changed_rows"]
            result.needs_review.extend(registered["needs_review"])
            result.completed.extend(registered["completed"])

        if "workbook" in selected:
            evidence = self._workbook_recovery_evidence(unfinished)
            check = DailyCheckWorkflow(self.settings).run(dry_run=dry_run)
            nested.append(
                {
                    "scope": "workbook",
                    "status": check.status.value,
                    "report": check.report_path,
                    "changed_rows": check.changed_rows,
                    "evidence": evidence,
                    "user_fields_overwritten": False,
                }
            )
            changed_rows += check.changed_rows
            result.needs_review.extend(
                item for item in check.needs_review if item not in result.needs_review
            )

        if "inbox" in selected and self._has_inbox_interruption(unfinished):
            inbox = InboxRegisterWorkflow(self.settings).run(
                dry_run=dry_run,
                offline=offline,
                refresh=refresh,
                cache_write=cache_write,
            )
            nested.append(
                {
                    "scope": "inbox",
                    "status": inbox.status.value,
                    "report": inbox.report_path,
                    "changed_files": inbox.changed_files,
                    "changed_rows": inbox.changed_rows,
                }
            )
            changed_files += inbox.changed_files
            changed_rows += inbox.changed_rows
            result.needs_review.extend(
                item for item in inbox.needs_review if item not in result.needs_review
            )
        elif "inbox" in selected:
            nested.append(
                {
                    "scope": "inbox",
                    "status": "no_changes",
                    "reason": "no unfinished Workflow 3 journal detected",
                }
            )

        if "publication-types" in selected:
            if self._has_publication_type_anomalies():
                repair = PublicationTypeRepairWorkflow(self.settings).run(dry_run=dry_run)
                nested.append(
                    {
                        "scope": "publication-types",
                        "status": repair.status.value,
                        "report": repair.report_path,
                        "changed_files": repair.changed_files,
                        "changed_rows": repair.changed_rows,
                    }
                )
                changed_files += repair.changed_files
                changed_rows += repair.changed_rows
                result.needs_review.extend(
                    item for item in repair.needs_review if item not in result.needs_review
                )
            else:
                nested.append(
                    {
                        "scope": "publication-types",
                        "status": "no_changes",
                        "reason": "no historical mixed publication_type values detected",
                    }
                )

        result.details["scope_results"] = nested
        result.changed_files = changed_files
        result.changed_rows = changed_rows

        if not dry_run:
            recovered_journals = self._close_resolved_journals()
            result.details["recovered_journals"] = recovered_journals
            if recovered_journals:
                result.completed.extend(
                    {"action": "marked_interrupted_journal_recovered", **item}
                    for item in recovered_journals
                )
            final = DailyCheckWorkflow(self.settings).run(final_check=True)
            result.details["final_check"] = {
                "status": final.status.value,
                "report": final.report_path,
            }
            result.state_committed = final.state_committed
            result.needs_review.extend(
                item for item in final.needs_review if item not in result.needs_review
            )

        if not selected:
            result.skipped.append({"reason": "no_recovery_action_detected"})
        result.finalize_status()
        ReportService(self.settings.reports_dir).write(result)
        return result

    def _selected_scopes(
        self, scope: str, unfinished: list[dict[str, Any]]
    ) -> set[str]:
        if scope != "auto":
            return {scope}
        selected: set[str] = set()
        workflows = {normalized_text(item.get("workflow")) for item in unfinished}
        if "inbox_register" in workflows and self._direct_documents(self.settings.inbox_dir):
            selected.add("inbox")
        if self._registered_orphans():
            selected.add("registered")
        if any(item not in {"inbox_register"} for item in workflows):
            selected.add("workbook")
        if self._has_publication_type_anomalies():
            selected.add("publication-types")
        return selected

    @staticmethod
    def _has_inbox_interruption(unfinished: list[dict[str, Any]]) -> bool:
        return any(
            normalized_text(item.get("workflow")) == "inbox_register"
            for item in unfinished
        )

    def _recover_registered(self, *, dry_run: bool) -> dict[str, Any]:
        catalogue = CatalogueService(self.settings.catalogue_path)
        records = catalogue.load()
        records_by_uuid = {
            normalized_text(record.get("paper_uuid")): record for record in records
        }
        documents_by_id = {
            normalized_text(document.get("document_id")): document
            for document in catalogue.documents
        }
        operations = self._journal_operations()
        completed: list[dict[str, Any]] = []
        review: list[dict[str, Any]] = []
        today = date.today().isoformat()

        for path in self._registered_orphans():
            relative = path.relative_to(self.settings.library_root).as_posix()
            matching = [
                operation
                for operation in operations
                if normalized_relative_path(operation.get("target"))
                == normalized_relative_path(relative)
                and normalized_text(operation.get("paper_uuid")) in records_by_uuid
            ]
            if len(matching) != 1:
                review.append(
                    {
                        "file": relative,
                        "issue": "registered_orphan_identity_ambiguous",
                        "journal_candidates": len(matching),
                    }
                )
                continue
            operation = matching[0]
            paper_uuid = str(operation.get("paper_uuid"))
            document_id = str(operation.get("document_id") or f"{paper_uuid}:main")
            if normalized_text(document_id) in documents_by_id:
                document = documents_by_id[normalized_text(document_id)]
                updates = {
                    "filename": path.name,
                    "relative_path": relative,
                    "extension": path.suffix,
                    "sha256": full_hash(path),
                    "file_status": "registered",
                    "date_updated": today,
                }
                if dry_run:
                    completed.append(
                        {
                            "action": "would_reconnect_registered_document",
                            "file": relative,
                            "document_id": document_id,
                        }
                    )
                else:
                    catalogue.update_document_fields(document, updates)
                    completed.append(
                        {
                            "action": "reconnected_registered_document",
                            "file": relative,
                            "document_id": document_id,
                        }
                    )
                continue

            planned = operation.get("planned_updates")
            values = dict(planned) if isinstance(planned, dict) else {}
            is_supplementary = ":supp:" in document_id.casefold()
            values.update(
                {
                    "document_id": document_id,
                    "paper_uuid": paper_uuid,
                    "uncertainty": "",
                    "document_type": (
                        values.get("document_type")
                        or ("supplementary" if is_supplementary else "main")
                    ),
                    "filename": path.name,
                    "relative_path": relative,
                    "extension": path.suffix,
                    "sha256": full_hash(path),
                    "file_status": "registered",
                    "source": values.get("source") or "recovery_journal",
                    "date_added": values.get("date_added") or today,
                    "date_updated": today,
                }
            )
            if dry_run:
                completed.append(
                    {
                        "action": "would_restore_registered_document",
                        "file": relative,
                        "paper_uuid": paper_uuid,
                        "document_id": document_id,
                    }
                )
            else:
                catalogue.add_document(values)
                completed.append(
                    {
                        "action": "restored_registered_document",
                        "file": relative,
                        "paper_uuid": paper_uuid,
                        "document_id": document_id,
                    }
                )

        changed_rows = 0
        if not dry_run and (catalogue.changes or catalogue.document_changes):
            recovery_journal = OperationJournal.create(
                self.settings.state_dir,
                [
                    {
                        "operation_type": "registered_binding_recovery",
                        "operation_id": item.get("document_id"),
                        "document_id": item.get("document_id"),
                        "paper_uuid": item.get("paper_uuid"),
                        "target": item.get("file"),
                        "execution_state": "planned",
                    }
                    for item in completed
                    if item["action"].startswith("restored")
                    or item["action"].startswith("reconnected")
                ],
                workflow="recover",
                suffix="recover-registered",
            )
            backup = catalogue.save_atomic()
            if backup:
                recovery_journal.payload["catalogue_backup"] = str(backup)
            recovery_journal.finish("recovered")
            changed_rows = len(
                {item.row_number for item in catalogue.changes}
                | {item.row_number for item in catalogue.document_changes}
            )
            append_change_log(
                self.settings.changes_log_path,
                workflow="Recovery",
                action="Reconnect Registered orphan Documents",
                files_changed=0,
                catalogue_rows_changed=changed_rows,
                reason="Operation journal uniquely identified the existing Registered file",
                uncertainty=f"{len(review)} orphan(s) remain unresolved",
            )
        return {
            "scope": "registered",
            "status": "needs_review" if review else ("success" if completed else "no_changes"),
            "changed_rows": changed_rows,
            "completed": completed,
            "needs_review": review,
        }

    def _registered_orphans(self) -> list[Path]:
        referenced: set[str] = set()
        try:
            catalogue = CatalogueService(self.settings.catalogue_path)
            catalogue.load()
            referenced = {
                normalized_relative_path(document.get("relative_path"))
                for document in catalogue.documents
                if document.get("relative_path")
            }
        except Exception:
            pass
        return [
            path
            for path in self._direct_documents(self.settings.registered_dir)
            if normalized_relative_path(
                path.relative_to(self.settings.library_root).as_posix()
            )
            not in referenced
        ]

    @staticmethod
    def _direct_documents(directory: Path) -> list[Path]:
        if not directory.is_dir():
            return []
        return sorted(
            [
                path
                for path in directory.iterdir()
                if path.is_file()
                and not path.name.startswith(".")
                and path.suffix.casefold() in MANAGED_DOCUMENT_EXTENSIONS
            ],
            key=lambda path: path.name.casefold(),
        )

    def _journal_operations(self) -> list[dict[str, Any]]:
        operations: list[dict[str, Any]] = []
        runs = self.settings.state_dir / "runs"
        if not runs.is_dir():
            return operations
        for path in runs.glob("*/operation_journal.json"):
            try:
                payload = json.loads(path.read_text(encoding="utf-8"))
            except Exception:
                continue
            for operation in payload.get("operations", []):
                operations.append({**operation, "_journal": str(path)})
        return operations

    def _has_publication_type_anomalies(self) -> bool:
        try:
            catalogue = CatalogueService(self.settings.catalogue_path)
            records = catalogue.load()
        except Exception:
            return False
        for record in records:
            raw = str(record.get("publication_type") or "").strip()
            canonical = canonicalize_publication_type(raw).canonical_type or ""
            if raw and normalized_text(raw) != normalized_text(canonical):
                return True
        return False

    def _workbook_recovery_evidence(
        self, unfinished: list[dict[str, Any]]
    ) -> dict[str, Any]:
        marker_path = self.settings.state_dir / "snapshot_commit.json"
        current_generation = None
        if marker_path.is_file():
            try:
                current_generation = json.loads(
                    marker_path.read_text(encoding="utf-8")
                ).get("generation_id")
            except Exception:
                current_generation = "unreadable"
        generations_root = self.settings.state_dir / "snapshot_generations"
        generations = (
            sorted(
                (path.name for path in generations_root.iterdir() if path.is_dir()),
                reverse=True,
            )
            if generations_root.is_dir()
            else []
        )
        previous_generation = next(
            (item for item in generations if item != current_generation), None
        )
        comparisons: dict[str, Any] = {}
        try:
            catalogue = CatalogueService(self.settings.catalogue_path)
            catalogue.load()
            current_workbook = catalogue.snapshot_payload()
            snapshots = SnapshotService(
                self.settings.library_root,
                self.settings.state_dir,
                self.settings.reserved_root_directories,
            )
            committed = snapshots.load_catalogue_snapshot()
            comparisons["committed_to_current"] = snapshots.compare_catalogue(
                committed, current_workbook
            ) if committed else []
            previous_path = (
                generations_root / previous_generation / "catalogue_snapshot.json"
                if previous_generation
                else None
            )
            previous = (
                json.loads(previous_path.read_text(encoding="utf-8"))
                if previous_path and previous_path.is_file()
                else {}
            )
            comparisons["previous_to_committed"] = snapshots.compare_catalogue(
                previous, committed
            ) if previous and committed else []
        except Exception as exc:
            comparisons = {"unavailable": type(exc).__name__}
        backups = sorted(
            self.settings.library_root.glob("catalogue.backup.*.xlsx"),
            key=lambda path: path.stat().st_mtime_ns,
            reverse=True,
        )
        valid_backup = next(
            (str(path) for path in backups if self._valid_backup(path)), None
        )
        return {
            "comparison": "current workbook vs committed snapshot vs previous snapshot",
            "comparison_results": comparisons,
            "current_workbook": str(self.settings.catalogue_path),
            "current_generation": current_generation,
            "previous_generation": previous_generation,
            "latest_valid_backup": valid_backup,
            "operation_journals": unfinished,
            "filesystem_observed_by": "Workflow 1 reconciliation",
            "automatic_repairs": "unambiguous machine fields only",
            "whole_backup_restore": False,
            "user_fields_protected": True,
        }

    @staticmethod
    def _valid_backup(path: Path) -> bool:
        try:
            workbook = load_workbook(path, read_only=True, data_only=False)
            valid = "Catalogue" in workbook.sheetnames
            workbook.close()
            return valid
        except Exception:
            return False

    def _close_resolved_journals(self) -> list[dict[str, Any]]:
        catalogue = CatalogueService(self.settings.catalogue_path)
        catalogue.load()
        document_paths = {
            normalized_relative_path(document.get("relative_path"))
            for document in catalogue.documents
            if document.get("relative_path")
        }
        resolved = []
        for item in incomplete_journals(self.settings.state_dir):
            path = Path(str(item.get("journal") or ""))
            try:
                payload = json.loads(path.read_text(encoding="utf-8"))
            except Exception:
                continue
            operations = payload.get("operations", [])
            if not operations:
                continue
            complete = True
            for operation in operations:
                target = normalized_relative_path(operation.get("target"))
                if target:
                    target_path = self.settings.library_root / target
                    document_id = normalized_text(operation.get("document_id"))
                    if not target_path.is_file():
                        complete = False
                        break
                    if document_id and target not in document_paths:
                        complete = False
                        break
                elif payload.get("workflow") not in {
                    "identifier_migration",
                    "document_migration",
                }:
                    complete = False
                    break
            if not complete:
                continue
            payload["status"] = "recovered"
            payload["recovered_at"] = datetime.now().astimezone().isoformat(
                timespec="seconds"
            )
            temporary = path.with_name(f".{path.name}.tmp")
            temporary.write_text(
                json.dumps(payload, ensure_ascii=False, indent=2, default=str) + "\n",
                encoding="utf-8",
                newline="\n",
            )
            os.replace(temporary, path)
            resolved.append(
                {"journal": str(path), "run_id": payload.get("run_id")}
            )
        return resolved
