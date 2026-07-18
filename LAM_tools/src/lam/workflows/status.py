from __future__ import annotations

import json
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from filelock import FileLock, Timeout

from ..config import Settings
from ..models import WorkflowResult
from ..schema import CATALOGUE_FIELDS, DOCUMENT_FIELDS, MANAGED_DOCUMENT_EXTENSIONS
from ..services.catalogue_service import CatalogueService
from ..services.journal_service import incomplete_journals
from ..services.report_service import ReportService
from ..services.run_workspace import temporary_inventory
from ..utils.normalize import normalized_relative_path
from ..versions import (
    JSON_SCHEMA_VERSION,
    LIBRARY_SCHEMA_VERSION,
    PACKAGE_VERSION,
)


class StatusWorkflow:
    def __init__(self, settings: Settings):
        self.settings = settings

    def library(self) -> WorkflowResult:
        result = WorkflowResult("status_library", mode="diagnostic")
        workbook_state = self._workbook_state()
        counts = self._file_counts()
        temporary = temporary_inventory(self.settings)
        details: dict[str, Any] = {
            "initialized": workbook_state["current_schema"],
            "package_version": PACKAGE_VERSION,
            "library_schema_version": LIBRARY_SCHEMA_VERSION,
            "json_schema_version": JSON_SCHEMA_VERSION,
            "schema_version": (
                LIBRARY_SCHEMA_VERSION if workbook_state["current_schema"] else None
            ),
            "workbook": workbook_state,
            "files": counts,
            "active_blockers": 0,
            "provisional_records": 0,
            "missing_documents": 0,
            "recent_run": self._recent_runtime_item(),
            "temporary": temporary,
        }
        if self.settings.catalogue_path.is_file():
            try:
                catalogue = CatalogueService(self.settings.catalogue_path)
                records = catalogue.load()
                catalogue_blockers = sum(
                    line.lstrip().upper().startswith("NEEDS_REVIEW:")
                    for record in records
                    for line in str(record.get("uncertainty") or "").splitlines()
                )
                document_blockers = sum(
                    any(
                        line.lstrip().upper().startswith("NEEDS_REVIEW:")
                        or line.strip().casefold()
                        in {
                            "document_file_missing",
                            "source_missing",
                            "topic_location_mismatch",
                            "document_target_collision",
                            "target_collision",
                            "supplementary_target_collision",
                        }
                        for line in str(document.get("uncertainty") or "").splitlines()
                    )
                    for document in catalogue.documents
                )
                details.update(
                    {
                        "catalogue_rows": len(records),
                        "document_rows": len(catalogue.documents),
                        "active_blockers": catalogue_blockers + document_blockers,
                        "provisional_records": sum(
                            str(record.get("source") or "").strip().casefold()
                            == "local_pdf"
                            or "metadata_identity_unconfirmed"
                            in str(record.get("uncertainty") or "")
                            for record in records
                        ),
                        "missing_documents": sum(
                            str(document.get("file_status") or "").strip().casefold()
                            == "missing"
                            for document in catalogue.documents
                        ),
                    }
                )
            except Exception as exc:
                result.needs_review.append(
                    {"issue": "workbook_schema_unavailable", "detail": str(exc)}
                )
        result.details = details
        result.counts = {
            "catalogue_rows": int(details.get("catalogue_rows", 0)),
            "document_rows": int(details.get("document_rows", 0)),
            "active_blockers": int(details.get("active_blockers", 0)),
            "provisional_records": int(details.get("provisional_records", 0)),
            "missing_documents": int(details.get("missing_documents", 0)),
            "temporary_directories": int(temporary["temporary_directories"]),
            "temporary_files": int(temporary["temporary_files"]),
            "temporary_bytes": int(temporary["temporary_bytes"]),
            "expired_temporary_artifacts": int(
                temporary["expired_temporary_artifacts"]
            ),
            "unreadable_temporary_artifacts": int(
                temporary["unreadable_temporary_artifacts"]
            ),
            "unknown_temporary_artifacts": int(
                temporary["unknown_temporary_artifacts"]
            ),
        }
        if not workbook_state["exists"]:
            result.needs_review.append(
                {"issue": "library_not_initialized", "recommendation": "lam init --dry-run"}
            )
        result.finalize_status()
        self._write_report(result)
        return result

    def recovery(self) -> WorkflowResult:
        result = WorkflowResult("status_recovery", mode="diagnostic")
        incomplete = incomplete_journals(self.settings.state_dir)
        snapshot = self._snapshot_state()
        orphans = self._orphans()
        backups = sorted(
            self.settings.library_root.glob("catalogue.backup.*.xlsx"),
            key=lambda path: path.stat().st_mtime_ns,
            reverse=True,
        )
        valid_backup = next((path for path in backups if self._valid_workbook(path)), None)
        result.details = {
            "lock": self._lock_state(),
            "incomplete_journals": incomplete,
            "snapshots": snapshot,
            "latest_valid_backup": str(valid_backup) if valid_backup else None,
            "orphan_documents": orphans,
            "recover_recommended": bool(incomplete or orphans),
        }
        result.counts = {
            "incomplete_journals": len(incomplete),
            "orphan_documents": len(orphans),
            "backups": len(backups),
        }
        if incomplete or orphans:
            result.needs_review.append(
                {
                    "issue": "recovery_recommended",
                    "recommendation": "lam recover --dry-run",
                }
            )
        result.finalize_status()
        self._write_report(result)
        return result

    def config(self) -> WorkflowResult:
        result = WorkflowResult("status_config", mode="diagnostic")
        result.details = {
            "library_root": str(self.settings.library_root),
            "catalogue_exists": self.settings.catalogue_path.is_file(),
            "providers": {
                "pubmed_enabled": self.settings.pubmed.enabled,
                "crossref_enabled": self.settings.crossref.enabled,
                "arxiv_enabled": self.settings.arxiv.enabled,
                "unpaywall_enabled": self.settings.unpaywall.enabled,
                "NCBI_EMAIL": self._configured(self.settings.pubmed.email),
                "NCBI_TOOL": self._configured(self.settings.pubmed.tool),
                "NCBI_API_KEY": self._configured(self.settings.pubmed.api_key),
                "CROSSREF_EMAIL": self._configured(self.settings.crossref.email),
                "UNPAYWALL_EMAIL": self._configured(self.settings.unpaywall.email),
            },
            "network": {
                "timeout_seconds": self.settings.network.timeout_seconds,
                "max_retries": self.settings.network.max_retries,
                "user_agent": self.settings.network.user_agent,
            },
            "ocr": {
                "enabled": self.settings.ocr.enabled,
                "languages": list(self.settings.ocr.languages),
                "gpu": self.settings.ocr.gpu,
                "POPPLER_PATH": self._configured(self.settings.ocr.poppler_path),
                "OCR_MODEL_STORAGE_DIR": self._configured(
                    self.settings.ocr.model_storage_dir
                ),
                "model_download_enabled": self.settings.ocr.download_enabled,
            },
            "document_analysis": {
                "backend": self.settings.document_analysis.backend,
                "fallbacks": list(self.settings.document_analysis.fallbacks),
                "doi_min_suffix_alnum": (
                    self.settings.document_analysis.doi_min_suffix_alnum
                ),
                "doi_max_length": self.settings.document_analysis.doi_max_length,
                "installed_backends": ["native", "easyocr"],
            },
            "secrets_exposed": False,
        }
        result.completed.append({"action": "reported_non_sensitive_effective_config"})
        self._write_report(result)
        return result

    def _workbook_state(self) -> dict[str, Any]:
        path = self.settings.catalogue_path
        if not path.is_file():
            return {
                "exists": False,
                "current_schema": False,
                "catalogue_columns": [],
                "document_columns": [],
            }
        try:
            workbook = load_workbook(path, read_only=True, data_only=False)
            catalogue = workbook["Catalogue"] if "Catalogue" in workbook.sheetnames else None
            documents = workbook["Documents"] if "Documents" in workbook.sheetnames else None
            catalogue_columns = (
                [cell.value for cell in next(catalogue.iter_rows(min_row=1, max_row=1))]
                if catalogue
                else []
            )
            document_columns = (
                [cell.value for cell in next(documents.iter_rows(min_row=1, max_row=1))]
                if documents
                else []
            )
            workbook.close()
            return {
                "exists": True,
                "current_schema": tuple(catalogue_columns) == CATALOGUE_FIELDS
                and tuple(document_columns) == DOCUMENT_FIELDS,
                "catalogue_columns": catalogue_columns,
                "document_columns": document_columns,
            }
        except Exception as exc:
            return {
                "exists": True,
                "current_schema": False,
                "error": str(exc),
                "catalogue_columns": [],
                "document_columns": [],
            }

    def _file_counts(self) -> dict[str, int]:
        return {
            "Inbox": self._count_documents(self.settings.inbox_dir, recursive=False),
            "Registered": self._count_documents(
                self.settings.registered_dir, recursive=False
            ),
            "Topics": self._count_documents(self.settings.topics_dir, recursive=True),
        }

    @staticmethod
    def _count_documents(directory: Path, *, recursive: bool) -> int:
        if not directory.is_dir():
            return 0
        paths = directory.rglob("*") if recursive else directory.iterdir()
        return sum(
            path.is_file()
            and path.suffix.casefold() in MANAGED_DOCUMENT_EXTENSIONS
            and not any(part.startswith(".") for part in path.relative_to(directory).parts)
            for path in paths
        )

    def _orphans(self) -> list[dict[str, str]]:
        referenced: set[str] = set()
        if self.settings.catalogue_path.is_file():
            try:
                catalogue = CatalogueService(self.settings.catalogue_path)
                catalogue.load()
                referenced = {
                    normalized_relative_path(document.get("relative_path"))
                    for document in catalogue.documents
                    if document.get("relative_path")
                }
            except Exception:
                pass
        results = []
        for directory in (self.settings.inbox_dir, self.settings.registered_dir):
            if not directory.is_dir():
                continue
            for path in directory.iterdir():
                if (
                    path.is_file()
                    and path.suffix.casefold() in MANAGED_DOCUMENT_EXTENSIONS
                    and not path.name.startswith(".")
                ):
                    relative = path.relative_to(self.settings.library_root).as_posix()
                    if normalized_relative_path(relative) not in referenced:
                        results.append(
                            {
                                "path": relative,
                                "location": directory.name,
                            }
                        )
        return results

    def _snapshot_state(self) -> dict[str, Any]:
        marker = self.settings.state_dir / "snapshot_commit.json"
        current = None
        if marker.is_file():
            try:
                current = json.loads(marker.read_text(encoding="utf-8")).get(
                    "generation_id"
                )
            except Exception:
                current = "unreadable"
        generations_dir = self.settings.state_dir / "snapshot_generations"
        generations = (
            sorted(
                [path.name for path in generations_dir.iterdir() if path.is_dir()],
                reverse=True,
            )
            if generations_dir.is_dir()
            else []
        )
        previous = next((item for item in generations if item != current), None)
        return {
            "current_generation": current,
            "previous_generation": previous,
            "available_generations": generations,
        }

    def _recent_runtime_item(self) -> str | None:
        candidates: list[Path] = []
        for directory in (
            self.settings.state_dir / "runs",
            self.settings.reports_dir,
            self.settings.invocations_dir,
        ):
            if directory.is_dir():
                candidates.extend(path for path in directory.rglob("*") if path.is_file())
        if not candidates:
            return None
        return str(max(candidates, key=lambda path: path.stat().st_mtime_ns))

    def _write_report(self, result: WorkflowResult) -> None:
        ReportService(self.settings.reports_dir).write(result)

    def _lock_state(self) -> dict[str, Any]:
        self.settings.lock_path.parent.mkdir(parents=True, exist_ok=True)
        lock = FileLock(self.settings.lock_path, timeout=0)
        try:
            lock.acquire()
        except Timeout:
            return {
                "path": str(self.settings.lock_path),
                "held": True,
                "probe": "non_blocking_acquire",
            }
        else:
            lock.release()
            return {
                "path": str(self.settings.lock_path),
                "held": False,
                "probe": "non_blocking_acquire",
            }

    @staticmethod
    def _valid_workbook(path: Path) -> bool:
        try:
            workbook = load_workbook(path, read_only=True, data_only=False)
            valid = "Catalogue" in workbook.sheetnames
            workbook.close()
            return valid
        except Exception:
            return False

    @staticmethod
    def _configured(value: object) -> str:
        return "configured" if str(value or "").strip() else "missing"
