from __future__ import annotations

import os
import shutil
import uuid
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook

from ..exceptions import CatalogueError


LEGACY_CATALOGUE_REQUIRED_FIELDS = frozenset(
    {
        "id",
        "title",
        "topic_folder",
        "pdf_status",
        "pdf_filename",
        "pdf_relative_path",
        "uncertainty",
    }
)

DUAL_CATALOGUE_REQUIRED_FIELDS = frozenset(
    {
        "paper_uuid",
        "title",
        "authors",
        "year",
        "journal",
        "journal_abbrev",
        "doi",
        "pmid",
        "arxiv_id",
        "publication_type",
        "abstract",
        "keywords",
        "manual_tags",
        "auto_tags",
        "suggested_topic",
        "topic_folder",
        "source",
        "notes",
        "uncertainty",
        "date_added",
        "date_updated",
    }
)

DOCUMENT_REQUIRED_FIELDS = frozenset(
    {
        "document_id",
        "paper_uuid",
        "document_type",
        "supplementary_type",
        "sequence",
        "filename",
        "relative_path",
        "extension",
        "sha256",
        "file_status",
        "source",
        "uncertainty",
        "date_added",
        "date_updated",
    }
)


@dataclass(frozen=True, slots=True)
class CatalogueStatToken:
    """A lightweight optimistic-concurrency token for catalogue.xlsx."""

    size: int
    mtime_ns: int

    @classmethod
    def capture(cls, path: Path) -> "CatalogueStatToken":
        try:
            stat = path.stat()
        except OSError as exc:
            raise CatalogueError(f"Cannot stat catalogue: {path}") from exc
        return cls(size=stat.st_size, mtime_ns=stat.st_mtime_ns)


@dataclass(frozen=True, slots=True)
class CataloguePreflightResult:
    path: Path
    schema_mode: str
    token: CatalogueStatToken


class CataloguePreflightService:
    """Validate catalogue write readiness before expensive work and commit.

    The service never changes the source workbook.  Its write probe operates on
    a short-lived sibling copy so directory permissions and openpyxl round trips
    are checked without creating a user-visible workbook revision.
    """

    def __init__(self, path: str | Path):
        self.path = Path(path)

    def before_modification(self) -> CataloguePreflightResult:
        """Run the initial modifying-command preflight and return its token."""
        return self._run(expected_token=None)

    def before_commit(
        self, expected_token: CatalogueStatToken
    ) -> CataloguePreflightResult:
        """Re-run preflight and refuse a workbook changed since initial load."""
        return self._run(expected_token=expected_token)

    # Explicit aliases make the intended call sites unambiguous while keeping
    # the shorter methods convenient for tests and workflows.
    preflight_before_modification = before_modification
    preflight_before_commit = before_commit

    def _run(
        self, *, expected_token: CatalogueStatToken | None
    ) -> CataloguePreflightResult:
        self._check_exists()
        self._check_excel_lock_file()
        initial_token = CatalogueStatToken.capture(self.path)
        if expected_token is not None and initial_token != expected_token:
            raise CatalogueError(
                "Catalogue changed after preflight; refusing to overwrite concurrent edits"
            )

        self._check_direct_write_access()
        schema_mode = self._validate_schema(self.path)
        self._probe_temporary_workbook()

        final_token = CatalogueStatToken.capture(self.path)
        if final_token != initial_token:
            raise CatalogueError(
                "Catalogue changed during preflight; retry after closing other editors"
            )
        if expected_token is not None and final_token != expected_token:
            raise CatalogueError(
                "Catalogue changed after preflight; refusing to overwrite concurrent edits"
            )
        return CataloguePreflightResult(self.path, schema_mode, final_token)

    def _check_exists(self) -> None:
        if not self.path.is_file():
            raise CatalogueError(f"Required catalogue is missing: {self.path}")

    def _check_excel_lock_file(self) -> None:
        lock_path = self.path.with_name(f"~${self.path.name}")
        if lock_path.exists():
            raise CatalogueError(
                f"Catalogue appears to be open in Excel: {lock_path.name}"
            )

    def _check_direct_write_access(self) -> None:
        try:
            # Opening r+b performs no write but verifies that the current
            # process can obtain a writable handle.  Excel's sharing mode also
            # causes this to fail on Windows when no ~$ marker is available.
            with self.path.open("r+b"):
                pass
        except OSError as exc:
            raise CatalogueError(f"Catalogue is not writable: {self.path}") from exc

    def _probe_temporary_workbook(self) -> None:
        temporary = self.path.with_name(
            f".{self.path.stem}.preflight-{uuid.uuid4().hex}.tmp.xlsx"
        )
        workbook = None
        validation = None
        try:
            shutil.copy2(self.path, temporary)
            workbook = load_workbook(temporary)
            workbook.save(temporary)
            workbook.close()
            workbook = None
            validation = load_workbook(temporary, read_only=True, data_only=False)
        except Exception as exc:
            raise CatalogueError(
                "Cannot create and validate a temporary catalogue workbook"
            ) from exc
        finally:
            if validation is not None:
                validation.close()
            if workbook is not None:
                workbook.close()
            try:
                temporary.unlink(missing_ok=True)
            except OSError as exc:
                raise CatalogueError(
                    f"Cannot remove catalogue preflight file: {temporary}"
                ) from exc

    @classmethod
    def _validate_schema(cls, path: Path) -> str:
        workbook = None
        try:
            workbook = load_workbook(path, read_only=True, data_only=False)
            if "Catalogue" not in workbook.sheetnames:
                raise CatalogueError("Catalogue worksheet is missing")
            catalogue_headers = cls._headers(workbook["Catalogue"])
            has_documents = "Documents" in workbook.sheetnames

            if has_documents:
                missing_catalogue = DUAL_CATALOGUE_REQUIRED_FIELDS - catalogue_headers
                documents_headers = cls._headers(workbook["Documents"])
                missing_documents = DOCUMENT_REQUIRED_FIELDS - documents_headers
                if missing_catalogue or missing_documents:
                    parts = []
                    if missing_catalogue:
                        parts.append(
                            f"Catalogue missing {sorted(missing_catalogue)}"
                        )
                    if missing_documents:
                        parts.append(
                            f"Documents missing {sorted(missing_documents)}"
                        )
                    raise CatalogueError("Invalid dual-sheet schema: " + "; ".join(parts))
                return "dual"

            missing_legacy = LEGACY_CATALOGUE_REQUIRED_FIELDS - catalogue_headers
            if missing_legacy:
                if "paper_uuid" in catalogue_headers:
                    raise CatalogueError(
                        "Partial dual-sheet migration: Documents worksheet is missing"
                    )
                raise CatalogueError(
                    f"Invalid legacy Catalogue schema; missing {sorted(missing_legacy)}"
                )
            return "legacy"
        except CatalogueError:
            raise
        except Exception as exc:
            raise CatalogueError(f"Cannot validate catalogue schema: {path}") from exc
        finally:
            if workbook is not None:
                workbook.close()

    @staticmethod
    def _headers(worksheet: object) -> set[str]:
        headers: set[str] = set()
        for cell in worksheet[1]:  # type: ignore[index]
            if cell.value is not None and str(cell.value).strip():
                headers.add(str(cell.value).strip())
        return headers

