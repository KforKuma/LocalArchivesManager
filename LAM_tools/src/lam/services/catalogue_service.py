from __future__ import annotations

import csv
import hashlib
import os
import re
import shutil
import uuid
from copy import copy
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook

from ..exceptions import CatalogueError
from ..models import CatalogueChange, CatalogueRecord, DocumentRecord
from ..schema import (
    CATALOGUE_FIELDS,
    DOCUMENT_FIELDS,
    DOCUMENT_TYPES,
    MACHINE_FILLABLE_FIELDS,
    MACHINE_MAINTAINED_FIELDS,
    LEGACY_CATALOGUE_REQUIRED_FIELDS,
    SNAPSHOT_FIELDS,
    SYSTEM_IDENTITY_FIELDS,
    SUPPLEMENTARY_TYPES,
    USER_CONTROLLED_FIELDS,
)
from ..utils.identifiers import normalize_arxiv_id, normalize_doi, normalize_pmid
from ..utils.normalize import normalized_text
from ..utils.publication_type import canonicalize_publication_type
from ..utils.uncertainty import (
    confirmed_value,
    has_user_confirmation,
    parse_user_confirmations,
)
from .catalogue_preflight_service import (
    CataloguePreflightService,
    CatalogueStatToken,
)


UNCERTAINTY_PREFIXES = ("NEEDS_REVIEW:", "USER_CONFIRMED:", "MACHINE_NOTE:", "RESOLVED:")


class CatalogueService:
    def __init__(self, path: Path, *, allow_legacy_schema: bool = False):
        self.path = path
        self.allow_legacy_schema = allow_legacy_schema
        self.workbook = None
        self.worksheet = None
        self.headers: dict[str, int] = {}
        self.records: list[CatalogueRecord] = []
        self.documents_worksheet = None
        self.document_headers: dict[str, int] = {}
        self.documents: list[DocumentRecord] = []
        self.changes: list[CatalogueChange] = []
        self.document_changes: list[CatalogueChange] = []
        self.review_decisions: set[str] = set()
        self.maintenance_actions: list[dict[str, Any]] = []
        self.preflight_token: CatalogueStatToken | None = None

    def load(self) -> list[CatalogueRecord]:
        self.preflight_token = CatalogueStatToken.capture(self.path)
        try:
            self.workbook = load_workbook(self.path)
        except Exception as exc:
            raise CatalogueError(f"Cannot open catalogue: {self.path}") from exc

        candidates: list[tuple[Any, dict[str, int]]] = []
        for sheet in self.workbook.worksheets:
            headers = self._read_headers(sheet)
            if set(CATALOGUE_FIELDS).issubset(headers) or (
                self.allow_legacy_schema
                and LEGACY_CATALOGUE_REQUIRED_FIELDS.issubset(headers)
            ):
                candidates.append((sheet, headers))
        if not candidates:
            available = {
                sheet.title: sorted(self._read_headers(sheet))
                for sheet in self.workbook.worksheets
            }
            raise CatalogueError(
                "No worksheet contains a valid 0.5.2 Catalogue schema. "
                f"required={list(CATALOGUE_FIELDS)}; available={available}. "
                "Use 'lam migrate-identifiers --dry-run' for a legacy workbook."
            )
        self.worksheet, self.headers = candidates[0]
        self._validate_duplicate_headers()
        self.records = []
        for row_number in range(2, self.worksheet.max_row + 1):
            values = {
                name: self.worksheet.cell(row=row_number, column=column).value
                for name, column in self.headers.items()
            }
            if any(value not in (None, "") for value in values.values()):
                self.records.append(CatalogueRecord(row_number=row_number, values=values))
        self._validate_duplicate_values()
        self._load_documents()
        if not self.allow_legacy_schema and not self.has_documents_sheet:
            raise CatalogueError(
                "The 0.5.2 Catalogue requires a Documents sheet. "
                "Run 'lam migrate-identifiers --dry-run' first."
            )
        return self.records

    @property
    def has_documents_sheet(self) -> bool:
        return self.documents_worksheet is not None

    def _load_documents(self) -> None:
        assert self.workbook is not None
        self.documents_worksheet = None
        self.document_headers = {}
        self.documents = []
        if "Documents" not in self.workbook.sheetnames:
            return
        sheet = self.workbook["Documents"]
        headers = self._read_headers(sheet)
        missing = set(DOCUMENT_FIELDS) - set(headers)
        if missing and not self.allow_legacy_schema:
            raise CatalogueError(
                f"Documents sheet is missing required fields: {sorted(missing)}"
            )
        self._validate_sheet_duplicate_headers(sheet, "Documents")
        self.documents_worksheet = sheet
        self.document_headers = headers
        for row_number in range(2, sheet.max_row + 1):
            values = {
                name: sheet.cell(row=row_number, column=column).value
                for name, column in headers.items()
            }
            if any(value not in (None, "") for value in values.values()):
                self.documents.append(DocumentRecord(row_number, values))
        if not self.allow_legacy_schema:
            self._validate_documents()

    @staticmethod
    def _validate_sheet_duplicate_headers(sheet: Any, name: str) -> None:
        seen: set[str] = set()
        duplicates: set[str] = set()
        for column in range(1, sheet.max_column + 1):
            value = sheet.cell(row=1, column=column).value
            if value is None or not str(value).strip():
                continue
            header = str(value).strip()
            if header in seen:
                duplicates.add(header)
            seen.add(header)
        if duplicates:
            raise CatalogueError(
                f"Duplicate {name} columns: {sorted(duplicates)}"
            )

    def _validate_documents(self) -> None:
        seen_ids: dict[str, int] = {}
        seen_paths: dict[str, int] = {}
        seen_slots: dict[tuple[str, str, str], int] = {}
        main_by_paper: dict[str, int] = {}
        problems: list[str] = []
        catalogue_uuids: set[str] = set()
        for record in self.records:
            raw_uuid = str(record.get("paper_uuid") or "").strip()
            try:
                parsed_uuid = uuid.UUID(raw_uuid)
            except ValueError:
                problems.append(
                    f"invalid paper_uuid={raw_uuid!r} at Catalogue row {record.row_number}"
                )
                continue
            if parsed_uuid.version != 4:
                problems.append(
                    f"paper_uuid must be UUID4 at Catalogue row {record.row_number}"
                )
                continue
            catalogue_uuids.add(normalized_text(str(parsed_uuid)))
        for document in self.documents:
            document_id = normalized_text(document.get("document_id"))
            paper_uuid = normalized_text(document.get("paper_uuid"))
            document_type = normalized_text(document.get("document_type"))
            relative_path = normalized_text(document.get("relative_path"))
            if not document_id or not paper_uuid:
                problems.append(f"missing document identity at row {document.row_number}")
                continue
            if document_id in seen_ids:
                problems.append(
                    f"document_id={document.get('document_id')!r} at rows "
                    f"{seen_ids[document_id]} and {document.row_number}"
                )
            seen_ids[document_id] = document.row_number
            if paper_uuid not in catalogue_uuids:
                problems.append(
                    f"paper_uuid={document.get('paper_uuid')!r} has no Catalogue row"
                )
            if document_type not in DOCUMENT_TYPES:
                problems.append(
                    f"invalid document_type={document.get('document_type')!r} at row "
                    f"{document.row_number}"
                )
            if relative_path:
                if relative_path in seen_paths:
                    problems.append(
                        f"relative_path={document.get('relative_path')!r} at rows "
                        f"{seen_paths[relative_path]} and {document.row_number}"
                    )
                seen_paths[relative_path] = document.row_number
            if document_type == "main":
                if paper_uuid in main_by_paper:
                    problems.append(
                        f"multiple main documents for paper_uuid={document.get('paper_uuid')!r}"
                    )
                main_by_paper[paper_uuid] = document.row_number
            elif document_type == "supplementary":
                supplementary_type = str(document.get("supplementary_type") or "Supplementary")
                if supplementary_type not in SUPPLEMENTARY_TYPES:
                    problems.append(
                        f"invalid supplementary_type={supplementary_type!r} at row "
                        f"{document.row_number}"
                    )
                slot = (
                    paper_uuid,
                    normalized_text(supplementary_type),
                    normalized_text(document.get("sequence")),
                )
                if slot in seen_slots:
                    problems.append(
                        "duplicate supplementary type/sequence at rows "
                        f"{seen_slots[slot]} and {document.row_number}"
                    )
                seen_slots[slot] = document.row_number
        if problems:
            raise CatalogueError("Invalid Documents sheet: " + "; ".join(problems))

    @staticmethod
    def _read_headers(sheet: Any) -> dict[str, int]:
        headers: dict[str, int] = {}
        for column in range(1, sheet.max_column + 1):
            raw = sheet.cell(row=1, column=column).value
            if raw is None or not str(raw).strip():
                continue
            name = str(raw).strip()
            if name not in headers:
                headers[name] = column
        return headers

    def _validate_duplicate_headers(self) -> None:
        assert self.worksheet is not None
        seen: dict[str, int] = {}
        duplicates: list[str] = []
        for column in range(1, self.worksheet.max_column + 1):
            raw = self.worksheet.cell(row=1, column=column).value
            if raw is None or not str(raw).strip():
                continue
            name = str(raw).strip()
            if name in seen:
                duplicates.append(name)
            seen[name] = column
        if duplicates:
            raise CatalogueError(f"Duplicate catalogue columns: {sorted(set(duplicates))}")

    def _validate_duplicate_values(self) -> None:
        problems: list[str] = []
        for field_name in ("paper_uuid", "doi", "pmid", "arxiv_id"):
            if field_name not in self.headers:
                continue
            seen: dict[str, int] = {}
            for record in self.records:
                key = self._field_key(field_name, record.get(field_name))
                if not key:
                    continue
                if key in seen:
                    problems.append(
                        f"{field_name}={record.get(field_name)!r} at rows "
                        f"{seen[key]} and {record.row_number}"
                    )
                else:
                    seen[key] = record.row_number
        if problems:
            raise CatalogueError("Duplicate catalogue identifiers/paths: " + "; ".join(problems))

    def find_by(self, field_name: str, value: object) -> list[CatalogueRecord]:
        key = self._field_key(field_name, value)
        if not key:
            return []
        return [
            record
            for record in self.records
            if self._field_key(field_name, record.get(field_name)) == key
        ]

    @staticmethod
    def _field_key(field_name: str, value: object) -> str:
        if field_name == "doi":
            return normalize_doi(value)
        if field_name == "pmid":
            return normalize_pmid(value)
        if field_name == "arxiv_id":
            return normalize_arxiv_id(value)
        return normalized_text(value)

    def update_fields(self, record: CatalogueRecord, updates: dict[str, Any]) -> list[CatalogueChange]:
        if self.worksheet is None:
            raise CatalogueError("Catalogue must be loaded before it can be updated")
        applied: list[CatalogueChange] = []
        for field_name, new_value in updates.items():
            if field_name in USER_CONTROLLED_FIELDS:
                raise CatalogueError(f"Refusing to overwrite user-controlled field: {field_name}")
            if field_name not in MACHINE_MAINTAINED_FIELDS | MACHINE_FILLABLE_FIELDS:
                raise CatalogueError(f"Workflow cannot update field: {field_name}")
            if field_name not in self.headers:
                raise CatalogueError(f"Catalogue field is missing: {field_name}")
            if field_name == "publication_type":
                new_value = canonicalize_publication_type(new_value).canonical_type
            old_value = record.get(field_name, None)
            if (
                field_name in MACHINE_FILLABLE_FIELDS
                and old_value not in (None, "")
                and not self._equivalent(old_value, new_value)
            ):
                raise CatalogueError(
                    f"Refusing to overwrite non-empty bibliographic field: {field_name}"
                )
            if self._equivalent(old_value, new_value):
                continue
            column = self.headers[field_name]
            self.worksheet.cell(row=record.row_number, column=column).value = new_value
            record.values[field_name] = new_value
            change = CatalogueChange(record.row_number, field_name, old_value, new_value)
            self.changes.append(change)
            applied.append(change)
        return applied

    def normalize_topic_folder_for_migration(
        self, record: CatalogueRecord, normalized_value: str
    ) -> CatalogueChange | None:
        """Remove only a historical Topics/ prefix during explicit migration."""
        if self.worksheet is None:
            raise CatalogueError("Catalogue must be loaded before it can be updated")
        if "topic_folder" not in self.headers:
            raise CatalogueError("Catalogue field is missing: topic_folder")
        old_value = str(record.get("topic_folder") or "").strip().replace("\\", "/")
        if not old_value.casefold().startswith("topics/"):
            if self._equivalent(old_value, normalized_value):
                return None
            raise CatalogueError(
                "Topic migration may only normalize an existing Topics/ prefix"
            )
        if old_value.split("/", 1)[1] != normalized_value:
            raise CatalogueError("Topic migration normalization changed topic semantics")
        column = self.headers["topic_folder"]
        self.worksheet.cell(row=record.row_number, column=column).value = normalized_value
        record.values["topic_folder"] = normalized_value
        change = CatalogueChange(
            record.row_number, "topic_folder", old_value, normalized_value
        )
        self.changes.append(change)
        return change

    def ensure_header(self, field_name: str) -> int:
        """Add one optional machine column in memory while preserving all sheets."""
        if self.worksheet is None:
            raise CatalogueError("Catalogue must be loaded before its schema can be extended")
        if field_name in self.headers:
            return self.headers[field_name]
        column = self.worksheet.max_column + 1
        target = self.worksheet.cell(row=1, column=column)
        target.value = field_name
        if column > 1:
            source = self.worksheet.cell(row=1, column=column - 1)
            target.font = copy(source.font)
            target.fill = copy(source.fill)
            target.border = copy(source.border)
            target.alignment = copy(source.alignment)
            target.number_format = source.number_format
            target.protection = copy(source.protection)
        self.headers[field_name] = column
        for record in self.records:
            record.values.setdefault(field_name, None)
        return column

    def ensure_documents_sheet(self) -> Any:
        """Create the Documents sheet without disturbing other sheets."""
        if self.workbook is None or self.worksheet is None:
            raise CatalogueError("Catalogue must be loaded before Documents can be created")
        if self.documents_worksheet is not None:
            return self.documents_worksheet
        sheet = self.workbook.create_sheet("Documents")
        for column, field_name in enumerate(DOCUMENT_FIELDS, start=1):
            target = sheet.cell(row=1, column=column, value=field_name)
            source = self.worksheet.cell(row=1, column=min(column, self.worksheet.max_column))
            target.font = copy(source.font)
            target.fill = copy(source.fill)
            target.border = copy(source.border)
            target.alignment = copy(source.alignment)
            target.number_format = source.number_format
            target.protection = copy(source.protection)
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = f"A1:{sheet.cell(row=1, column=len(DOCUMENT_FIELDS)).coordinate}1"
        self.documents_worksheet = sheet
        self.document_headers = {
            field_name: index for index, field_name in enumerate(DOCUMENT_FIELDS, start=1)
        }
        self.documents = []
        return sheet

    def ensure_paper_uuid(self, record: CatalogueRecord) -> str:
        current = str(record.get("paper_uuid") or "").strip()
        if current:
            try:
                parsed = uuid.UUID(current)
            except ValueError as exc:
                raise CatalogueError(
                    f"Invalid paper_uuid at row {record.row_number}: {current!r}"
                ) from exc
            if parsed.version != 4:
                raise CatalogueError(
                    f"paper_uuid must be UUID4 at row {record.row_number}: {current!r}"
                )
            return str(parsed)
        column = self.ensure_header("paper_uuid")
        value = str(uuid.uuid4())
        self.worksheet.cell(row=record.row_number, column=column).value = value
        record.values["paper_uuid"] = value
        self.changes.append(CatalogueChange(record.row_number, "paper_uuid", None, value))
        return value

    def find_documents_by(self, field_name: str, value: object) -> list[DocumentRecord]:
        key = normalized_text(value)
        if not key:
            return []
        return [
            document
            for document in self.documents
            if normalized_text(document.get(field_name)) == key
        ]

    def documents_for_paper(self, paper_uuid: object) -> list[DocumentRecord]:
        return self.find_documents_by("paper_uuid", paper_uuid)

    def add_document(self, values: dict[str, Any]) -> DocumentRecord:
        if self.documents_worksheet is None:
            self.ensure_documents_sheet()
        assert self.documents_worksheet is not None
        supplied = {
            key: value
            for key, value in values.items()
            if key in self.document_headers and value not in (None, "")
        }
        required = {"document_id", "paper_uuid", "document_type"}
        missing = required - set(supplied)
        if missing:
            raise CatalogueError(f"Document is missing required fields: {sorted(missing)}")
        try:
            paper_uuid = str(uuid.UUID(str(supplied["paper_uuid"])))
        except ValueError as exc:
            raise CatalogueError(f"Invalid document paper_uuid: {supplied['paper_uuid']!r}") from exc
        supplied["paper_uuid"] = paper_uuid
        if not self.find_by("paper_uuid", paper_uuid):
            raise CatalogueError(f"Document paper_uuid has no Catalogue row: {paper_uuid}")
        document_type = normalized_text(supplied["document_type"])
        if document_type not in DOCUMENT_TYPES:
            raise CatalogueError(f"Invalid document_type: {supplied['document_type']!r}")
        supplied["document_type"] = document_type
        if self.find_documents_by("document_id", supplied["document_id"]):
            raise CatalogueError(
                f"Refusing duplicate document_id: {supplied['document_id']!r}"
            )
        relative_path = supplied.get("relative_path")
        if relative_path and self.find_documents_by("relative_path", relative_path):
            raise CatalogueError(f"Refusing duplicate document path: {relative_path!r}")
        if document_type == "main" and any(
            normalized_text(item.get("document_type")) == "main"
            for item in self.documents_for_paper(paper_uuid)
        ):
            raise CatalogueError(f"Paper already has a main document: {paper_uuid}")
        if document_type == "supplementary":
            supplementary_type = str(supplied.get("supplementary_type") or "Supplementary")
            if supplementary_type not in SUPPLEMENTARY_TYPES:
                raise CatalogueError(f"Invalid supplementary_type: {supplementary_type!r}")
            supplied["supplementary_type"] = supplementary_type
            sequence = normalized_text(supplied.get("sequence"))
            for item in self.documents_for_paper(paper_uuid):
                if normalized_text(item.get("document_type")) != "supplementary":
                    continue
                if (
                    normalized_text(item.get("supplementary_type"))
                    == normalized_text(supplementary_type)
                    and normalized_text(item.get("sequence")) == sequence
                ):
                    raise CatalogueError(
                        "Refusing duplicate supplementary type/sequence: "
                        f"{supplementary_type} {sequence or '(unsequenced)'}"
                    )
        row_number = self.documents_worksheet.max_row + 1
        for field_name, value in supplied.items():
            self.documents_worksheet.cell(
                row=row_number, column=self.document_headers[field_name]
            ).value = value
        record_values = {
            name: self.documents_worksheet.cell(row=row_number, column=column).value
            for name, column in self.document_headers.items()
        }
        record = DocumentRecord(row_number, record_values)
        self.documents.append(record)
        self.document_changes.append(
            CatalogueChange(row_number, "__row__", None, supplied)
        )
        return record

    def update_document_fields(
        self, document: DocumentRecord, updates: dict[str, Any]
    ) -> list[CatalogueChange]:
        if self.documents_worksheet is None:
            raise CatalogueError("Documents sheet is not available")
        immutable = {"document_id", "paper_uuid", "document_type"}
        if immutable & set(updates):
            raise CatalogueError(
                f"Refusing to change immutable document fields: {sorted(immutable & set(updates))}"
            )
        applied: list[CatalogueChange] = []
        for field_name, new_value in updates.items():
            if field_name not in self.document_headers:
                raise CatalogueError(f"Documents field is missing: {field_name}")
            old_value = document.get(field_name, None)
            if self._equivalent(old_value, new_value):
                continue
            self.documents_worksheet.cell(
                row=document.row_number,
                column=self.document_headers[field_name],
            ).value = new_value
            document.values[field_name] = new_value
            change = CatalogueChange(document.row_number, field_name, old_value, new_value)
            self.document_changes.append(change)
            applied.append(change)
        return applied

    def update_canonical_fields(
        self, record: CatalogueRecord, updates: dict[str, Any]
    ) -> list[CatalogueChange]:
        """Apply a prevalidated canonical record without overwriting user decisions."""
        if self.worksheet is None:
            raise CatalogueError("Catalogue must be loaded before it can be updated")
        confirmed_fields = {
            item.field.strip().casefold()
            for item in parse_user_confirmations(record.get("uncertainty"))
        }
        applied: list[CatalogueChange] = []
        for field_name, new_value in updates.items():
            if field_name in USER_CONTROLLED_FIELDS:
                raise CatalogueError(f"Refusing to overwrite user-controlled field: {field_name}")
            if field_name not in (
                SYSTEM_IDENTITY_FIELDS | MACHINE_MAINTAINED_FIELDS | MACHINE_FILLABLE_FIELDS
            ):
                raise CatalogueError(f"Workflow cannot canonicalize field: {field_name}")
            if field_name == "paper_uuid":
                current_uid = str(record.get("paper_uuid") or "").strip()
                if current_uid and normalized_text(current_uid) != normalized_text(new_value):
                    raise CatalogueError("Refusing to change immutable paper_uuid")
            if field_name not in self.headers:
                raise CatalogueError(f"Catalogue field is missing: {field_name}")
            if field_name.casefold() in confirmed_fields and field_name != "paper_uuid":
                continue
            if field_name == "publication_type":
                new_value = canonicalize_publication_type(new_value).canonical_type
            old_value = record.get(field_name, None)
            if self._equivalent(old_value, new_value):
                continue
            column = self.headers[field_name]
            self.worksheet.cell(row=record.row_number, column=column).value = new_value
            record.values[field_name] = new_value
            change = CatalogueChange(record.row_number, field_name, old_value, new_value)
            self.changes.append(change)
            applied.append(change)
        return applied

    def update_provisional_fields(
        self, record: CatalogueRecord, updates: dict[str, Any]
    ) -> list[CatalogueChange]:
        """Upgrade machine-owned values on one provisional row after confirmation."""
        if self.worksheet is None:
            raise CatalogueError("Catalogue must be loaded before it can be updated")
        current_uncertainty = str(record.get("uncertainty") or "")
        confirmed_fields = {
            item.field.strip().casefold()
            for item in parse_user_confirmations(current_uncertainty)
        }
        applied: list[CatalogueChange] = []
        for field_name, new_value in updates.items():
            if field_name in USER_CONTROLLED_FIELDS:
                raise CatalogueError(f"Refusing to overwrite user-controlled field: {field_name}")
            if field_name not in MACHINE_MAINTAINED_FIELDS | MACHINE_FILLABLE_FIELDS:
                raise CatalogueError(f"Workflow cannot update field: {field_name}")
            if field_name not in self.headers:
                raise CatalogueError(f"Catalogue field is missing: {field_name}")
            if field_name == "publication_type":
                new_value = canonicalize_publication_type(new_value).canonical_type
            if field_name.casefold() in confirmed_fields:
                continue
            old_value = record.get(field_name, None)
            if self._equivalent(old_value, new_value):
                continue
            column = self.headers[field_name]
            self.worksheet.cell(row=record.row_number, column=column).value = new_value
            record.values[field_name] = new_value
            change = CatalogueChange(record.row_number, field_name, old_value, new_value)
            self.changes.append(change)
            applied.append(change)
        return applied

    def add_record(self, values: dict[str, Any]) -> CatalogueRecord:
        """Append one machine-created row without changing existing row order."""
        if self.worksheet is None:
            raise CatalogueError("Catalogue must be loaded before a row can be added")
        normalized_values = dict(values)
        normalized_values.setdefault("paper_uuid", str(uuid.uuid4()))
        if "publication_type" in normalized_values:
            normalized_values["publication_type"] = canonicalize_publication_type(
                normalized_values["publication_type"]
            ).canonical_type
        supplied = {
            key: value for key, value in normalized_values.items() if key in self.headers and value not in (None, "")
        }
        unsupported = set(supplied) - (
            SYSTEM_IDENTITY_FIELDS | MACHINE_FILLABLE_FIELDS | MACHINE_MAINTAINED_FIELDS
        )
        if unsupported:
            raise CatalogueError(
                f"Refusing unsupported fields in new catalogue row: {sorted(unsupported)}"
            )
        if any(supplied.get(field) for field in USER_CONTROLLED_FIELDS):
            raise CatalogueError("Machine-created rows cannot set user-controlled fields")
        for field_name in ("paper_uuid", "doi", "pmid", "arxiv_id"):
            value = supplied.get(field_name)
            if value and self.find_by(field_name, value):
                raise CatalogueError(
                    f"Refusing duplicate catalogue record: {field_name}={value!r}"
                )
        row_number = self.worksheet.max_row + 1
        for field_name, value in supplied.items():
            self.worksheet.cell(row=row_number, column=self.headers[field_name]).value = value
        record_values = {
            name: self.worksheet.cell(row=row_number, column=column).value
            for name, column in self.headers.items()
        }
        record = CatalogueRecord(row_number=row_number, values=record_values)
        self.records.append(record)
        self.changes.append(CatalogueChange(row_number, "__row__", None, supplied))
        return record

    def repair_publication_type(
        self, record: CatalogueRecord, raw_value: Any
    ) -> list[CatalogueChange]:
        """Normalize the machine-fillable type during the explicit repair workflow."""
        if self.worksheet is None:
            raise CatalogueError("Catalogue must be loaded before it can be updated")
        if "publication_type" not in self.headers:
            raise CatalogueError("Catalogue field is missing: publication_type")
        new_value = canonicalize_publication_type(raw_value).canonical_type
        old_value = record.get("publication_type", None)
        if self._equivalent(old_value, new_value):
            return []
        column = self.headers["publication_type"]
        self.worksheet.cell(row=record.row_number, column=column).value = new_value
        record.values["publication_type"] = new_value
        change = CatalogueChange(record.row_number, "publication_type", old_value, new_value)
        self.changes.append(change)
        return [change]

    @staticmethod
    def _equivalent(left: Any, right: Any) -> bool:
        if left in (None, "") and right in (None, ""):
            return True
        return left == right

    def add_uncertainty(
        self,
        record: CatalogueRecord,
        prefix: str,
        field_name: str,
        issue: str,
        *,
        conflict_with_confirmation: bool = False,
        issue_key: str | None = None,
    ) -> bool:
        if prefix not in UNCERTAINTY_PREFIXES:
            raise CatalogueError(f"Unsupported uncertainty prefix: {prefix}")
        if prefix == "NEEDS_REVIEW:":
            outcome = self.ensure_review_blocker(
                record,
                field_name,
                issue,
                issue_key=issue_key or "review",
                conflict_with_confirmation=conflict_with_confirmation,
            )
            return outcome == "added"
        current = str(record.get("uncertainty") or "")
        lines = [line.rstrip() for line in current.splitlines() if line.strip()]
        issue_part = f"; issue_key={issue_key}" if issue_key else ""
        line = f"{prefix} field={field_name}{issue_part}; issue={issue}"
        normalized_line = normalized_text(line)
        if any(normalized_text(existing) == normalized_line for existing in lines):
            return False
        lines.append(line)
        self.update_fields(record, {"uncertainty": "\n".join(lines)})
        return True

    def ensure_review_blocker(
        self,
        record: CatalogueRecord,
        field_name: str,
        issue: str,
        *,
        issue_key: str,
        conflict_with_confirmation: bool = False,
    ) -> str:
        """Return added, existing, confirmed, or cleared for one row/field blocker."""
        current = str(record.get("uncertainty") or "")
        lines = [line.rstrip() for line in current.splitlines() if line.strip()]
        if not conflict_with_confirmation and self._has_user_confirmation(lines, field_name):
            retained = [
                line for line in lines if not self._is_review_for_field(line, field_name)
            ]
            if retained != lines:
                self.update_fields(record, {"uncertainty": "\n".join(retained)})
            return "confirmed"

        decision_key = self._review_decision_key(record, field_name, issue_key, issue)
        if decision_key in self.review_decisions:
            return "cleared"
        if any(self._is_review_for_field(line, field_name) for line in lines):
            return "existing"

        line = (
            f"NEEDS_REVIEW: field={field_name}; issue_key={issue_key}; issue={issue}"
        )
        lines.append(line)
        self.update_fields(record, {"uncertainty": "\n".join(lines)})
        return "added"

    def configure_review_state(self, previous_snapshot: dict[str, Any]) -> None:
        """Carry approvals and treat a user-cleared blocker as a one-time decision."""
        self.review_decisions.update(previous_snapshot.get("review_decisions", []))
        old_rows = {row.get("row_number"): row for row in previous_snapshot.get("rows", [])}
        old_by_uid = {
            normalized_text(
                row.get("paper_uuid") or row.get("fields", {}).get("paper_uuid")
            ): row
            for row in previous_snapshot.get("rows", [])
            if normalized_text(
                row.get("paper_uuid") or row.get("fields", {}).get("paper_uuid")
            )
        }
        for record in self.records:
            old_row = old_by_uid.get(
                normalized_text(record.get("paper_uuid"))
            ) or old_rows.get(
                record.row_number, {}
            )
            old_uncertainty = str(
                old_row.get("fields", {}).get("uncertainty") or ""
            )
            current_lines = {
                line.strip()
                for line in str(record.get("uncertainty") or "").splitlines()
                if line.strip()
            }
            for old_line in old_uncertainty.splitlines():
                old_line = old_line.strip()
                parsed = self._parse_review_line(old_line)
                if not parsed or old_line in current_lines:
                    continue
                field_name, issue_key, issue = parsed
                if any(self._is_review_for_field(line, field_name) for line in current_lines):
                    continue
                self.review_decisions.add(
                    self._review_decision_key(
                        record, field_name, issue_key, issue
                    )
                )

    def has_user_confirmation(self, record: CatalogueRecord, field_name: str) -> bool:
        return has_user_confirmation(record.get("uncertainty"), field_name)

    def confirmed_value(self, record: CatalogueRecord, field_name: str) -> str:
        return confirmed_value(record.get("uncertainty"), field_name)

    def resolve_confirmed_reviews(self, record: CatalogueRecord) -> list[str]:
        """Remove only machine blockers covered by a user confirmation."""
        current = str(record.get("uncertainty") or "")
        lines = [line.rstrip() for line in current.splitlines() if line.strip()]
        removed: list[str] = []
        retained: list[str] = []
        for line in lines:
            parsed = self._parse_review_line(line)
            if parsed and self.has_user_confirmation(record, parsed[0]):
                removed.append(parsed[1])
                continue
            retained.append(line)
        if retained != lines:
            self.update_fields(record, {"uncertainty": "\n".join(retained)})
        return removed

    def resolve_review_blockers(
        self,
        record: CatalogueRecord,
        field_name: str,
        issue_keys: set[str],
        *,
        resolution: str,
    ) -> list[str]:
        """Resolve only named machine blockers while preserving all user text."""
        current = str(record.get("uncertainty") or "")
        lines = [line.rstrip() for line in current.splitlines() if line.strip()]
        removed: list[str] = []
        retained: list[str] = []
        for line in lines:
            parsed = self._parse_review_line(line)
            if parsed and normalized_text(parsed[0]) == normalized_text(field_name) and parsed[1] in issue_keys:
                removed.append(parsed[1])
                continue
            retained.append(line)
        if not removed:
            return []
        resolved_line = (
            f"RESOLVED: field={field_name}; issue_key={','.join(sorted(set(removed)))}; "
            f"issue={resolution}"
        )
        if not any(normalized_text(line) == normalized_text(resolved_line) for line in retained):
            retained.append(resolved_line)
        self.update_fields(record, {"uncertainty": "\n".join(retained)})
        return removed

    def active_review_lines(self, record: CatalogueRecord) -> list[str]:
        active: list[str] = []
        for line in str(record.get("uncertainty") or "").splitlines():
            parsed = self._parse_review_line(line.strip())
            if not parsed:
                continue
            field_name, issue_key, issue = parsed
            if self.has_user_confirmation(record, field_name):
                continue
            if self._review_decision_key(record, field_name, issue_key, issue) in self.review_decisions:
                continue
            active.append(line.rstrip())
        return active

    @staticmethod
    def _parse_review_line(line: str) -> tuple[str, str, str] | None:
        if not line.lstrip().upper().startswith("NEEDS_REVIEW:"):
            return None
        field_match = re.search(r"(?:NEEDS_REVIEW:|;)\s*field=([^;]+)", line, re.I)
        issue_key_match = re.search(r"(?:^|;)\s*issue_key=([^;]+)", line, re.I)
        issue_match = re.search(r"(?:^|;)\s*issue=(.*)$", line, re.I)
        if not field_match or not issue_match:
            return None
        return (
            field_match.group(1).strip(),
            issue_key_match.group(1).strip() if issue_key_match else "review",
            issue_match.group(1).strip(),
        )

    @classmethod
    def _is_review_for_field(cls, line: str, field_name: str) -> bool:
        parsed = cls._parse_review_line(line)
        return bool(parsed and normalized_text(parsed[0]) == normalized_text(field_name))

    @staticmethod
    def _review_decision_key(
        record: CatalogueRecord,
        field_name: str,
        issue_key: str,
        issue: str,
    ) -> str:
        identity = (
            record.get("paper_uuid")
            or f"row:{record.row_number}"
        )
        evidence = "|".join(
            normalized_text(record.get(field))
            for field in ("pmid", "doi", "title", "authors", "year", "journal")
        )
        payload = "|".join(
            normalized_text(value)
            for value in (identity, field_name, issue_key, issue, evidence)
        )
        return hashlib.sha256(payload.encode("utf-8")).hexdigest()

    @staticmethod
    def _has_user_confirmation(lines: Iterable[str], field_name: str) -> bool:
        return has_user_confirmation("\n".join(lines), field_name)

    def snapshot_payload(self) -> dict[str, Any]:
        rows = []
        for record in self.records:
            rows.append(
                {
                    "row_number": record.row_number,
                    "paper_uuid": record.get("paper_uuid", None),
                    "fields": {field: record.get(field, None) for field in SNAPSHOT_FIELDS},
                }
            )
        documents = [
            {
                "row_number": document.row_number,
                "document_id": document.get("document_id", None),
                "paper_uuid": document.get("paper_uuid", None),
                "fields": {
                    field: document.get(field, None) for field in DOCUMENT_FIELDS
                },
            }
            for document in self.documents
        ]
        return {
            "sheet": self.worksheet.title if self.worksheet else None,
            "rows": rows,
            "documents_sheet": (
                self.documents_worksheet.title if self.documents_worksheet else None
            ),
            "documents": documents,
            "review_decisions": sorted(self.review_decisions),
        }

    def save_atomic(self) -> Path | None:
        if not self.changes and not self.document_changes:
            return None
        if self.workbook is None:
            raise CatalogueError("Catalogue must be loaded before it can be saved")
        if self.preflight_token is not None:
            CataloguePreflightService(self.path).before_commit(self.preflight_token)
        timestamp = datetime.now().astimezone().strftime("%Y%m%d-%H%M%S")
        backup = self._unique_backup_path(timestamp)
        temporary = self.path.with_name(f".{self.path.stem}.{timestamp}.tmp.xlsx")
        replaced = False
        try:
            shutil.copy2(self.path, backup)
            self.workbook.save(temporary)
            check = load_workbook(temporary, read_only=True)
            check.close()
            os.replace(temporary, self.path)
            replaced = True
            verification = CatalogueService(self.path)
            verification.load()
            if verification.workbook is not None:
                verification.workbook.close()
            # Retention is post-commit maintenance. A cleanup failure must be
            # reported, but must never roll back an already validated save.
            try:
                self._prune_valid_backups(keep=5)
            except Exception as exc:
                self.maintenance_actions.append(
                    {
                        "action": "catalogue_backup_cleanup_failed",
                        "path": str(self.path.parent),
                        "error": str(exc),
                    }
                )
            return backup
        except Exception as exc:
            if temporary.exists():
                temporary.unlink(missing_ok=True)
            if replaced and backup.is_file():
                try:
                    shutil.copy2(backup, self.path)
                except OSError:
                    pass
            pending = self.path.with_name("catalogue_pending_updates.csv")
            self._write_pending_updates(pending)
            if self.document_changes:
                self._write_pending_updates(
                    self.path.with_name("documents_pending_updates.csv"),
                    changes=self.document_changes,
                )
            raise CatalogueError(
                f"Catalogue write failed; proposed changes exported to {pending}"
            ) from exc

    def _unique_backup_path(self, timestamp: str) -> Path:
        candidate = self.path.with_name(f"catalogue.backup.{timestamp}.xlsx")
        counter = 1
        while candidate.exists():
            candidate = self.path.with_name(
                f"catalogue.backup.{timestamp}-{counter:02d}.xlsx"
            )
            counter += 1
        return candidate

    def _write_pending_updates(
        self, path: Path, *, changes: Iterable[CatalogueChange] | None = None
    ) -> None:
        with path.open("w", newline="", encoding="utf-8-sig") as handle:
            writer = csv.DictWriter(
                handle,
                fieldnames=["row_number", "field_name", "old_value", "new_value"],
            )
            writer.writeheader()
            for change in changes if changes is not None else self.changes:
                writer.writerow(
                    {
                        "row_number": change.row_number,
                        "field_name": change.field_name,
                        "old_value": change.old_value,
                        "new_value": change.new_value,
                    }
                )

    def _prune_valid_backups(self, *, keep: int) -> None:
        pattern = re.compile(
            r"^catalogue\.backup\.\d{8}-\d{6}(?:-\d{2})?\.xlsx$",
            re.IGNORECASE,
        )
        protected = self._journal_protected_backups()
        valid: list[Path] = []
        for candidate in self.path.parent.glob("catalogue.backup.*.xlsx"):
            if not pattern.fullmatch(candidate.name):
                continue
            try:
                workbook = load_workbook(candidate, read_only=True)
                workbook.close()
            except Exception:
                continue
            valid.append(candidate)
        valid.sort(key=lambda item: (item.stat().st_mtime_ns, item.name), reverse=True)
        retained = set(valid[: max(0, keep)]) | protected
        for candidate in valid:
            if candidate in retained:
                continue
            try:
                size = candidate.stat().st_size
                candidate.unlink()
                self.maintenance_actions.append(
                    {
                        "action": "deleted_catalogue_backup",
                        "path": str(candidate),
                        "bytes": size,
                        "reason": f"valid_backup_beyond_recent_{keep}",
                    }
                )
            except OSError as exc:
                self.maintenance_actions.append(
                    {
                        "action": "catalogue_backup_cleanup_failed",
                        "path": str(candidate),
                        "error": str(exc),
                    }
                )

    def _journal_protected_backups(self) -> set[Path]:
        protected: set[Path] = set()
        runs_dir = self.path.parent / ".library_state" / "runs"
        if not runs_dir.is_dir():
            return protected
        for journal_path in runs_dir.glob("*/operation_journal.json"):
            try:
                import json

                payload = json.loads(journal_path.read_text(encoding="utf-8"))
            except Exception:
                continue
            if payload.get("status") == "final_check_committed":
                continue
            values = [payload]
            while values:
                value = values.pop()
                if isinstance(value, dict):
                    values.extend(value.values())
                elif isinstance(value, list):
                    values.extend(value)
                elif isinstance(value, str) and "catalogue.backup." in value:
                    candidate = Path(value)
                    if not candidate.is_absolute():
                        candidate = self.path.parent / candidate.name
                    try:
                        candidate.resolve().relative_to(self.path.parent.resolve())
                    except ValueError:
                        continue
                    protected.add(candidate)
        return protected
